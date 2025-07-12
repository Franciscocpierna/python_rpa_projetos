# Importa o módulo tkinter como 'tk' para criar interfaces 
        # gráficas baseadas em janelas.
import tkinter as tk

# Importa submódulos específicos do tkinter: filedialog para 
        # dialogo de seleção de arquivos, ttk para widgets 
        # temáticos e messagebox para caixas de mensagem.
from tkinter import filedialog, ttk, messagebox

# Importa o módulo pandas como 'pd', uma biblioteca poderosa 
        # para manipulação e análise de dados estruturados.
import pandas as pd

# Importa o módulo openpyxl para leitura e escrita de arquivos 
        # Excel (.xlsx), permitindo manipulação detalhada das planilhas.
import openpyxl

# Importa a classe PatternFill do módulo openpyxl.styles, 
        # usada para aplicar preenchimento de padrão às células do Excel.
from openpyxl.styles import PatternFill

# Importa o módulo os, que fornece uma maneira portátil de usar 
        # funcionalidades dependentes do sistema operacional, 
        # como ler ou escrever arquivos.
import os


# Define uma classe chamada Aplicativo. 
        # Em Python, uma classe é como um molde que permite criar 
        # objetos com atributos (dados) e métodos (funções) específicos.
class Aplicativo:
    
    # '__init__' é um método especial chamado de construtor. Ele é 
            # automaticamente chamado quando um novo objeto da 
            # classe Aplicativo é criado.
    def __init__(self, raiz):
        
        # 'self.raiz' é um atributo do objeto criado. 'raiz' é um 
                # parâmetro passado durante a criação do objeto. Esse 
                # parâmetro é geralmente uma janela principal da interface gráfica.
        self.raiz = raiz

        # 'Define o título da janela principal para 
                # "Formatação Condicional em Tabela".
        self.raiz.title("Formatação Condicional em Tabela")

        # Define o tamanho da janela principal como 950 pixels de 
                # largura e 650 pixels de altura.
        self.raiz.geometry("950x650")
        
        # 'Inicializa o atributo 'dados_excel' como None. Esse 
                # atributo será usado para armazenar os 
                # dados do arquivo Excel que será carregado.
        self.dados_excel = None

        # Inicializa o atributo 'caminho_arquivo' como None. Este 
                # atributo será usado para armazenar o caminho do 
                # arquivo Excel selecionado pelo usuário.
        self.caminho_arquivo = None

        # Cria um contêiner na janela principal usando um Frame, 
                # que é um retângulo que pode conter outros widgets. 
        # 'bg="#f7f7f7"' define a cor de fundo desse contêiner.
        self.container_principal = tk.Frame(raiz, bg="#f7f7f7")

        # Coloca o contêiner na janela e permite que ele expanda em ambos 
                # os eixos (horizontal e vertical). 
        # 'padx=20' e 'pady=20' adicionam espaçamento interno de 20 
                # pixels em todas as direções.
        self.container_principal.pack(fill=tk.BOTH, 
                                      expand=True, 
                                      padx=20, 
                                      pady=20)

        # Cria um rótulo com o texto "Formatação Condicional em Tabela". 
        # A fonte é Arial, tamanho 20 e negrito. A cor de fundo é a mesma do contêiner.
        self.titulo = tk.Label(self.container_principal, 
                               text="Formatação Condicional em Tabela", 
                               font=('Arial', 20, 'bold'), bg="#f7f7f7")

        # Posiciona o rótulo no grid (grade) do contêiner. 
        # 'row=0' significa que ele está na primeira linha. 'columnspan=4' 
                # faz com que ele se estenda por quatro colunas. 
        # 'pady=10' adiciona espaçamento vertical de 10 pixels.
        self.titulo.grid(row=0, columnspan=4, pady=10)


        # Cria um novo contêiner dentro do contêiner principal. 
        # Este contêiner específico será usado para organizar elementos 
                # relacionados ao carregamento de arquivos, como botões e 
                # rótulos. 'bg="#f7f7f7"' define a cor de fundo deste contêiner.
        self.container_arquivo = tk.Frame(self.container_principal, 
                                          bg="#f7f7f7")
        
        # Posiciona o contêiner de arquivo na janela principal. 
        # 'row=1' coloca o contêiner na segunda linha (a contagem começa do 0). 
        # 'columnspan=4' faz com que o contêiner se estenda por quatro 
                # colunas. 'pady=10' adiciona um espaçamento vertical 
                # de 10 pixels acima e abaixo do contêiner.
        self.container_arquivo.grid(row=1, columnspan=4, pady=10)
        
        # Cria um botão e o coloca dentro do contêiner de arquivo. 
        # 'text="Escolher Arquivo"' é o texto exibido no botão. 
        # 'command=self.carregar_arquivo' define a função que será 
                # chamada quando o botão for clicado. 
        # 'bg="#4CAF50"' e 'fg="white"' definem a cor de fundo do botão e 
                # a cor do texto, respectivamente. 
        # 'font=('Arial', 12, 'bold')' configura o estilo da fonte.
        self.botao_carregar_arquivo = tk.Button(self.container_arquivo, 
                                                text="Escolher Arquivo", 
                                                command=self.carregar_arquivo, 
                                                bg="#4CAF50", 
                                                fg="white", 
                                                font=('Arial', 12, 'bold'))
        
        # Organiza o botão dentro do contêiner usando o método 'pack'. 
        # 'side=tk.LEFT' coloca o botão à esquerda dentro do contêiner. 
        # 'padx=5' adiciona um espaçamento horizontal de 5 pixels à 
                # esquerda e à direita do botão.
        self.botao_carregar_arquivo.pack(side=tk.LEFT, padx=5)
        
        # Cria um rótulo dentro do contêiner de arquivo. Este rótulo será 
                # usado para mostrar o nome do arquivo selecionado 
                # ou uma mensagem padrão. 
        # 'text="Nenhum arquivo escolhido"' é o texto inicial mostrado 
                # no rótulo. 
        # 'bg="#f7f7f7"' define a cor de fundo igual ao do contêiner 
                # para manter a consistência visual. 
        # 'font=('Arial', 12)' define o estilo da fonte.
        self.rotulo_nome_arquivo = tk.Label(self.container_arquivo, 
                                            text="Nenhum arquivo escolhido", 
                                            bg="#f7f7f7", font=('Arial', 12))
        
        # Organiza o rótulo ao lado do botão dentro do contêiner usando o método 'pack'. 
        # 'side=tk.LEFT' continua alinhando elementos à esquerda. 
        # 'padx=5' adiciona um espaçamento horizontal, mantendo um 
                # pequeno espaço entre o botão e o rótulo.
        self.rotulo_nome_arquivo.pack(side=tk.LEFT, padx=5)


        # Cria um novo contêiner chamado 'menu_selecao' dentro do contêiner principal. 
                # Este contêiner será usado para organizar os elementos relacionados à 
                # seleção de colunas e condições para a formatação condicional. 
        # 'bg="#f7f7f7"' configura a cor de fundo para ser a mesma 
                # do contêiner principal.
        self.menu_selecao = tk.Frame(self.container_principal, bg="#f7f7f7")
        
        # Posiciona o contêiner 'menu_selecao' dentro do layout da janela principal. 
        # 'row=2' coloca o contêiner na terceira linha (a contagem 
                # começa do zero). 'columnspan=4' faz com que o contêiner 
                # se estenda por quatro colunas, ocupando o espaço horizontalmente. 
        # 'pady=10' adiciona um espaçamento vertical de 10 pixels acima e 
                # abaixo do contêiner para separá-lo dos outros elementos visualmente.
        self.menu_selecao.grid(row=2, columnspan=4, pady=10)
        
        # Cria um rótulo dentro do contêiner 'menu_selecao'. Este rótulo serve 
                # para instruir o usuário a escolher uma coluna do arquivo Excel
                # para a formatação condicional. 
        # 'text="Selecione a Coluna:"' é o texto que será exibido.
        # 'bg="#f7f7f7"' define a cor de fundo igual ao contêiner para consistência, e 
                # 'font=('Arial', 12)' estabelece o estilo da fonte.
        self.rotulo_selecao_coluna = tk.Label(self.menu_selecao, 
                                              text="Selecione a Coluna:", 
                                              bg="#f7f7f7", font=('Arial', 12))
        
        # Posiciona o rótulo dentro do grid do contêiner 'menu_selecao'. 
        # 'row=0' e 'column=0' colocam o rótulo na primeira linha e primeira 
                # coluna, respectivamente. 
        # 'padx=5' adiciona um espaçamento horizontal de 5 pixels para 
                # não tocar nas bordas do contêiner. 
        # 'sticky="e"' faz com que o rótulo alinhe-se ao lado direito 
                # da célula do grid para uma aparência mais organizada.
        self.rotulo_selecao_coluna.grid(row=0, 
                                        column=0, 
                                        padx=5, 
                                        sticky="e")
                
        # Cria um widget Combobox, que é um menu dropdown que permite ao 
                # usuário selecionar uma opção de uma lista, neste caso, as 
                # colunas do arquivo Excel. 
        # 'font=('Arial', 12)' define o estilo da fonte para combinar 
                # com os outros elementos da interface.
        self.selecao_coluna = ttk.Combobox(self.menu_selecao, font=('Arial', 12))
        
        # Posiciona o Combobox no grid do contêiner 'menu_selecao'. 
        # 'row=0' indica que está na mesma linha que o rótulo de seleção de 
                # coluna, e 'column=1' coloca-o na segunda coluna. 'padx=5' 
                # mantém um pequeno espaço horizontal entre o rótulo e o Combobox 
                # para evitar que os elementos fiquem visualmente apertados.
        self.selecao_coluna.grid(row=0, column=1, padx=5)


        # Cria um rótulo dentro do contêiner 'menu_selecao'. Este rótulo serve 
                # para instruir o usuário a escolher uma condição de formatação condicional. 
        # 'text="Condição:"' é o texto que será exibido. 
        # 'bg="#f7f7f7"' define a cor de fundo igual ao contêiner 
                # para consistência visual.
        # 'font=('Arial', 12)' estabelece o estilo da fonte.
        self.rotulo_condicao = tk.Label(self.menu_selecao, 
                                        text="Condição:", 
                                        bg="#f7f7f7", 
                                        font=('Arial', 12))
        
        # Posiciona o rótulo dentro do grid do contêiner 'menu_selecao'. 
        # 'row=0' e 'column=2' colocam o rótulo na primeira linha e terceira 
                # coluna, respectivamente. 'padx=5' adiciona um espaçamento 
                # horizontal de 5 pixels para não tocar nas bordas do contêiner. 
        # 'sticky="e"' faz com que o rótulo alinhe-se ao lado direito da 
                # célula do grid para uma apresentação mais organizada.
        self.rotulo_condicao.grid(row=0, 
                                  column=2, 
                                  padx=5, 
                                  sticky="e")
                
        # Cria um widget Combobox, que é um menu dropdown que permite ao 
                # usuário selecionar uma condição de formatação condicional. 
        # O parâmetro 'values' lista as opções disponíveis no menu. 
        # 'font=('Arial', 12)' define o estilo da fonte para combinar 
                # com os outros elementos da interface.
        self.selecao_condicao = ttk.Combobox(self.menu_selecao, 
                                             values=["Maior que", "Menor que", "Igual a", "Entre", "Texto que Contém...", "Uma Data que Ocorre...", "Valores Duplicados"], 
                                             font=('Arial', 12))
        
        # Posiciona o Combobox no grid do contêiner 'menu_selecao'. 
        # 'row=0' indica que está na mesma linha que o rótulo de 
                # seleção de condição, e 'column=3' coloca-o na quarta coluna. 
        # 'padx=5' mantém um pequeno espaço horizontal entre o rótulo e o 
                # Combobox para evitar que os elementos fiquem visualmente apertados.
        self.selecao_condicao.grid(row=0, column=3, padx=5)
        
        # Adiciona um evento de vinculação ao Combobox 'selecao_condicao'. 
                # Isso significa que quando o usuário selecionar uma opção no 
                # Combobox, a função 'self.condicao_selecionada' será chamada 
                # automaticamente. 
        # Este método é útil para executar ações baseadas na escolha 
                # do usuário, como atualizar a interface ou aplicar filtros.
        self.selecao_condicao.bind("<<ComboboxSelected>>", self.condicao_selecionada)


        # Cria um rótulo (label) dentro do contêiner 'menu_selecao'. Este rótulo será 
                # usado para indicar ao usuário que ele deve inserir um valor no campo 
                # ao lado, referente à condição de formatação condicional escolhida. 
        # 'text="Valor 1:"' é o texto exibido no rótulo. 
        # 'bg="#f7f7f7"' define a cor de fundo igual ao do contêiner 
                # para manter a consistência visual. 
        # 'font=('Arial', 12)' especifica o estilo da fonte, 
                # garantindo que o texto seja claro e legível.
        self.rotulo_valor1 = tk.Label(self.menu_selecao, 
                                      text="Valor 1:", 
                                      bg="#f7f7f7", 
                                      font=('Arial', 12))
        
        # Posiciona o rótulo dentro do layout grid do contêiner 'menu_selecao'. 
        # 'row=1' coloca o rótulo na segunda linha, já que a contagem de linhas 
                # começa em 0. 'column=0' coloca o rótulo na primeira coluna. 
        # 'padx=5' adiciona um espaçamento horizontal de 5 pixels à esquerda e 
                # à direita do rótulo para evitar que ele toque diretamente as 
                # bordas do contêiner. 'sticky="e"' faz com que o rótulo alinhe-se 
                # ao lado direito da célula do grid, o que ajuda a criar uma 
                # aparência alinhada e organizada quando combinado com outros elementos.
        self.rotulo_valor1.grid(row=1, 
                                column=0, 
                                padx=5, 
                                sticky="e")
        
        # Cria um campo de entrada (Entry) dentro do contêiner 'menu_selecao'. 
                # Este campo será utilizado pelo usuário para digitar um valor 
                # que será usado na condição de formatação condicional. 
        # 'font=('Arial', 12)' define o estilo da fonte, assegurando que seja 
                # consistente com os outros elementos da interface e fácil de ler.
        self.campo_valor1 = tk.Entry(self.menu_selecao, 
                                     font=('Arial', 12))
        
        # Posiciona o campo de entrada no layout grid do contêiner 'menu_selecao'. 
        # 'row=1' indica que o campo de entrada está na mesma linha que o rótulo "Valor 1". 
        # 'column=1' coloca o campo de entrada na segunda coluna, diretamente ao 
                # lado do rótulo, facilitando a associação visual entre o rótulo e o campo. 
        # 'padx=5' adiciona um espaçamento horizontal de 5 pixels à esquerda e à 
                # direita do campo para proporcionar espaço suficiente entre 
                # este e outros elementos, evitando uma interface apertada.
        self.campo_valor1.grid(row=1, column=1, padx=5)


        # Cria um rótulo no contêiner 'menu_selecao' para indicar onde o 
                # usuário deve inserir o segundo valor necessário para 
                # algumas condições de formatação, como a condição "Entre". 
        # 'text="Valor 2:"' define o texto exibido. 
        # 'bg="#f7f7f7"' garante que a cor de fundo seja a mesma que a 
                # do contêiner para consistência visual.
        # 'font=('Arial', 12)' especifica o estilo da fonte.
        self.rotulo_valor2 = tk.Label(self.menu_selecao, 
                                      text="Valor 2:", 
                                      bg="#f7f7f7", 
                                      font=('Arial', 12))
        
        # Cria um campo de entrada no contêiner 'menu_selecao'. Este campo é 
                # utilizado pelo usuário para inserir o segundo valor nas condições que 
                # requerem dois valores, como a condição "Entre". 
        # A definição da fonte ('Arial', 12) assegura que o texto no campo de 
                # entrada seja consistente com os outros elementos da interface.
        self.campo_valor2 = tk.Entry(self.menu_selecao, font=('Arial', 12))
        
        # Cria uma variável booleana que será usada para armazenar o estado do 
                # checkbox a seguir. Este tipo de variável é útil para rastrear 
                # estados verdadeiro/falso (booleanos) dentro de interfaces gráficas.
        self.caixa_mostrar_filtrados = tk.BooleanVar()
        
        # Cria um checkbox que permite ao usuário escolher se deseja ou 
                # não ver apenas os registros filtrados conforme sua condição de 
                # formatação. 
        # 'text="Mostrar apenas filtrados"' define o texto ao lado do checkbox. 
        # 'variable=self.caixa_mostrar_filtrados' associa o estado do checkbox à 
                # variável booleana criada anteriormente. 
        # 'bg="#f7f7f7"' e 'font=('Arial', 12)' garantem a consistência 
                # visual e a legibilidade do texto.
        self.checkbox_filtrados = tk.Checkbutton(self.menu_selecao, 
                                                 text="Mostrar apenas filtrados", 
                                                 variable=self.caixa_mostrar_filtrados, 
                                                 bg="#f7f7f7", 
                                                 font=('Arial', 12))
        
        # Posiciona o checkbox dentro do layout grid do contêiner 'menu_selecao'. 
        # 'row=2' coloca o checkbox na terceira linha, aumentando a 
                # organização vertical. 
        # 'column=0', com 'columnspan=2', faz o checkbox estender-se por 
                # duas colunas, proporcionando uma aparência equilibrada e centralizada. 
        # 'padx=5' adiciona um espaçamento horizontal para manter uma margem entre o 
                # checkbox e outros elementos da interface.
        self.checkbox_filtrados.grid(row=2, 
                                     column=0, 
                                     columnspan=2, 
                                     padx=5)


        # Cria um botão no contêiner 'menu_selecao'. Este botão será usado pelo usuário 
                # para aplicar a formatação condicional selecionada aos dados. 
        # 'text="Aplicar Formatação"' define o texto no botão. 
        # 'command=self.aplicar_formatacao_condicional' associa este 
                # botão à função que executa a lógica de formatação condicional. 
        # 'bg="#008CBA"' e 'fg="white"' definem, respectivamente, a cor de fundo 
                # e a cor do texto do botão.
        # 'font=('Arial', 12, 'bold')' especifica o estilo da fonte, 
                # destacando o botão com texto em negrito.
        self.botao_aplicar_formatacao = tk.Button(self.menu_selecao, 
                                                  text="Aplicar Formatação", 
                                                  command=self.aplicar_formatacao_condicional, 
                                                  bg="#008CBA", 
                                                  fg="white", 
                                                  font=('Arial', 12, 'bold'))
        
        # Posiciona o botão "Aplicar Formatação" no grid do contêiner 'menu_selecao'. 
        # 'row=2' coloca o botão na terceira linha, alinhando-o com os outros 
                # elementos na mesma linha. 'column=2' coloca o botão na 
                # terceira coluna. 
        # 'padx=5' adiciona espaçamento horizontal para manter uma margem 
                # adequada entre este botão e outros elementos.
        self.botao_aplicar_formatacao.grid(row=2, column=2, padx=5)
        
        # Cria outro botão no contêiner 'menu_selecao', destinado a permitir 
                # que o usuário exporte os dados filtrados para um arquivo. 
        # 'text="Exportar Dados Filtrados"' é o texto exibido no botão. 
        # 'command=self.exportar_dados_filtrados' liga o botão à função 
                # que maneja a exportação de dados. 
        # 'bg="#f44336"' e 'fg="white"' definem as cores de fundo e do 
                # texto, respectivamente, e 'font=('Arial', 12, 'bold')' 
                # realça o botão com um estilo de fonte em negrito.
        self.botao_exportar_dados = tk.Button(self.menu_selecao, 
                                              text="Exportar Dados Filtrados", 
                                              command=self.exportar_dados_filtrados, 
                                              bg="#f44336", 
                                              fg="white", 
                                              font=('Arial', 12, 'bold'))
        
        # Posiciona o botão "Exportar Dados Filtrados" dentro do grid do contêiner 'menu_selecao'. 
        # 'row=2' garante que ele esteja alinhado com o botão "Aplicar Formatação" na 
                # mesma linha. 
        # 'column=3' coloca-o na quarta coluna, separando-o adequadamente do 
                # botão de aplicar formatação. 'padx=5' adiciona um espaçamento 
                # horizontal, proporcionando um layout equilibrado e acessível.
        self.botao_exportar_dados.grid(row=2, column=3, padx=5)


        # Cria um contêiner chamado 'container_tabela' dentro do contêiner 
                # principal 'container_principal'. 
        # Um Frame em Tkinter é usado para agrupar e organizar visualmente outros 
                # widgets. Este contêiner específico será usado para 
                # exibir a tabela de dados.
        self.container_tabela = tk.Frame(self.container_principal)
        
        # Posiciona o 'container_tabela' dentro do layout grid do contêiner principal. 
        # 'row=3' coloca o contêiner na quarta linha, 'column=0' inicia o 
                # posicionamento na primeira coluna, e 'columnspan=4' faz com que o 
                # contêiner se estenda por todas as quatro colunas disponíveis, maximizando o 
                # espaço horizontal. 
        # 'pady=10' adiciona um espaçamento vertical de 10 pixels acima e abaixo do 
                # contêiner para separação visual. 
        # 'sticky="nsew"' faz com que o contêiner se expanda em todas as 
                # direções (norte, sul, leste, oeste), assegurando que ele 
                # preencha completamente o espaço alocado.
        self.container_tabela.grid(row=3, 
                                   column=0, 
                                   columnspan=4, 
                                   pady=10, 
                                   sticky="nsew")
        
        # Configura a linha 3 do grid no contêiner principal para ter 
                # um 'weight' de 1. 
        # Isso significa que essa linha pode expandir e contrair 
                # proporcionalmente com a janela quando ela é redimensionada. 
        # A configuração de peso é crucial para uma interface responsiva, 
                # garantindo que a tabela se ajuste dinamicamente ao 
                # tamanho da janela.
        self.container_principal.grid_rowconfigure(3, weight=1)
        
        # Configura a coluna 0 do grid no contêiner principal para também 
                # ter um 'weight' de 1. 
        # Isso permite que a coluna se expanda e ocupe o espaço disponível, 
                # assegurando que a tabela se ajuste horizontalmente ao 
                # redimensionar a janela. A combinação das configurações de 
                # peso nas linhas e colunas promove uma flexibilidade na 
                # disposição dos componentes da interface, adaptando-se 
                # melhor a diferentes tamanhos de tela.
        self.container_principal.grid_columnconfigure(0, weight=1)


    
    # Definição do método 'carregar_arquivo' que é um 
                # componente da classe 'Aplicativo'.
    def carregar_arquivo(self):
        
        # Abre uma caixa de diálogo que permite ao usuário selecionar um 
                # arquivo para abrir. 'filetypes' restringe os tipos de arquivo 
                # que podem ser escolhidos para arquivos Excel ('.xlsx' e '.xls').
        self.caminho_arquivo = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        
        # Verifica se um caminho de arquivo foi realmente selecionado (ou seja, o 
                # usuário não cancelou a operação). 
        # Se um arquivo foi escolhido, o bloco de código dentro 
                # do 'if' será executado.
        if self.caminho_arquivo:
            
            # Atualiza o texto do rótulo 'rotulo_nome_arquivo' para mostrar o nome 
                    # do arquivo selecionado, usando 'os.path.basename' para 
                    # extrair apenas o nome do arquivo do caminho completo.
            self.rotulo_nome_arquivo.config(text=os.path.basename(self.caminho_arquivo))
            
            # Usa a biblioteca pandas para ler o arquivo Excel selecionado e 
                    # armazenar os dados na variável 'dados_excel'. 
            # Isso transforma os dados do Excel em um DataFrame do pandas, 
                    # que é uma estrutura de dados bidimensional, semelhante a 
                    # uma tabela, que é altamente otimizada para operações de dados.
            self.dados_excel = pd.read_excel(self.caminho_arquivo)
            
            # Chama o método 'preencher_tabela', que é responsável 
                    # por preencher a tabela na interface gráfica com os 
                    # dados carregados do arquivo Excel. Esse método irá 
                    # configurar a visualização da tabela com os dados 
                    # recém-carregados.
            self.preencher_tabela()
            
            # Chama o método 'preencher_selecao_coluna', que atualiza o 
                    # ComboBox de seleção de coluna na interface para listar as 
                    # colunas do arquivo Excel carregado. Isso permite ao usuário 
                    # escolher uma coluna para aplicar a formatação condicional.
            self.preencher_selecao_coluna()


    # Definição do método 'preencher_tabela' dentro da classe 'Aplicativo'.
    def preencher_tabela(self):
        
        # Itera sobre todos os widgets (componentes gráficos) que estão 
                # dentro do contêiner 'container_tabela'. 
        # 'winfo_children()' retorna uma lista de todos os widgets 
                # filhos dentro deste contêiner.
        for widget in self.container_tabela.winfo_children():
            
            # Remove cada widget encontrado na iteração anterior do contêiner. 
            # Isso é útil para limpar quaisquer tabelas ou elementos 
                    # antigos antes de preencher o contêiner com novos dados.
            widget.destroy()
    
        # Cria um widget Treeview dentro do 'container_tabela'. 
        # 'Treeview' é usado para mostrar os dados em forma de tabela. 
        # 'columns=list(self.dados_excel.columns)' define as colunas da 
                # tabela com base nas colunas do DataFrame 'dados_excel'. 
        # 'show='headings'' configura o Treeview para mostrar apenas os 
                # cabeçalhos das colunas, sem a coluna de índice padrão.
        self.tabela = ttk.Treeview(self.container_tabela, 
                                   columns=list(self.dados_excel.columns), 
                                   show='headings')
    
        # Posiciona a tabela dentro do 'container_tabela'. 
        # 'pack' é um gerenciador de geometria que organiza widgets 
                # em blocos antes de colocá-los no contêiner pai. 
        # 'fill=tk.BOTH' faz com que o widget expanda tanto horizontal 
                # quanto verticalmente para preencher todo o espaço 
                # disponível no contêiner. 
        # 'expand=True' permite que o widget se expanda para ocupar 
                # qualquer espaço extra na janela principal.
        self.tabela.pack(fill=tk.BOTH, expand=True)
    
        # Itera sobre cada coluna no DataFrame 'dados_excel'.
        for col in self.dados_excel.columns:
            
            # Configura o cabeçalho de cada coluna na Treeview 
                    # com o nome da coluna correspondente.
            self.tabela.heading(col, text=col)
            
            # Define a largura inicial de cada coluna na 
                    # Treeview para 100 pixels.
            self.tabela.column(col, width=100)
    
        # Itera sobre o DataFrame 'dados_excel'. 
        # 'iterrows()' retorna um gerador que produz índice e 
                # linha, onde 'idx' é o índice da linha e 'row' é 
                # a série de dados para essa linha.
        for idx, row in self.dados_excel.iterrows():
            
            # Insere cada linha de dados no Treeview. '""' e '"end"' especificam 
                    # que a linha deve ser adicionada ao final da lista de linhas. 
            # 'values=list(row)' converte os dados da linha atual em uma lista 
                    # para serem exibidos corretamente. 'iid=str(idx)' define o 
                    # identificador interno da linha como uma string do índice, o 
                    # que é útil para referenciar a linha especificamente 
                    # mais tarde, se necessário.
            self.tabela.insert("", "end", values=list(row), iid=str(idx))


    # Definição do método 'preencher_selecao_coluna' dentro 
            # da classe 'Aplicativo'.
    def preencher_selecao_coluna(self):
        
        # Atualiza o widget Combobox 'selecao_coluna' com as colunas do
                # DataFrame 'dados_excel'. 
        # Isso permite ao usuário selecionar de uma lista de colunas do 
                # arquivo Excel carregado. 
        # 'list(self.dados_excel.columns)' converte os nomes das 
                # colunas em uma lista para ser usada no Combobox.
        self.selecao_coluna['values'] = list(self.dados_excel.columns)
        
        # Define o item atualmente selecionado no Combobox para o 
                # primeiro item da lista (índice 0), garantindo que 
                # uma coluna esteja sempre selecionada por padrão 
                # quando os dados forem carregados.
        self.selecao_coluna.current(0)

    

    # Definição do método 'condicao_selecionada', que é chamado 
            # quando um item é selecionado no Combobox 'selecao_condicao'.
    def condicao_selecionada(self, event):
        
        # Recupera a condição selecionada no Combobox 'selecao_condicao'.
        condicao = self.selecao_condicao.get()
        
        # 'Verifica se a condição selecionada é "Entre", 
                # que requer dois valores para a comparação.
        if condicao == "Entre":
            
            # Exibe o campo de entrada 'campo_valor2' no grid. 
            # É posicionado na segunda linha (row=1), quarta 
                    # coluna (column=3) e com um espaçamento horizontal 
                    # de 5 pixels (padx=5). 
            # Isso torna o campo visível apenas quando necessário.
            self.campo_valor2.grid(row=1, column=3, padx=5)
            
            # Exibe o rótulo 'rotulo_valor2' no grid, posicionando-o 
                    # próximo ao 'campo_valor2'. É colocado na segunda linha (row=1), 
                    # terceira coluna (column=2), com um espaçamento horizontal de 5 
                    # pixels (padx=5), e alinhado à direita da célula (sticky="e").
            self.rotulo_valor2.grid(row=1, 
                                    column=2, 
                                    padx=5, 
                                    sticky="e")
            
        else:
            
            # Oculta o campo de entrada 'campo_valor2' se a condição 
                    # selecionada não for "Entre". 
            # Isso remove o widget do layout sem destruí-lo, tornando-o 
                    # invisível e não ocupando espaço.
            self.campo_valor2.grid_remove()
            
            # Oculta o rótulo 'rotulo_valor2' se a condição selecionada 
                    # não for "Entre". 
            # Funciona da mesma maneira que 'grid_remove()' para o campo de 
                    # entrada, removendo o rótulo do layout.
            self.rotulo_valor2.grid_remove()



    # Definição do método 'aplicar_formatacao_condicional' 
            # dentro da classe 'Aplicativo'.
    def aplicar_formatacao_condicional(self):
        
        # Chamado para garantir que a tabela seja atualizada e limpa 
                # de qualquer formatação anterior antes de aplicar uma 
                # nova formatação condicional. 
        # Isso ajuda a evitar sobreposições ou erros de exibição 
                # causados por dados antigos ainda presentes.
        self.preencher_tabela()
    
        # Obtém a coluna que o usuário selecionou no ComboBox 
                # 'selecao_coluna' para aplicar a formatação condicional.
        coluna_selecionada = self.selecao_coluna.get()
        
        # Obtém a condição de formatação condicional selecionada 
                # pelo usuário no ComboBox 'selecao_condicao'.
        condicao = self.selecao_condicao.get()
        
        # Recupera o valor inserido pelo usuário no campo de entrada 'campo_valor1', 
                # que será usado para comparar com os dados da coluna 
                # selecionada conforme a condição especificada.
        valor1 = self.campo_valor1.get()
        
        # Recupera o segundo valor inserido pelo usuário no campo 
                # de entrada 'campo_valor2', usado principalmente para 
                # condições que exigem um intervalo, como a condição "Entre".
        valor2 = self.campo_valor2.get()
    
        # Converte os dados da coluna selecionada em uma série do pandas. 
        # Uma série é uma estrutura de dados do pandas que é como 
                # uma coluna em uma tabela de dados, facilitando 
                # operações e comparações.
        serie_coluna = self.dados_excel[coluna_selecionada]
    
        # Cria uma lista vazia que será usada para armazenar os 
                # identificadores das linhas que não cumprem a condição 
                # especificada e que devem ser ocultadas na visualização da tabela.
        itens_para_ocultar = []

        # Itera sobre cada linha (ou "item") na tabela. 
        # 'get_children()' retorna todos os itens presentes no widget 
                # Treeview, que neste caso são as linhas da tabela.
        for item in self.tabela.get_children():
            
            # Converte o identificador do item (que é uma string) 
                    # para um inteiro. Este identificador é usado 
                    # para indexar as linhas na série de dados.
            idx = int(item)
            
            # Obtém o valor da série de dados na posição indexada. 
            # 'iloc' é um método do pandas usado para acessar o 
                    # elemento na posição indicada.
            valor = serie_coluna.iloc[idx]
            
            # Inicializa uma variável booleana que determinará se uma 
                    # formatação específica (como mudança de cor) 
                    # deve ser aplicada à linha corrente.
            aplicar_cor = False
        
            # Inicia um loop sobre todos os itens na tabela 
                    # para verificar condições individuais em cada linha.
            for item in self.tabela.get_children():
                
                # Converte o identificador do item de string para 
                        # inteiro, pois é armazenado como string no Treeview.
                idx = int(item)
                
                # Acessa o valor na coluna selecionada para o 
                        # item atual usando o índice.
                valor = serie_coluna.iloc[idx]
                
                # Inicializa a variável que indica se a 
                        # formatação especial (cor) será aplicada.
                aplicar_cor = False
            
                # Tenta executar as operações que podem lançar 
                        # um erro de tipo ou valor.
                try:
                    
                    # Verifica se a condição selecionada é "Maior que".
                    if condicao == "Maior que":
                        
                        # Compara numericamente o valor da coluna com o 'valor1' 
                                # fornecido, usando conversão para tratar não números.
                        aplicar_cor = pd.to_numeric(valor, errors='coerce') > pd.to_numeric(valor1, errors='coerce')
                    
                    # Verifica se a condição selecionada é "Menor que".
                    elif condicao == "Menor que":
                        
                        # Compara numericamente se o valor da coluna é menor que o 'valor1'.
                        aplicar_cor = pd.to_numeric(valor, errors='coerce') < pd.to_numeric(valor1, errors='coerce')
                    
                    # Verifica se a condição selecionada é "Igual a".
                    elif condicao == "Igual a":
                        
                        # Compara diretamente o valor da coluna (como string) 
                                # com o 'valor1'.
                        aplicar_cor = (str(valor) == valor1)
                    
                    # Verifica se a condição selecionada é "Entre".
                    elif condicao == "Entre":
                        
                        # Verifica se o valor da coluna está entre 'valor1' e 'valor2'.
                        aplicar_cor = pd.to_numeric(valor1, errors='coerce') <= pd.to_numeric(valor, errors='coerce') <= pd.to_numeric(valor2, errors='coerce')
                    
                    # Verifica se a condição é "Texto que Contém...".
                    elif condicao == "Texto que Contém...":
                        
                        # Checa se o 'valor1' está contido no valor da 
                                # coluna (considerado como string).
                        aplicar_cor = valor1 in str(valor)
                    
                    # Verifica se a condição é "Uma Data que Ocorre...".
                    elif condicao == "Uma Data que Ocorre...":
                        
                        # Compara se a data no valor da coluna corresponde à 
                                # 'valor1', tratando ambos como datas.
                        aplicar_cor = pd.to_datetime(valor, errors='coerce', dayfirst=True) == pd.to_datetime(valor1, errors='coerce', dayfirst=True)
                    
                    # Verifica se a condição é "Valores Duplicados".
                    elif condicao == "Valores Duplicados":
                        
                        # Determina se o valor na coluna está duplicado 
                                # no conjunto de dados.
                        aplicar_cor = serie_coluna.duplicated(keep=False).iloc[idx]
                
                # Captura erros de valor que possam ocorrer durante a 
                        # conversão de tipos ou comparações.
                except ValueError:
                    
                    # Ignora a iteração atual e continua para a próxima.
                    continue
            
                # Verifica se a condição para aplicar a formatação 
                        # especial foi satisfeita.
                if aplicar_cor:
                    
                    # Aplica uma tag de destaque à linha atual, que pode 
                            # ser usada para mudar a cor de fundo.
                    self.tabela.item(item, tags=("highlight",))
                    
                else:
                    
                    # Adiciona o identificador do item à lista de 
                            # itens a serem ocultados.
                    itens_para_ocultar.append(item)

        
            # Se 'aplicar_cor' é True, aplica uma formatação 
                    # especial à linha (como mudança de cor).
            if aplicar_cor:
                
                self.tabela.item(item, tags=("highlight",))
                
            else:
                
                # Se a condição não for verdadeira, adiciona o identificador 
                        # do item à lista de itens para ocultar.
                itens_para_ocultar.append(item)


        # Verifica se a opção 'Mostrar apenas filtrados' 
                # está ativada.
        if self.caixa_mostrar_filtrados.get():
            
            # Se a opção está ativada, itera sobre os itens que não 
                    # atendem ao critério de formatação condicional.
            for item in itens_para_ocultar:
                
                # 'self.tabela.detach(item)' remove temporariamente a linha 
                        # especificada da visualização sem deletá-la permanentemente. 
                # Isso oculta as linhas que não atendem ao critério especificado, 
                        # permitindo que apenas as linhas que atendem sejam mostradas.
                self.tabela.detach(item)
                
        else:
            
            # Se a opção 'Mostrar apenas filtrados' não está ativada, 
                    # todas as linhas devem ser mostradas.
            for item in self.tabela.get_children():
                
                # 'self.tabela.reattach(item, '', 'end')' reinsere todas as linhas na 
                        # tabela. O método 'reattach' é usado para adicionar novamente 
                        # linhas que foram previamente 'detachadas'. O primeiro argumento é 
                        # o identificador da linha, o segundo é a nova localização parente, e 
                        # o terceiro é a posição na qual a linha deve ser inserida.
                self.tabela.reattach(item, '', 'end')
        
        # Configura um estilo visual para as linhas que são destacadas 
                # pela formatação condicional.
        # 'Define a configuração de uma tag chamada "highlight". Qualquer 
                # linha marcada com esta tag terá um fundo amarelo. 
        # O método 'tag_configure' permite personalizar o estilo de itens 
                # individuais dentro de um widget Treeview.
        self.tabela.tag_configure("highlight", background="yellow")


        
    # Definição do método 'exportar_dados_filtrados' dentro da classe 'Aplicativo'.
    def exportar_dados_filtrados(self):
        
        # Verifica se há dados carregados no DataFrame 'dados_excel'.
        if self.dados_excel is not None:
            
            # Abre uma caixa de diálogo que permite ao usuário escolher o 
                    # local e o nome do arquivo para salvar. 
            # 'defaultextension=".xlsx"' define a extensão padrão do arquivo 
                    # como .xlsx. 
            # 'filetypes=[("Excel files", "*.xlsx *.xls")]' limita os 
                    # tipos de arquivos que o usuário pode salvar.
            caminho_exportacao = filedialog.asksaveasfilename(defaultextension=".xlsx", 
                                                              filetypes=[("Excel files", "*.xlsx *.xls")])
            
            # Verifica se um caminho de arquivo foi escolhido. 
            # Se o usuário cancelar a operação, 'caminho_exportacao' 
                    # será uma string vazia, e esse bloco de código não será executado.
            if caminho_exportacao:
                
                # Cria um novo objeto Workbook usando openpyxl, que é uma 
                        # biblioteca para ler e escrever arquivos Excel. 
                # Um Workbook é um arquivo Excel completo.
                workbook = openpyxl.Workbook()
                
                # 'workbook.active' obtém a planilha ativa no Workbook 
                        # recém-criado. 
                # Por padrão, um novo Workbook vem com uma planilha.
                planilha = workbook.active


                # Itera sobre as colunas do DataFrame 'dados_excel', 
                        # usando a função 'enumerate' para obter tanto o índice 
                        # da coluna (começando de 1) quanto o nome da coluna. 
                # A enumeração começa em 1 para corresponder à convenção de 
                        # numeração de colunas do Excel, onde a primeira coluna é 1 e não 0.
                for col_num, col_nome in enumerate(self.dados_excel.columns, 1):
                    
                    # Acessa a célula na primeira linha e na coluna especificada 
                            # por 'col_num'. 
                    # O método 'cell()' é usado para acessar ou criar uma 
                            # célula na planilha. 'value = col_nome' define o 
                            # valor dessa célula para o nome da coluna, 
                            # efetivamente criando os cabeçalhos da tabela no arquivo Excel.
                    planilha.cell(row=1, column=col_num).value = col_nome


                # Inicia um loop para iterar sobre cada linha (item) na tabela, 
                        # começando a contagem de linhas a partir de 2, 
                        # pois a linha 1 contém os cabeçalhos.
                for row_num, item in enumerate(self.tabela.get_children(), 2):
                    
                    # Recupera os valores armazenados em cada célula da linha 
                            # corrente. 
                    # O método 'item' do widget Treeview é usado para 
                            # acessar o conteúdo de cada item (linha).
                    valores = self.tabela.item(item, 'values')
                    
                    # Itera sobre cada valor na linha, juntamente com seu índice 
                            # de coluna (começando de 1, para corresponder à 
                            # numeração de colunas do Excel).
                    for col_num, valor in enumerate(valores, 1):
                        
                        # Acessa a célula na planilha no local especificado 
                                # pelo número da linha e da coluna. Se a célula não 
                                # existir, ela será criada.
                        celula = planilha.cell(row=row_num, column=col_num)
                        
                        # Define o valor da célula para o dado contido 
                                # na linha da tabela.
                        celula.value = valor
                        
                        # Verifica se a linha atual está marcada com a tag 'highlight', 
                                # que indica que a linha deve ser formatada de maneira especial.
                        if "highlight" in self.tabela.item(item, 'tags'):
                            
                            # Aplica uma formatação de cor amarela à célula. 'PatternFill' é 
                                    # usado para preencher a célula com uma cor sólida, onde 
                                    # 'start_color' e 'end_color' definem a cor do preenchimento.
                            celula.fill = PatternFill(start_color="FFFF00", 
                                                      end_color="FFFF00", 
                                                      fill_type="solid")


                # Salva o arquivo Excel criado e preenchido no caminho 
                        # especificado pelo usuário. Este é o caminho que o 
                        # usuário escolheu ou confirmou na caixa de diálogo de salvar arquivo.
                workbook.save(caminho_exportacao)
                
                # Exibe uma caixa de mensagem informativa ao usuário final. 
                # "Exportação" é o título da janela da caixa de mensagem.
                # "Dados exportados com sucesso!" é a mensagem mostrada. Isso fornece um feedback visual 
                        # de que o processo de exportação foi concluído sem erros.
                messagebox.showinfo("Exportação", "Dados exportados com sucesso!")
            


# Cria uma nova janela raiz Tkinter. 
        # 'raiz' será a janela principal da aplicação.
raiz = tk.Tk()

# Cria uma instância da classe 'Aplicativo', passando a 
        # janela raiz como argumento. Isso inicializa a aplicação, 
        # configurando todos os widgets e lógicas associadas.
aplicativo = Aplicativo(raiz)

# Inicia o loop principal da aplicação Tkinter.
raiz.mainloop()