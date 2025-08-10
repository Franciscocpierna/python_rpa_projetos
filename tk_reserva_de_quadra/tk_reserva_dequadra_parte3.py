import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
from datetime import date, datetime, timedelta
import openpyxl
from openpyxl import Workbook

import os
import pandas as pd

#Lista que armazena as reservas
reservas = []

#C:\Users\55119\Desktop\Reserva Quadra

def verificar_conflito_reserva(data, hora_inicial, hora_final):
    
    #Caminho para o arquivo do Excel que contém as reservas
    caminho = r"C:\Users\55119\Desktop\Reserva Quadra\Dados.xlsx"
    
    #Nome da planilha que contém as reservas
    planilha = "Reservas"
    
    try:
        
        #Carrega o arquivo do Excel
        workbook = openpyxl.load_workbook(caminho)
        
        #Obtém a planilha especifica
        sheet = workbook[planilha]
    
    except FileNotFoundError:
        
        return False  # Se o arquivo não existir, não há conflitos
    
    # Itera sobre as linhas da planilha, a partir da segunda linha (min_row=2)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        
        # Obtém os valores da data, hora inicial e hora final da reserva atual
        reserva_data = row[0]
        reserva_hora_inicial = row[1]
        reserva_hora_final = row[2]

        # Verifica se a data da nova reserva é igual à data da reserva atual
        # e se há sobreposição de horários entre a nova reserva e a reserva atual
        if data == reserva_data and hora_inicial < reserva_hora_final and hora_final > reserva_hora_inicial:
            
            return True  # Há um conflito de horários
    
    return False  # Não há conflito de horários

# Função que calcula o valor total a ser pago com base nas horas inicial e final da reserva
def calcular_valor(hora_inicial, hora_final):
    
        #09
        # Obtém a parte das horas da hora inicial
        hora_inicial_str = hora_inicial[:-3]

        #30
        # Obtém a parte dos minutos da hora inicial
        minuto_inicial_str = hora_inicial[-2:]

        #10
        # Obtém a parte das horas da hora final
        hora_final_str = hora_final[:-3]

        #30
        # Obtém a parte dos minutos da hora final
        minuto_final_str = hora_final[-2:]
        
        """
        Digite a hora inicial da reserva (entre 09:00 e 23:00): 09:00
        Digite a hora final da reserva (entre 09:00 e 23:00): 10:00
        hora_inicial_str: 09
        minuto_inicial_str: 00
        hora_final_str: 10
        minuto_final_str: 00
        """
        
        """
        (int(hora_final_str) * 60 + int(minuto_final_str)): Essa expressão realiza o 
        cálculo para converter a hora final da reserva em minutos. Primeiro, convertemos
        a parte das horas, hora_final_str, em um número inteiro usando int(hora_final_str). 
        Em seguida, multiplicamos esse valor por 60 para converter as horas em minutos. 
        Por fim, adicionamos a parte dos minutos, minuto_final_str, também convertida em um 
        número inteiro usando int(minuto_final_str). Assim, obtemos o total de minutos 
        correspondente à hora final da reserva.

        (int(hora_inicial_str) * 60 + int(minuto_inicial_str)): Essa expressão realiza o 
        mesmo cálculo descrito anteriormente, mas para a hora inicial da reserva. Convertemos 
        a parte das horas, hora_inicial_str, em um número inteiro usando int(hora_inicial_str). 
        Em seguida, multiplicamos esse valor por 60 para converter as horas em minutos. 
        Por fim, adicionamos a parte dos minutos, minuto_inicial_str, também convertida em um 
        número inteiro usando int(minuto_inicial_str). Assim, obtemos o total de minutos 
        correspondente à hora inicial da reserva.

        (int(hora_final_str) * 60 + int(minuto_final_str)) - (int(hora_inicial_str) * 60 + int(minuto_inicial_str)): 
        Nessa etapa, subtraímos o valor total de minutos da hora inicial do valor total de 
        minutos da hora final. Isso nos dá a duração da reserva em minutos. Por exemplo, se a hora 
        inicial for 09:00 e a hora final for 10:30, teremos (10 * 60 + 30) - (9 * 60 + 0), que 
        resultará em 90 minutos, correspondendo à duração da reserva.

        valor_total = duracao_minutos * 8 / 60: Após calcular a duração da reserva em 
        minutos, multiplicamos esse valor por 8, que é o valor por minuto para calcular 
        o valor total da reserva. Como o preço é dado por minuto, dividimos por 60 para 
        converter os minutos em horas. O resultado é armazenado na variável valor_total.
        """
        duracao_minutos = (int(hora_final_str) * 60 + int(minuto_final_str)) - (int(hora_inicial_str) * 60 + int(minuto_inicial_str))
        valor_total = duracao_minutos * 8 / 60
        
        return valor_total
    

def fazer_reserva(data):
    
    # Cria uma nova janela usando Tkinter
    janela_fazer_reserva = tk.Toplevel()
    janela_fazer_reserva.title("Fazer Reserva")
    janela_fazer_reserva.configure(bg = "#FFFFFF")
    
    # Cria um frame para o título e o coloca na janela
    frame_titulo = tk.Frame(janela_fazer_reserva, bg = "#FFFFFF")
    frame_titulo.pack(pady = 20)
    
    label_titulo = tk.Label(frame_titulo, 
                            text="Fazer Reserva",
                            font = "Arial 24 bold", 
                            bg = "#FFFFFF")
    label_titulo.pack()
    
    # Cria um frame para o formulario e o coloca na janela
    frame_formulario = tk.Frame(janela_fazer_reserva, bg = "#FFFFFF")
    frame_formulario.pack(pady = 20)
    
    
    # Cria um label para a data, configura sua fonte e o coloca no frame do formulário
    label_data = tk.Label(frame_formulario, 
                            text="Digite a data da reserva (dd/mm/aaaa)",
                            font = "Arial 18", 
                            bg = "#FFFFFF")
    label_data.grid(row = 0, column = 0, pady = 10, stick = "e")
    
    # Cria uma entrada de texto para a data, configura sua fonte e a coloca no frame do formulário
    entry_data = tk.Entry(frame_formulario,
                            font = "Arial 18")
    entry_data.grid(row = 0, column = 1, pady = 10, padx = 10)
    entry_data.insert(0, data)
    
    #-----------------------------------------------------------------
    
    # Cria um label para a data, configura sua fonte e o coloca no frame do formulário
    label_inicial = tk.Label(frame_formulario, 
                            text="Digite a hora inicial da reserva (entre 09:00 e 23:00)",
                            font = "Arial 18", 
                            bg = "#FFFFFF")
    label_inicial.grid(row = 1, column = 0, pady = 10, stick = "e")
    
    # Cria uma entrada de texto para a data, configura sua fonte e a coloca no frame do formulário
    entry_inicial = tk.Entry(frame_formulario,
                            font = "Arial 18")
    entry_inicial.grid(row = 1, column = 1, pady = 10, padx = 10)
    
    #-----------------------------------------------------------------
    
    # Cria um label para a data, configura sua fonte e o coloca no frame do formulário
    label_final = tk.Label(frame_formulario, 
                            text="Digite a hora final da reserva (entre 09:00 e 23:00)",
                            font = "Arial 18", 
                            bg = "#FFFFFF")
    label_final.grid(row = 2, column = 0, pady = 10, stick = "e")
    
    # Cria uma entrada de texto para a data, configura sua fonte e a coloca no frame do formulário
    entry_final = tk.Entry(frame_formulario,
                            font = "Arial 18")
    entry_final.grid(row = 2, column = 1, pady = 10, padx = 10)
    
    #-----------------------------------------------------------------
    
    # Cria um label para a data, configura sua fonte e o coloca no frame do formulário
    label_cpf = tk.Label(frame_formulario, 
                            text="Digite o CPF",
                            font = "Arial 18", 
                            bg = "#FFFFFF")
    label_cpf.grid(row = 3, column = 0, pady = 10, stick = "e")
    
    # Cria uma entrada de texto para a data, configura sua fonte e a coloca no frame do formulário
    entry_cpf = tk.Entry(frame_formulario,
                            font = "Arial 18")
    entry_cpf.grid(row = 3, column = 1, pady = 10, padx = 10)
    
    #-----------------------------------------------------------------
    
    # Cria um label para a data, configura sua fonte e o coloca no frame do formulário
    label_nome = tk.Label(frame_formulario, 
                            text="Digite o Nome",
                            font = "Arial 18", 
                            bg = "#FFFFFF")
    label_nome.grid(row = 4, column = 0, pady = 10, stick = "e")
    
    # Cria uma entrada de texto para a data, configura sua fonte e a coloca no frame do formulário
    entry_nome = tk.Entry(frame_formulario,
                            font = "Arial 18")
    entry_nome.grid(row = 4, column = 1, pady = 10, padx = 10)
    
    
    # Cria uma entrada de texto para a data, configura sua fonte e a coloca no frame do formulário
    entry_valor = tk.Entry(frame_formulario,
                            font = "Arial 18")
    entry_valor.grid(row = 5, column = 1, pady = 10, padx = 10)
    
    def preecher_valor():
        
        #Obtém as horas iniciais e finais a partir das entradas de texto
        hora_inicial = entry_inicial.get()
        hora_final = entry_final.get()
        
        #Calcula o valor total com base nas horas iniciais e finais
        valor_total = calcular_valor(hora_inicial, hora_final)
        
        #Limpa o campo de valor
        entry_valor.delete(0, tk.END)
        
        #Insere o valor total calculado no campo de valor
        entry_valor.insert(0, valor_total)
        
    
    btn_calcular_valor = tk.Button(frame_formulario,
                            text = "Calcular Valor",
                            font = "Arial 20",
                            command = preecher_valor)
    btn_calcular_valor.grid(row = 5, column = 0, pady = 10, padx = 10, stick = "e")
    
    
    def fazer_reserva_click():
        
        #Obtém os valores das entradas de texto
        data_da_reserva = entry_data.get()
        hora_inicial = entry_inicial.get()
        hora_final = entry_final.get()
        cpf = entry_cpf.get()
        nome = entry_nome.get()
        valor_total = float(entry_valor.get())

        # Verifica se o horário é inválido (fora dos limites permitidos ou com hora inicial maior ou igual à hora final)
        if hora_inicial < "09:00" or hora_final > "23:00" or hora_inicial >= hora_final:
            
            tk.messagebox.showerror("Horário inválido", "Horário inválido. Tente novamente.")
            
            return

        # Verifica se há conflito de horários com reservas existentes
        if verificar_conflito_reserva(data_da_reserva, hora_inicial, hora_final):
            
            tk.messagebox.showerror("Conflito de horários", "Já existe uma reserva neste horário. Escolha outro horário.")
            
            return
        
        # Caminho do arquivo do Excel e nome da planilha
        caminho = r"C:\Users\55119\Desktop\Reserva Quadra\Dados.xlsx"
        planilha = "Reservas"
        
        try:
            
            # Tenta carregar o arquivo do Excel e obter a planilha
            workbook = openpyxl.load_workbook(caminho)
            sheet = workbook[planilha]
            
            
        except FileNotFoundError:
            
            print("Arquivo não encontrado!")
            
        #Adiciona os dados da reserva na planilha
        sheet.append([data_da_reserva, hora_inicial, hora_final, cpf, nome, valor_total])
        workbook.save(caminho)
        
        #Exibe uma mensagem de sucesso com os detalhes da reserva
        mensagem = f"\nA reserva das {hora_inicial} às {hora_final} na data {data_da_reserva} foi feita com sucesso."
        mensagem += "\nO valor total da reserva é de R${:.2f}\n".format(valor_total)
        
        tk.messagebox.showinfo("Reserva feita com sucesso", mensagem)
        
        #Fecha a janela de reservas
        janela_fazer_reserva.destroy()
            
    
    # Cria um botão "Confirmar" que chama a função fazer_reserva_click quando clicado
    btn_confirmar = tk.Button(janela_fazer_reserva,
                            text = "Confirmar",
                            font = "Arial 20",
                            width = 50,
                            command = fazer_reserva_click)
    btn_confirmar.pack(pady = 20)
    
    #Centraliza a janela
    janela_fazer_reserva.update_idletasks()

    # Obtém a largura atual da janela
    largura = janela_fazer_reserva.winfo_width()

    # Obtém a altura atual da janela
    altura = janela_fazer_reserva.winfo_height()

    # Calcula a posição x para centralizar a janela na tela
    posicao_x = (janela_fazer_reserva.winfo_screenwidth() // 2) - (largura // 2)

    # Calcula a posição y para centralizar a janela na tela
    posicao_y = (janela_fazer_reserva.winfo_screenheight() // 2) - (altura // 2)

    # Define a geometria da janela com a largura, altura, posição x e posição y
    janela_fazer_reserva.geometry('{}x{}+{}+{}'.format(largura, altura, posicao_x, posicao_y))
    

def ver_reservas():
    
    janela_ver_reservas = tk.Toplevel()
    janela_ver_reservas.title("Ver Reservas")
    janela_ver_reservas.configure(bg="#FFFFFF")
    
    
    # Cria um frame para conter os botões
    frame_titulo = tk.Frame(janela_ver_reservas,
                           bg = "#FFFFFF")
    frame_titulo.pack(pady = 20) # Coloca o frame na janela com um espaçamento vertical

    
    label_titulo = tk.Label(frame_titulo,
                       text = "Visualização de Reservas",
                       font = "Arial 24 bold",
                       bg = "#FFFFFF")
    label_titulo.pack()
    
    # Cria um frame para conter os botões
    frame_reservas = tk.Frame(janela_ver_reservas,
                           bg = "#FFFFFF")
    frame_reservas.pack(pady = 20) # Coloca o frame na janela com um espaçamento vertical

    # Cria um canvas para exibir os horários
    largura_canvas = 950  # Aumente a largura do canvas conforme necessário
    canvas = tk.Canvas(frame_reservas, bg="#FFFFFF", width=largura_canvas)
    canvas.pack(side="left", fill="both", expand=True)
    
    
    # Cria uma barra de rolagem vertical
    """
    A barra de rolagem é criada utilizando a biblioteca ttk com a classe ttk.Scrollbar. 
    É passado o parâmetro frame_reservas como o pai (parent) da barra de rolagem.

    O parâmetro orient="vertical" indica que a barra de rolagem deve ser vertical.

    O parâmetro command=canvas.yview especifica que a barra de rolagem deve controlar 
    o eixo y (vertical) do canvas.

    Por fim, a barra de rolagem é colocada no lado direito (side="right") do frame de 
    reservas e preenchida verticalmente (fill="y"). Isso faz com que a barra de rolagem 
    seja exibida ao lado direito do canvas e se ajuste verticalmente para ocupar o espaço 
    disponível.
    """
    scrollbar = ttk.Scrollbar(frame_reservas, orient="vertical", command=canvas.yview)
    scrollbar.pack(side="right", fill="y")
    
        
    # Configura a barra de rolagem para controlar o canvas
    """
    Essa linha configura o comando de rolagem vertical do canvas para a barra de rolagem 
    que foi criada. O método configure é usado para definir opções de configuração para um 
    widget. No caso do canvas, estamos configurando a opção yscrollcommand para ser igual ao 
    método set da barra de rolagem.

    O yscrollcommand é usado para associar a barra de rolagem vertical ao canvas, de modo 
    que a barra de rolagem seja atualizada quando o canvas é rolado. O método set é chamado 
    pela barra de rolagem para atualizar sua posição com base no comando de rolagem recebido.
    """
    canvas.configure(yscrollcommand=scrollbar.set)
    
    """
    Essa linha faz o bind (vinculação) do evento <Configure> com uma função lambda 
    que atualiza a região de rolagem (scrollregion) do canvas quando ocorrem alterações de 
    tamanho.

    O evento <Configure> é acionado quando ocorrem alterações no tamanho do widget, 
    como redimensionamento da janela ou mudanças no conteúdo do canvas.

    Dentro da função lambda, é chamado o método configure do canvas para atualizar a 
    scrollregion com base na caixa delimitadora (bbox) do canvas. A caixa delimitadora "all" 
    retorna as coordenadas que englobam todos os itens desenhados no canvas.

    Essa atualização da região de rolagem garante que a barra de rolagem funcione 
    corretamente ao ajustar o tamanho e o conteúdo do canvas, permitindo rolar adequadamente 
    toda a área do canvas.
    """
    canvas.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    
    #Cria um frame dentro do canvas para conter os horários
    frame_horarios = tk.Frame(canvas, bg="#FFFFFF")
    
    """
    O método create_window é usado para criar uma janela no canvas onde outros 
    widgets podem ser colocados. Ele recebe vários parâmetros:

    (0, 0): Esses valores definem a posição inicial da janela dentro do canvas. 
    Nesse caso, (0, 0) indica que a janela será posicionada no canto superior 
    esquerdo do canvas.
    
    window=frame_horarios: O parâmetro window especifica o widget que será inserido 
    na janela. Aqui, o frame_horarios é o widget que será colocado na janela criada.
    
    anchor="nw": O parâmetro anchor especifica o ponto de ancoragem para a janela. Nesse 
    caso, "nw" indica que a janela será ancorada no canto superior esquerdo, o que significa 
    que o widget frame_horarios será posicionado no canto superior esquerdo da janela.
    """
    canvas.create_window((0, 0), window=frame_horarios, anchor="nw")   
    
    def incrementar_horario(horario):
        
        # Função auxiliar para incrementar o horário em 30 minutos
        dt = datetime.strptime(horario, "%H:%M")
        dt += timedelta(minutes=30)
        return dt.strftime("%H:%M")
    
    def preencher_horarios(reservas=None):
        
        #Limpa os horários existes
        for child in frame_horarios.winfo_children():
            
            child.destroy()
            
        #Obter todos os horários disponiveis a partir das 09:00 às 23:00
        horario = "09:00"
        
        row = 0
        row_2 = 0
        
        while horario <= "23:00":
            
            #Inicializa a variável reservado como False antes de verificar se o horário está reservado
            reservado = False
            
            #Verifica se o horário está reservado
            #Verica se a lista de reservas não está vazia
            if reservas is not None:
                
                # Verificar se o horário está reservado
                for reserva in reservas:  # Itera sobre cada reserva na lista de reservas
                    
                    hora_inicio = str(reserva["Hora Inicio"])  # Obtém a hora de início da reserva atual e a converte para string
                    hora_fim = str(reserva["Hora Fim"])  # Obtém a hora de término da reserva atual e a converte para string
                    if hora_inicio <= horario < hora_fim:  # Verifica se o horário atual está dentro do intervalo da reserva
                        reservado = True  # Define a variável reservado como True se o horário estiver reservado
                        break  # Interrompe o loop for, pois não é necessário verificar as reservas restantes
                        
            
            #Coluna 1
            if "16:00" >= horario >= "09:00":
                
                
                    if reservado:
                        # Criação de um label para horário reservado
                        label_status = tk.Label(frame_horarios, text="das {} às {}".format(horario, incrementar_horario(horario)),
                                            font=("Arial 16"), bg="#FFFF00", width=20, padx=50)
                    else:
                        # Criação de um label para horário reservado
                        label_status = tk.Label(frame_horarios, text="das {} às {}".format(horario, incrementar_horario(horario)),
                                            font=("Arial 16"), bg="#00FF00", width=20, padx=50)
                        
                    #Posicionamento dos horários nas colunas 0 e 1
                    if horario <= "16:00":
                        label_status.grid(row=row, column=0, pady=5, sticky="w", padx=50)
                    else:
                        label_status.grid(row=row, column=1, pady=5, sticky="w", padx=50)

                    row += 1
                 
            #Coluna 2
            elif "23:00" >= horario >= "16:00":
                
                    if reservado:

                        # Criação de um label para horário reservado
                        label_status = tk.Label(frame_horarios, text="das {} às {}".format(horario, incrementar_horario(horario)),
                                                font=("Arial 16"), bg="#FFFF00", width=20, padx=50)
                    else:
                        # Criação de um label para horário reservado
                        label_status = tk.Label(frame_horarios, text="das {} às {}".format(horario, incrementar_horario(horario)),
                                                font=("Arial 16"), bg="#00FF00", width=20, padx=50)

                    #Posicionar os horários nas colunas 2 e 3
                    if horario <= "23:00":
                        label_status.grid(row=row_2, column=2, pady=5, sticky="w", padx=50)
                    else:
                        label_status.grid(row=row_2, column=3, pady=5, sticky="w", padx=50)

                    row_2 += 1
                
            #Incrementar o horário em 30 minutos
            horario = incrementar_horario(horario)
            
        # Atualiza o tamanho da área de visualização do canvas
        #Atualiza as tarefas pendentes do canvas e garante que o canvas esteja pronto 
        #para ser exibido corretamente.
        canvas.update_idletasks()
        
        #Configura a área de rolagem do canvas (scrollregion) para abranger 
        #toda a área contida no canvas. Isso permite que a barra de rolagem funcione 
        #adequadamente e abranja todos os elementos dentro do canvas.
        canvas.configure(scrollregion=canvas.bbox("all"))
        
    def filtrar_reservas():
        
        """
        Filtra as reservas com base em uma data específica e preenche os horários filtrados.
        """

        data_filtrada = entry_data.get()  # Obtém a data filtrada do campo de entrada (entry_data)

        # Filtrar as reservas com base na data
        df = pd.read_excel("C:/Users/55119/Desktop/Reserva Quadra/Dados.xlsx")  # Lê o arquivo Excel contendo as reservas
        reservas_filtradas = df[df["Data"] == data_filtrada]  # Filtra as reservas com base na coluna "Data" igual à data filtrada

        reservas = []  # Lista vazia para armazenar as reservas filtradas
        for _, row in reservas_filtradas.iterrows():
            # Itera sobre as linhas das reservas filtradas
            hora_inicio = row["Hora Inicio"]  # Obtém a hora de início da reserva atual
            hora_fim = row["Hora Fim"]  # Obtém a hora de fim da reserva atual
            reservas.append({"Hora Inicio": hora_inicio, "Hora Fim": hora_fim})  # Adiciona a reserva atual à lista de reservas

        preencher_horarios(reservas)  # Chama a função preencher_horarios passando as reservas filtradas como argumento
        
    #Cria um frame para conter os widgets de filtragem
    frame_filtragem = tk.Frame(janela_ver_reservas, bg="#FFFFFF")
    frame_filtragem.pack(pady=10)
    
    label_data = tk.Label(frame_filtragem,
                         text="Data",
                         font="Arial 16",
                         bg="#FFFFFF")
    label_data.pack(side="left")
    
    entry_data = tk.Entry(frame_filtragem,
                         font="Arial 16")
    entry_data.pack(side="left")
    
    #Preeche o campo de entry_data com a data atual do computador
    data_atual = date.today().strftime("%d/%m/%Y")
    entry_data.insert(tk.END, data_atual)
    
    button_filtrar = tk.Button(frame_filtragem,
                         text="Filtrar",
                         font="Arial 16",
                         command = filtrar_reservas)
    button_filtrar.pack(side="left", padx=10)
    
    button_reservar = tk.Button(frame_filtragem,
                         text="Reservas",
                         font="Arial 16",
                         command= lambda: fazer_reserva(entry_data.get()) )
    button_reservar.pack(side="left", padx=10)
    
    #Preenche os horários iniciais
    #preencher_horarios()
    
    #Filtra a reserva
    filtrar_reservas()
    
    #-------------------------------
    
    #Centraliza a janela
    janela_ver_reservas.update_idletasks()

    # Obtém a largura atual da janela
    largura = janela_ver_reservas.winfo_width()

    # Obtém a altura atual da janela
    altura = janela_ver_reservas.winfo_height()

    # Calcula a posição x para centralizar a janela na tela
    posicao_x = (janela_ver_reservas.winfo_screenwidth() // 2) - (largura // 2)

    # Calcula a posição y para centralizar a janela na tela
    posicao_y = (janela_ver_reservas.winfo_screenheight() // 2) - (altura // 2)

    # Define a geometria da janela com a largura, altura, posição x e posição y
    janela_ver_reservas.geometry('{}x{}+{}+{}'.format(largura, altura, posicao_x, posicao_y))

    janela_ver_reservas.mainloop()
    
    
def detalhe_reservas():
    
    janela_detalhe_reservas = tk.Toplevel()  # Cria uma nova janela de detalhes de reservas
    janela_detalhe_reservas.title("Detalhe de Reservas")  # Define o título da janela
    janela_detalhe_reservas.configure(bg="#FFFFFF")  # Define a cor de fundo da janela como branco

    frame_entries = tk.Frame(janela_detalhe_reservas, 
                             bg="#FFFFFF")  # Cria um frame para os campos Entry
    frame_entries.pack(pady=10)  # Empacota o frame e adiciona um espaçamento vertical de 10 pixels

    campos = ["Data", "Hora Início", "Hora Fim", "CPF", "Nome", "Preço"]  # Lista de campos a serem exibidos
    entries = {}  # Dicionário para armazenar os campos Entry
    
    for i, campo in enumerate(campos):
        
        """
        row=i // 3 calcula a linha em que o frame deve ser posicionado. 
        O operador // realiza a divisão inteira de i por 3, resultando em 
        um número inteiro que representa a linha. Isso significa que a 
        cada 3 campos, uma nova linha será iniciada.
        
        column=i % 3 calcula a coluna em que o frame deve ser posicionado. O 
        operador % realiza o cálculo do resto da divisão de i por 3, resultando em 
        um número inteiro que representa a coluna. Isso garante que o campo seja 
        posicionado em uma das três colunas disponíveis.
        
        Os parâmetros padx=10 e pady=5 definem o espaçamento interno (em pixels) 
        horizontal e vertical, respectivamente, para o frame atual dentro do 
        frame frame_entries. Esses valores adicionam um espaçamento de 10 pixels 
        à esquerda e à direita, e um espaçamento de 5 pixels acima e abaixo do frame.
        """
        frame = tk.Frame(frame_entries, bg="#FFFFFF")  # Cria um frame para o campo Entry atual
        frame.grid(row=i // 3, column=i % 3, padx=10, pady=5)  # Posiciona o frame dentro do frame_entries
    
        label = tk.Label(frame, text=campo + ":", font=("Arial", 14), bg="#FFFFFF")  # Cria um label para o campo
        label.pack(side="left")  # Posiciona o label à esquerda dentro do frame

        entry = tk.Entry(frame, font=("Arial", 14))  # Cria um campo Entry para o usuário inserir dados
        entry.pack(side="left", padx=10)  # Posiciona o campo Entry à esquerda com um espaçamento de 10 pixels
        entries[campo] = entry  # Adiciona o campo Entry ao dicionário entries com o nome do campo como chave
        
    frame_soma = tk.Frame(janela_detalhe_reservas, bg="#FFFFFF")  # Cria um frame para a soma da coluna Preço
    frame_soma.pack(pady=10)  # Empacota o frame e adiciona um espaçamento vertical de 10 pixels

    label_soma = tk.Label(frame_soma, text="Soma da coluna Preço:", font=("Arial", 14), bg="#FFFFFF")  # Cria um label para a soma
    label_soma.pack(side="left", padx=10)  # Posiciona o label à esquerda com um espaçamento de 10 pixels

    label_valor_soma = tk.Label(frame_soma, text="0.00", font=("Arial", 14), bg="#FFFFFF")  # Cria um label para exibir o valor da soma
    label_valor_soma.pack(side="left")  # Posiciona o label à esquerda

    label_status = tk.Label(frame_soma, text="", font=("Arial", 14), bg="#FFFFFF")  # Cria um label para exibir o status
    label_status.pack()  # Empacota o label

        
    frame_opcoes = tk.Frame(janela_detalhe_reservas, bg="#FFFFFF")  # Cria um frame
    frame_opcoes.pack(pady=10)  # Empacota o frame e adiciona um espaçamento vertical de 10 pixels
    
    # Cria os campos Entry para as datas do filtro
    frame_datas = tk.Frame(janela_detalhe_reservas, bg="#FFFFFF")
    frame_datas.pack(pady=10)

    label_data_inicio = tk.Label(frame_datas, text="Data Início:", font=("Arial", 14), bg="#FFFFFF")
    label_data_inicio.pack(side="left", padx=10)

    entry_data_inicio = tk.Entry(frame_datas, font=("Arial", 14))
    entry_data_inicio.pack(side="left", padx=10)

    label_data_fim = tk.Label(frame_datas, text="Data Fim:", font=("Arial", 14), bg="#FFFFFF")
    label_data_fim.pack(side="left", padx=10)

    entry_data_fim = tk.Entry(frame_datas, font=("Arial", 14))
    entry_data_fim.pack(side="left", padx=10)
    
    def filtrar():
        
        print("...")
    
    botao_filtrar = tk.Button(janela_detalhe_reservas,
                               text="Filtrar", 
                               font=("Arial", 14),
                               command = filtrar)
    botao_filtrar.pack(padx=10)
    
    # Cria o Treeview com o layout melhorado
    tree = ttk.Treeview(janela_detalhe_reservas, show="headings")
    
    """
    tree é um objeto treeview, que é uma estrutura de dados usada para exibir 
    informações em forma de tabela com colunas.

    "columns" é uma propriedade do treeview tree que define as colunas a serem exibidas.

    tuple(campos) cria uma tupla contendo os elementos da lista campos. A função 
    tuple() é usada para converter a lista em uma tupla.
    """
    tree["columns"] = tuple(campos)
    
    # Define os cabeçalhos das colunas
    for campo in campos:
        
        tree.heading(campo, text=campo, anchor=tk.CENTER)
        tree.column(campo, anchor=tk.CENTER, width=100)
        
    #Obtém o caminho do arquivo
    caminho = r"C:\Users\55119\Desktop\Reserva Quadra\Dados.xlsx"
    planilha = "Reservas"
    
    try:
        
        workbook = openpyxl.load_workbook(caminho)  # Carrega o arquivo Excel
        sheet = workbook[planilha]  # Obtém a planilha desejada do arquivo
        
    except FileNotFoundError:
        
        messagebox.showerror("Arquivo não encontrado", "O arquivo Dados.xlsx não foi encontrado.")  # Exibe uma mensagem de erro indicando que o arquivo não foi encontrado
        janela_detalhe_reservas.destroy()  # Fecha a janela janela_detalhe_reservas
        return
        
    dados = []  #Lista para armazenar os dados do arquivo
    for row in sheet.iter_rows(values_only=True):  # Itera sobre as linhas da planilha
        dados.append(row)  # Adiciona a linha atual à lista dados
        
    # Adiciona a treeview à janela usando o scrollbars
    """
    As duas linhas estão relacionadas à criação e configuração de uma barra de 
    rolagem vertical para o widget treeview tree dentro da janela janela_detalhe_reservas.

    scrollbar_y = ttk.Scrollbar(janela_detalhe_reservas, orient=tk.VERTICAL, 
    command=tree.yview) cria um objeto Scrollbar vertical do pacote ttk associado à 
    janela janela_detalhe_reservas. A propriedade orient é definida como tk.VERTICAL 
    para criar uma barra de rolagem vertical. O argumento command é configurado como 
    tree.yview para vincular a barra de rolagem à visualização vertical do widget 
    treeview tree.

    scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y) posiciona a barra de rolagem vertical à 
    direita do widget tree. O argumento side é definido como tk.RIGHT para posicionar a 
    barra de rolagem à direita. O argumento fill é definido como tk.Y para permitir 
    que a barra de rolagem preencha a altura do widget treeview tree, para que ela possa 
    ser rolada verticalmente.
    """
    scrollbar_y = ttk.Scrollbar(janela_detalhe_reservas, orient=tk.VERTICAL, command=tree.yview)
    scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
        
        
    """
    tree.configure(yscrollcommand=scrollbar_y.set) configura a propriedade 
    yscrollcommand do widget treeview tree para utilizar o método set() da barra de 
    rolagem vertical scrollbar_y. Isso permite que a barra de rolagem controle a 
    visualização vertical do conteúdo no widget treeview.

    tree.pack(pady=10, expand=True, fill=tk.BOTH) posiciona o widget treeview tree 
    dentro da janela janela_detalhe_reservas. O argumento pady=10 define um espaçamento 
    de 10 pixels entre o widget e suas bordas. O argumento expand=True permite que o 
    widget se expanda para preencher qualquer espaço extra disponível na janela. O 
    argumento fill=tk.BOTH faz com que o widget preencha tanto na direção vertical quanto 
    horizontal, ocupando todo o espaço disponível na janela.
    """
    tree.configure(yscroll=scrollbar_y.set)
    tree.pack(pady=10, expand=True, fill=tk.BOTH)

    # Insere os dados na treeview
    for linha in dados[1:]:
        
        tree.insert("", tk.END, values=linha)
        
    
    
    
    
        
janela_inicial = tk.Tk()  # Cria uma instância da classe Tk para a janela inicial
janela_inicial.title("Menu")  # Define o título da janela
janela_inicial.configure(bg="#FFFFFF")  # Define a cor de fundo da janela como branco

fonte_titulo = ("Arial", 24, "bold")  # Define a fonte para o título
fonte_botoes = ("Arial", 20)  # Define a fonte para os botões

label_titulo = tk.Label(janela_inicial,
                       text = "Sistema de Reservas",
                       font = fonte_titulo,
                       bg = "#FFFFFF")
label_titulo.pack(pady = 20)

# Cria um frame para conter os botões
frame_botoes = tk.Frame(janela_inicial,
                       bg = "#FFFFFF")
frame_botoes.pack(pady = 20) # Coloca o frame na janela com um espaçamento vertical

# Obtém a data atual no formato "dd/mm/aaaa"
data_atual = date.today().strftime("%d/%m/%Y")

btn_fazer_reserva = tk.Button(frame_botoes,
                             text = "Fazer Reserva",
                             font = "Arial 20",
                             command = lambda: fazer_reserva(data_atual),
                             width = 30)
btn_fazer_reserva.pack(pady = 10) # Coloca o frame na janela com um espaçamento vertical



btn_ver_reserva = tk.Button(frame_botoes,
                             text = "Ver Reserva",
                             font = "Arial 20",
                             command = ver_reservas,
                             width = 30)
btn_ver_reserva.pack(pady = 10) # Coloca o frame na janela com um espaçamento vertical


btn_detalhe_reserva = tk.Button(frame_botoes,
                             text = "Detalhe Reservas",
                             font = "Arial 20",
                             command = detalhe_reservas,
                             width = 30)
btn_detalhe_reserva.pack(pady = 10) # Coloca o frame na janela com um espaçamento vertical



btn_sair = tk.Button(frame_botoes,
                             text = "Sair",
                             font = "Arial 20",
                             command = janela_inicial.destroy,
                             width = 30)
btn_sair.pack(pady = 10) # Coloca o frame na janela com um espaçamento vertical


#-------------------------------
    
#Centraliza a janela
janela_inicial.update_idletasks()
    
# Obtém a largura atual da janela
largura = janela_inicial.winfo_width()

# Obtém a altura atual da janela
altura = janela_inicial.winfo_height()

# Calcula a posição x para centralizar a janela na tela
posicao_x = (janela_inicial.winfo_screenwidth() // 2) - (largura // 2)

# Calcula a posição y para centralizar a janela na tela
posicao_y = (janela_inicial.winfo_screenheight() // 2) - (altura // 2)

# Define a geometria da janela com a largura, altura, posição x e posição y
janela_inicial.geometry('{}x{}+{}+{}'.format(largura, altura, posicao_x, posicao_y))


janela_inicial.mainloop()

