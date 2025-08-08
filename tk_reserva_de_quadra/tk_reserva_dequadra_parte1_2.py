import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
from datetime import date, datetime, timedelta
import openpyxl
from openpyxl import Workbook

import os
import pandas as pd

#Lista que armazena as reservas
reservas = []

#C:\Users\55119\Desktop\Reserva Quadra

def verificar_conflito_reserva(data, hora_inicial, hora_final):
    
    #Caminho para o arquivo do Excel que contém as reservas
    caminho = r"C:\Users\55119\Desktop\Reserva Quadra\Dados.xlsx"
    
    #Nome da planilha que contém as reservas
    planilha = "Reservas"
    
    try:
        
        #Carrega o arquivo do Excel
        workbook = openpyxl.load_workbook(caminho)
        
        #Obtém a planilha especifica
        sheet = workbook[planilha]
    
    except FileNotFoundError:
        
        return False  # Se o arquivo não existir, não há conflitos
    
    # Itera sobre as linhas da planilha, a partir da segunda linha (min_row=2)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        
        # Obtém os valores da data, hora inicial e hora final da reserva atual
        reserva_data = row[0]
        reserva_hora_inicial = row[1]
        reserva_hora_final = row[2]

        # Verifica se a data da nova reserva é igual à data da reserva atual
        # e se há sobreposição de horários entre a nova reserva e a reserva atual
        if data == reserva_data and hora_inicial < reserva_hora_final and hora_final > reserva_hora_inicial:
            
            return True  # Há um conflito de horários
    
    return False  # Não há conflito de horários

def fazer_reserva(data):
    
    # Cria uma nova janela usando Tkinter
    janela_fazer_reserva = tk.Toplevel()
    janela_fazer_reserva.title("Fazer Reserva")
    janela_fazer_reserva.configure(bg = "#FFFFFF")
    
    # Cria um frame para o título e o coloca na janela
    frame_titulo = tk.Frame(janela_fazer_reserva, bg = "#FFFFFF")
    frame_titulo.pack(pady = 20)
    
    label_titulo = tk.Label(frame_titulo, 
                            text="Fazer Reserva",
                            font = "Arial 24 bold", 
                            bg = "#FFFFFF")
    label_titulo.pack()
    
    # Cria um frame para o formulario e o coloca na janela
    frame_formulario = tk.Frame(janela_fazer_reserva, bg = "#FFFFFF")
    frame_formulario.pack(pady = 20)
    
    
    # Cria um label para a data, configura sua fonte e o coloca no frame do formulário
    label_data = tk.Label(frame_formulario, 
                            text="Digite a data da reserva (dd/mm/aaaa)",
                            font = "Arial 18", 
                            bg = "#FFFFFF")
    label_data.grid(row = 0, column = 0, pady = 10, stick = "e")
    
    # Cria uma entrada de texto para a data, configura sua fonte e a coloca no frame do formulário
    entry_data = tk.Entry(frame_formulario,
                            font = "Arial 18")
    entry_data.grid(row = 0, column = 1, pady = 10, padx = 10)
    entry_data.insert(0, data)
    
    #-----------------------------------------------------------------
    
    # Cria um label para a data, configura sua fonte e o coloca no frame do formulário
    label_inicial = tk.Label(frame_formulario, 
                            text="Digite a hora inicial da reserva (entre 09:00 e 23:00)",
                            font = "Arial 18", 
                            bg = "#FFFFFF")
    label_inicial.grid(row = 1, column = 0, pady = 10, stick = "e")
    
    # Cria uma entrada de texto para a data, configura sua fonte e a coloca no frame do formulário
    entry_inicial = tk.Entry(frame_formulario,
                            font = "Arial 18")
    entry_inicial.grid(row = 1, column = 1, pady = 10, padx = 10)
    
    #-----------------------------------------------------------------
    
    # Cria um label para a data, configura sua fonte e o coloca no frame do formulário
    label_final = tk.Label(frame_formulario, 
                            text="Digite a hora final da reserva (entre 09:00 e 23:00)",
                            font = "Arial 18", 
                            bg = "#FFFFFF")
    label_final.grid(row = 2, column = 0, pady = 10, stick = "e")
    
    # Cria uma entrada de texto para a data, configura sua fonte e a coloca no frame do formulário
    entry_final = tk.Entry(frame_formulario,
                            font = "Arial 18")
    entry_final.grid(row = 2, column = 1, pady = 10, padx = 10)
    
    #-----------------------------------------------------------------
    
    # Cria um label para a data, configura sua fonte e o coloca no frame do formulário
    label_cpf = tk.Label(frame_formulario, 
                            text="Digite o CPF",
                            font = "Arial 18", 
                            bg = "#FFFFFF")
    label_cpf.grid(row = 3, column = 0, pady = 10, stick = "e")
    
    # Cria uma entrada de texto para a data, configura sua fonte e a coloca no frame do formulário
    entry_cpf = tk.Entry(frame_formulario,
                            font = "Arial 18")
    entry_cpf.grid(row = 3, column = 1, pady = 10, padx = 10)
    
    #-----------------------------------------------------------------
    
    # Cria um label para a data, configura sua fonte e o coloca no frame do formulário
    label_nome = tk.Label(frame_formulario, 
                            text="Digite o Nome",
                            font = "Arial 18", 
                            bg = "#FFFFFF")
    label_nome.grid(row = 4, column = 0, pady = 10, stick = "e")
    
    # Cria uma entrada de texto para a data, configura sua fonte e a coloca no frame do formulário
    entry_nome = tk.Entry(frame_formulario,
                            font = "Arial 18")
    entry_nome.grid(row = 4, column = 1, pady = 10, padx = 10)
    
    
    # Cria uma entrada de texto para a data, configura sua fonte e a coloca no frame do formulário
    entry_valor = tk.Entry(frame_formulario,
                            font = "Arial 18")
    entry_valor.grid(row = 5, column = 1, pady = 10, padx = 10)
    
    
    
    btn_calcular_valor = tk.Button(frame_formulario,
                            text = "Calcular Valor",
                            font = "Arial 20")
    btn_calcular_valor.grid(row = 5, column = 0, pady = 10, padx = 10, stick = "e")
    
    
    def fazer_reserva_click():
        
        #Obtém os valores das entradas de texto
        data_da_reserva = entry_data.get()
        hora_inicial = entry_inicial.get()
        hora_final = entry_final.get()
        cpf = entry_cpf.get()
        nome = entry_nome.get()
        valor_total = float(entry_valor.get())

        # Verifica se o horário é inválido (fora dos limites permitidos ou com hora inicial maior ou igual à hora final)
        if hora_inicial < "09:00" or hora_final > "23:00" or hora_inicial >= hora_final:
            
            tk.messagebox.showerror("Horário inválido", "Horário inválido. Tente novamente.")
            
            return

        # Verifica se há conflito de horários com reservas existentes
        if verificar_conflito_reserva(data_da_reserva, hora_inicial, hora_final):
            
            tk.messagebox.showerror("Conflito de horários", "Já existe uma reserva neste horário. Escolha outro horário.")
            
            return
        
        # Caminho do arquivo do Excel e nome da planilha
        caminho = r"C:\Users\55119\Desktop\Reserva Quadra\Dados.xlsx"
        planilha = "Reservas"
        
        try:
            
            # Tenta carregar o arquivo do Excel e obter a planilha
            workbook = openpyxl.load_workbook(caminho)
            sheet = workbook[planilha]
            
            
        except FileNotFoundError:
            
            print("Arquivo não encontrado!")
            
        #Adiciona os dados da reserva na planilha
        sheet.append([data_da_reserva, hora_inicial, hora_final, cpf, nome, valor_total])
        workbook.save(caminho)
        
        #Exibe uma mensagem de sucesso com os detalhes da reserva
        mensagem = f"\nA reserva das {hora_inicial} às {hora_final} na data {data_da_reserva} foi feita com sucesso."
        mensagem += "\nO valor total da reserva é de R${:.2f}\n".format(valor_total)
        
        tk.messagebox.showinfo("Reserva feita com sucesso", mensagem)
        
        #Fecha a janela de reservas
        janela_fazer_reserva.destroy()
            
    
    # Cria um botão "Confirmar" que chama a função fazer_reserva_click quando clicado
    btn_confirmar = tk.Button(janela_fazer_reserva,
                            text = "Confirmar",
                            font = "Arial 20",
                            width = 50,
                            command = fazer_reserva_click)
    btn_confirmar.pack(pady = 20)
    
    #Centraliza a janela
    janela_fazer_reserva.update_idletasks()

    # Obtém a largura atual da janela
    largura = janela_fazer_reserva.winfo_width()

    # Obtém a altura atual da janela
    altura = janela_fazer_reserva.winfo_height()

    # Calcula a posição x para centralizar a janela na tela
    posicao_x = (janela_fazer_reserva.winfo_screenwidth() // 2) - (largura // 2)

    # Calcula a posição y para centralizar a janela na tela
    posicao_y = (janela_fazer_reserva.winfo_screenheight() // 2) - (altura // 2)

    # Define a geometria da janela com a largura, altura, posição x e posição y
    janela_fazer_reserva.geometry('{}x{}+{}+{}'.format(largura, altura, posicao_x, posicao_y))
    
    
    
    

janela_inicial = tk.Tk()  # Cria uma instância da classe Tk para a janela inicial
janela_inicial.title("Menu")  # Define o título da janela
janela_inicial.configure(bg="#FFFFFF")  # Define a cor de fundo da janela como branco

fonte_titulo = ("Arial", 24, "bold")  # Define a fonte para o título
fonte_botoes = ("Arial", 20)  # Define a fonte para os botões

label_titulo = tk.Label(janela_inicial,
                       text = "Sistema de Reservas",
                       font = fonte_titulo,
                       bg = "#FFFFFF")
label_titulo.pack(pady = 20)

# Cria um frame para conter os botões
frame_botoes = tk.Frame(janela_inicial,
                       bg = "#FFFFFF")
frame_botoes.pack(pady = 20) # Coloca o frame na janela com um espaçamento vertical

# Obtém a data atual no formato "dd/mm/aaaa"
data_atual = date.today().strftime("%d/%m/%Y")

btn_fazer_reserva = tk.Button(frame_botoes,
                             text = "Fazer Reserva",
                             font = "Arial 20",
                             command = lambda: fazer_reserva(data_atual),
                             width = 30)
btn_fazer_reserva.pack(pady = 10) # Coloca o frame na janela com um espaçamento vertical



btn_ver_reserva = tk.Button(frame_botoes,
                             text = "Ver Reserva",
                             font = "Arial 20",
                             command = lambda: fazer_reserva(data_atual),
                             width = 30)
btn_ver_reserva.pack(pady = 10) # Coloca o frame na janela com um espaçamento vertical


btn_cancelar_reserva = tk.Button(frame_botoes,
                             text = "Detalhe Reservas",
                             font = "Arial 20",
                             command = lambda: fazer_reserva(data_atual),
                             width = 30)
btn_cancelar_reserva.pack(pady = 10) # Coloca o frame na janela com um espaçamento vertical



btn_sair = tk.Button(frame_botoes,
                             text = "Sair",
                             font = "Arial 20",
                             command = janela_inicial.destroy,
                             width = 30)
btn_sair.pack(pady = 10) # Coloca o frame na janela com um espaçamento vertical


#-------------------------------
    
#Centraliza a janela
janela_inicial.update_idletasks()
    
# Obtém a largura atual da janela
largura = janela_inicial.winfo_width()

# Obtém a altura atual da janela
altura = janela_inicial.winfo_height()

# Calcula a posição x para centralizar a janela na tela
posicao_x = (janela_inicial.winfo_screenwidth() // 2) - (largura // 2)

# Calcula a posição y para centralizar a janela na tela
posicao_y = (janela_inicial.winfo_screenheight() // 2) - (altura // 2)

# Define a geometria da janela com a largura, altura, posição x e posição y
janela_inicial.geometry('{}x{}+{}+{}'.format(largura, altura, posicao_x, posicao_y))


janela_inicial.mainloop()

