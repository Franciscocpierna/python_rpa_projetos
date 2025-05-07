from openpyxl import load_workbook
import os

#Caminho do arquivo
caminho_nome_arquivo = "C:\\python_projetos\\python_rpa_projetos\\pythonexcel_openpyxl\\Vendedores.xlsx"
planilha_aberta = load_workbook(filename=caminho_nome_arquivo)

#Seleciona a sheet de Vendas
sheet_selecionada = planilha_aberta['Vendas']

somarAmandaMartins = 0
somarElianeMoreira = 0
somarLeonardoAlmeida = 0
somarNicolasPereira = 0

#Inserindo a somatoria da venda  de cada vendedor
for linha in range(2, len(sheet_selecionada['A']) + 1):
    
    if sheet_selecionada['A%s' % linha].value == "Amanda Martins":
        somarAmandaMartins = somarAmandaMartins + sheet_selecionada['C%s' % linha].value
        
    elif sheet_selecionada['A%s' % linha].value == "Eliane Moreira":
        somarElianeMoreira = somarElianeMoreira + sheet_selecionada['C%s' % linha].value
        
    elif sheet_selecionada['A%s' % linha].value == "Nicolas Pereira":
        somarNicolasPereira = somarNicolasPereira + sheet_selecionada['C%s' % linha].value
        
    elif sheet_selecionada['A%s' % linha].value == "Leonardo Almeida":
        somarLeonardoAlmeida = somarLeonardoAlmeida + sheet_selecionada['C%s' % linha].value
        

#Inserindo uma nova sheet de resumo
sheet_resumo = planilha_aberta.create_sheet(title="Resumo")

sheet_resumo['A1'] = "Vendedores"
sheet_resumo['B1'] = "Vendas"

#Preenche dados do primeiro vendedor
sheet_resumo['A2'] = "Amanda Martins"
sheet_resumo['B2'] = somarAmandaMartins

#Preenche dados do segundo vendedor
sheet_resumo['A3'] = "Eliane Moreira"
sheet_resumo['B3'] = somarElianeMoreira

#Preenche dados do terceiro vendedor
sheet_resumo['A4'] = "Leonardo Almeida"
sheet_resumo['B4'] = somarLeonardoAlmeida

#Preenche dados do quarto vendedor
sheet_resumo['A5'] = "Nicolas Pereira"
sheet_resumo['B5'] = somarNicolasPereira




#Salva a planilha com as alterações
planilha_aberta.save(filename=caminho_nome_arquivo)

#Abre a planilha
os.startfile(caminho_nome_arquivo)