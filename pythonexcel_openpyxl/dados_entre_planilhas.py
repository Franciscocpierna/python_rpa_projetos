from openpyxl import load_workbook
from openpyxl import Workbook
import os


caminhoArquivoDadosSistema = "C:\\python_projetos\\python_rpa_projetos\\pythonexcel_openpyxl\\DadosSistema.xlsx"
planilhaDadosSistema = load_workbook(filename=caminhoArquivoDadosSistema)

#Seleciona a sheet de Professor
sheetPlanilhaDadosSistema = planilhaDadosSistema['Dados']

#Novo arquivo
criandoNovoArquivoExcel = Workbook()
nova_planilha = criandoNovoArquivoExcel.active

for linha in range(1, len(sheetPlanilhaDadosSistema['A']) + 1):
    for coluna in range(1, 10):
        nova_planilha.cell(row=linha, column=coluna).value = sheetPlanilhaDadosSistema.cell(row=linha, column=coluna).value


#Deleta a linha 2
nova_planilha.delete_rows(2)

#Deleta a coluna 2 e 3
nova_planilha.delete_cols(2)
nova_planilha.delete_cols(2)
        
#Renomeia a Sheet
nova_planilha.title = 'Dados Funcionarios'

#Cria uma nova sheet e renomeia como Resumo
criandoNovoArquivoExcel.create_sheet('Resumo')

#Seleciona a Sheet de Resuno da planilha que criamos
selecionaSheetResumoNovaPlanilha = criandoNovoArquivoExcel['Resumo']

#Preenche os titulos
selecionaSheetResumoNovaPlanilha['A1'] = "Vendedor"
selecionaSheetResumoNovaPlanilha['B1'] = "Total Vendas"

#Preenche dados do primeiro vendedor
selecionaSheetResumoNovaPlanilha['A2'] = "Amanda Martins"
selecionaSheetResumoNovaPlanilha['B2'] = "=SUMIF('Dados Funcionarios'!A:C,A2,'Dados Funcionarios'!C:C)"

#Preenche dados do segundo vendedor
selecionaSheetResumoNovaPlanilha['A3'] = "Eliane Moreira"
selecionaSheetResumoNovaPlanilha['B3'] = "=SUMIF('Dados Funcionarios'!A:C,A3,'Dados Funcionarios'!C:C)"

#Preenche dados do terceiro vendedor
selecionaSheetResumoNovaPlanilha['A4'] = "Leonardo Almeida"
selecionaSheetResumoNovaPlanilha['B4'] = "=SUMIF('Dados Funcionarios'!A:C,A4,'Dados Funcionarios'!C:C)"

#Preenche dados do quarto vendedor
selecionaSheetResumoNovaPlanilha['A5'] = "Nicolas Pereira"
selecionaSheetResumoNovaPlanilha['B5'] = "=SUMIF('Dados Funcionarios'!A:C,A5,'Dados Funcionarios'!C:C)"

#Nova variavel para armazenar o caminho e dar um nome para o novo arquivo de Excel criado
caminhoNovaPlanilha = "C:\\python_projetos\\python_rpa_projetos\\pythonexcel_openpyxl\\RelatorioSistema.xlsx"

#Salva a planilha com as alterações
criandoNovoArquivoExcel.save(filename=caminhoNovaPlanilha)

#Abre a planilha
os.startfile(caminhoNovaPlanilha)