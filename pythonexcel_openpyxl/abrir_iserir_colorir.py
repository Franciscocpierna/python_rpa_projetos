from openpyxl import load_workbook
import os

from openpyxl.styles import Color, PatternFill, Font, Border, Side
from openpyxl.styles import colors
from openpyxl.cell import Cell

caminho_nome_arquivo = "C:\\python_projetos\\python_rpa_projetos\\pythonexcel_openpyxl\\InserirDadosPintarCelulas.xlsx"
planilha_aberta = load_workbook(filename=caminho_nome_arquivo)

#Seleciona a sheet de Professor
sheet_selecionada = planilha_aberta['Professor']




#Popula as informações que vão para a planilha
dadosTabela = [
    ['Nome', 'Idade'],
    ['Berenice', 28],
    ['Caio', 32],
    ['Nicole', 34],
    ['Leonardo', 19],
    ['Amanda', 25]
]

    
#O append pega toda a lista e passa para a sheet
for linhaPlanilha in dadosTabela:
    sheet_selecionada.append(linhaPlanilha)
    

corTitulo = PatternFill(start_color='00FFFF00',
                      end_color='00FFFF00',
                      fill_type='solid')


corCelulas = PatternFill(start_color='00CCFFCC',
                      end_color='00CCFFCC',
                      fill_type='solid')


sheet_selecionada["A1"].fill = corTitulo
sheet_selecionada["B1"].fill = corTitulo

for linha in range(2, len(sheet_selecionada['A']) + 1):
    
    celulaColunaA = "A" + str(linha)
    celulaColunaB = "B" + str(linha)
    
    sheet_selecionada[celulaColunaA].fill = corCelulas
    sheet_selecionada[celulaColunaB].fill = corCelulas


#Salva a planilha com as alterações
planilha_aberta.save(filename=caminho_nome_arquivo)

#Abre a planilha
os.startfile(caminho_nome_arquivo)