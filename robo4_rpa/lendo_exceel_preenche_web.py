# Importa a função 'load_workbook' do módulo 'openpyxl'.
# Esta função é usada para abrir arquivos do Excel e trabalhar com eles.
from openpyxl import load_workbook

# Importa o módulo 'webdriver' do pacote 'selenium' e o renomeia como 'opcoesSelenium'.
# O 'webdriver' é uma ferramenta essencial do Selenium para interagir com navegadores web.
from selenium import webdriver as opcoesSelenium

# Importa a classe 'By' do módulo 'selenium.webdriver.common.by'.
# 'By' é usado para definir os métodos de localização de elementos em uma página web.
from selenium.webdriver.common.by import By

# Importa 'WebDriverWait' do módulo 'selenium.webdriver.support.ui'.
# 'WebDriverWait' permite a implementação de esperas explícitas, uma
# forma de esperar por uma condição específica antes de prosseguir com a execução do código.
from selenium.webdriver.support.ui import WebDriverWait

# Importa 'expected_conditions' do módulo 'selenium.webdriver.support' e o renomeia como 'EC'.
# 'expected_conditions' fornece um conjunto de condições predefinidas para usar com 'WebDriverWait'.
from selenium.webdriver.support import expected_conditions as EC

# A variável 'nomeCaminhoArquivo' é definida como o nome do arquivo Excel que será utilizado.
nomeCaminhoArquivo = "DadosFormulario.xlsx"

# 'planilha_aberta' é uma variável que armazena o arquivo Excel aberto pela função 'load_workbook'.
# 'filename=nomeCaminhoArquivo' indica que a função deve abrir o arquivo especificado em 'nomeCaminhoArquivo'.
planilha_aberta = load_workbook(filename=nomeCaminhoArquivo)

# A variável 'sheet_selecionada' armazena a aba (sheet) de nome 'Dados' do arquivo Excel.
# Essa aba é onde os dados a serem utilizados estão localizados.
sheet_selecionada = planilha_aberta['Dados']


# Este loop 'for' itera sobre as linhas da planilha, começando da segunda linha (linha=2) 
# até a última linha com dados na coluna 'A'. O 'range(2, len(sheet_selecionada['A']) + 1)'
# garante que a primeira linha (geralmente um cabeçalho) seja ignorada.
for linha in range(2, len(sheet_selecionada['A']) + 1):

    # Esta linha está dentro de um loop 'for', que itera sobre as linhas da planilha Excel.
    # Para cada iteração do loop, 'linha' representa o número da linha atual.
    # Aqui, 'sheet_selecionada[f'A{linha}']' acessa a célula na coluna 'A' da linha atual.
    # Por exemplo, se 'linha' é 2, então 'sheet_selecionada['A2']' é a célula referenciada.
    # '.value' é usado para obter o conteúdo (valor) da célula referenciada.
    # O valor obtido é então atribuído à variável 'nome'.
    # Portanto, 'nome' contém o dado que está na coluna 'A' da linha atual da planilha.
    nome = sheet_selecionada[f'A{linha}'].value

    # Um processo semelhante é realizado para a coluna 'B', onde o e-mail está armazenado.
    # 'sheet_selecionada[f'B{linha}'].value' acessa o valor da célula na coluna 'B' e na linha atual.
    # Este valor é então armazenado na variável 'email'.
    # Cada vez que o loop 'for' avança para a próxima linha, um novo e-mail é lido e armazenado em 'email'.
    email = sheet_selecionada[f'B{linha}'].value

    # A mesma lógica é aplicada para obter o número de telefone da coluna 'C'.
    # 'sheet_selecionada[f'C{linha}'].value' acessa o valor da célula correspondente na linha atual.
    # O valor obtido (neste caso, o número de telefone) é armazenado na variável 'telefone'.
    telefone = sheet_selecionada[f'C{linha}'].value

    # Para a coluna 'D', que contém informações sobre o sexo, o processo é repetido.
    # 'sheet_selecionada[f'D{linha}'].value' lê o valor da célula na coluna 'D' (sexo) da linha atual.
    # Este valor (que pode ser, por exemplo, 'Masculino' ou 'Feminino') é então armazenado na variável 'sexo'.
    sexo = sheet_selecionada[f'D{linha}'].value

    # Finalmente, o valor da célula na coluna 'E' (que contém informações 'sobre' o indivíduo) é lido.
    # 'sheet_selecionada[f'E{linha}'].value' acessa o valor dessa célula específica.
    # O conteúdo dessa célula é atribuído à variável 'sobre', que pode conter um texto descritivo.
    sobre = sheet_selecionada[f'E{linha}'].value


    # Inicializa uma nova instância do navegador Chrome. 
    # Esta linha cria um novo navegador para cada linha de dados processada.
    navegadorFormulario = opcoesSelenium.Chrome()

    # Comanda o navegador para abrir a URL especificada.
    # Este comando carrega a página do formulário do SurveyMonkey.
    navegadorFormulario.get("https://pt.surveymonkey.com/r/WLXYDX2")

    # Cria um objeto 'WebDriverWait' chamado 'espera'.
    # Este objeto é configurado para esperar por até 10 segundos para que certas condições sejam atendidas.
    # 'navegadorFormulario' é o navegador onde as esperas serão aplicadas.
    espera = WebDriverWait(navegadorFormulario, 10)

    # Esta linha de código estabelece uma espera explícita. 
    # A função 'espera.until()' é usada aqui.
    # Ela instrui o Selenium a aguardar até que uma determinada condição 
    # seja verdadeira antes de prosseguir.
    # A condição especificada é 'EC.presence_of_element_located()'.
    # 'EC.presence_of_element_located()' espera até que o elemento 
    # especificado esteja presente na página da web.
    # O elemento é especificado usando 'By.NAME, "166517069"', que 
    # procura um elemento pelo seu atributo 'name'.
    # Uma vez que o elemento (campo 'Nome') é encontrado, ele é 
    # armazenado na variável 'campo_nome'.
    campo_nome = espera.until(EC.presence_of_element_located((By.NAME, "166517069")))

    # Após o campo 'Nome' ser localizado e armazenado em 'campo_nome', o método 'send_keys(nome)' é chamado.
    # 'send_keys(nome)' envia o valor armazenado na variável 'nome' (extraído da planilha Excel) para o campo 'Nome' no formulário web.
    campo_nome.send_keys(nome)

    # Processo similar ao anterior, mas aplicado ao campo 'Email'.
    # Aguarda até que o campo 'Email' (identificado pelo seu atributo 'name') esteja presente na página.
    # Após ser localizado, o valor de 'email' é enviado para este campo.
    campo_email = espera.until(EC.presence_of_element_located((By.NAME, "166517072")))
    campo_email.send_keys(email)

    # Espera pelo campo 'Telefone' da mesma maneira que os campos anteriores.
    # Uma vez localizado, o número de telefone extraído da planilha é inserido neste campo.
    campo_telefone = espera.until(EC.presence_of_element_located((By.NAME, "166517070")))
    campo_telefone.send_keys(telefone)

    # Espera pelo campo 'Sobre', e após sua localização, insere o texto correspondente.
    campo_sobre = espera.until(EC.presence_of_element_located((By.NAME, "166517073")))
    campo_sobre.send_keys(sobre)

    # Esta seção do código lida com a seleção de um botão de rádio
    # com base no valor da variável 'sexo'.
    # Se 'sexo' for igual a "Masculino", aguarda até que o botão 
    # de rádio correspondente (identificado por 'By.ID') esteja clicável.
    # Uma vez clicável, o botão é acionado com o método '.click()'.
    if sexo == "Masculino":
        #botao_masculino = espera.until(EC.element_to_be_clickable((By.ID, "166517071_1215509812_label")))
        botao_masculino = espera.until(EC.element_to_be_clickable((By.ID, "1215509812-label")))
        botao_masculino.click()
        #<span id="1215509812-label" class="smqr-richTextContent-0-2-25" translate="no">Masculino</span>
    # Se 'sexo' não for "Masculino" (ou seja, for "Feminino" ou outro valor), o botão de rádio "Feminino" é selecionado e clicado.
    else:
        #botao_feminino = espera.until(EC.element_to_be_clickable((By.ID, "166517071_1215509813_label")))
        botao_feminino = espera.until(EC.element_to_be_clickable((By.ID, "1215509813-label")))
        botao_feminino.click()

    # Finalmente, espera pelo botão de envio do formulário.
    # Uma vez que o botão (identificado por um XPath específico) esteja disponível para 
    # clique, ele é acionado.
    # Isso submete as informações preenchidas no formulário.
    #botao_enviar = espera.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="patas"]/main/article/section/form/div[2]/button')))#antigo
    botao_enviar = espera.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="view-pageNavigation"]/div/button')))

    # botao_enviar.click()
    
    
print("Pronto!") 