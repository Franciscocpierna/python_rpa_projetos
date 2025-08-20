import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook

ARQUIVO_CAMINHO = r"C:\python_projetos\python_rpa_projetos\tk_eleicoes\Urna.xlsx"
NOME_PLANILHA = "Dados"


def contabilizar_votos():
    
    
    #Carrega o arquivo
    workbook = load_workbook(filename = ARQUIVO_CAMINHO)
    
    #Obtém a planilha especifica a ser carregada
    sheet = workbook[NOME_PLANILHA]
    
    #Dicionário para armazenar a contagem de votos por candidato
    contagem = {}
    
    #Itera sobre as linhas da planilha, começando da segunda linha
    for row in sheet.iter_rows(min_row=2, values_only=True):
        
        #Obtem o valor do candidato na coluna 0 da linha atual
        candidato = row[0]
        
        #Obtem o valor dos votos na coluna 1 da linha atual
        votos = row[1]
        
        #Verifica se o candidato não é nulo (vazio)
        if candidato:
            
            #Armazena a contagem de votos para o candidato no dicionário
            contagem[candidato] = votos
            
    #Retorna o dicionário com a contagem de votos
    return contagem
            
    
def adicionar_candidato():
    
    #Obtem o candidato digitado
    candidato = entrada_texto.get()
    
    #Limpa o campo
    entrada_texto.delete(0, tk.END)
    
    #Insere o candidato na lista de opções
    lista_opcoes.insert(tk.END, candidato)
    
    #Exibe o resultado da votação atualizado
    exibir_resultado(contabilizar_votos())
    
    #Atualiza o arquivo de contagem de votos com o novo candidato
    atualizar_arquivo(candidato)
    
    #Imprimindo no campo de texto o candidato e o total de votos
    resultado_texto.insert(tk.END, "Candidato {}: 0 voto(s)\n".format(candidato))

def votar() :
    
    #Obtem o candidato digitado
    candidato = entrada_texto.get()
    
    #Limpa o campo
    entrada_texto.delete(0, tk.END)
    
    #Verifica se o candidato está na lista de opções
    if candidato in lista_opcoes.get(0, tk.END):
        
        #Atualiza o arquivo da planilha com o voto para o candidato
        atualizar_arquivo(candidato)
        
        #Pega todos os candidato e votos e imprime na tela
        exibir_resultado(contabilizar_votos())
        
def remover_candidato():
    
    #Obtem o candidato digitado
    candidato = entrada_texto.get()
    
    #Limpa o campo
    entrada_texto.delete(0, tk.END)
    
    #Verifica se o candidato está na lista de opções
    if candidato in lista_opcoes.get(0, tk.END):
        
        #Obtem o índice do candidato na lista de opções
        indice = lista_opcoes.get(0, tk.END).index(candidato)
        
        #Remove o candidato da lista de opções
        lista_opcoes.delete(indice)
        
        #Pega todos os candidato e votos e imprime na tela
        exibir_resultado(contabilizar_votos())
        
        #Remove o candidato do arquivo de contagem de votos
        remover_do_arquivo(candidato)
        
        remover_do_resultado_texto(candidato)
        
        
def remover_do_resultado_texto(candidato):
    
    #Obtém todas as linhas de texto do campo resultado e divide em uma lista
    linhas_resultado = resultado_texto.get(1.0, tk.END).split('\n')
    
    nova_linha_resultado = ""
    
    for linha in linhas_resultado:
        
        #Verifica se a linha começa com o formato Candidato {candidato} :
        if linha.startswith("Candidato {}: ".format(candidato)):
            
            #Se sim, pula para a próxima linha ignorando a linha atual
            continue
            
        #nova_linha_resultado = nova_linha_resultado + novo texto
        nova_linha_resultado += linha + "\n"
        
    
    #Limpa o campo de resultado
    resultado_texto.delete(1.0, tk.END)
    
    #Imprimindo no campo de texto o candidato e o total de votos
    resultado_texto.insert(tk.END, nova_linha_resultado)
  
    
        
def remover_do_arquivo(candidato):
    
    #Carrega o arquivo
    workbook = load_workbook(filename = ARQUIVO_CAMINHO)
    
    #Obtém a planilha especifica a ser carregada
    sheet = workbook[NOME_PLANILHA]
    
    #Itera sobre as linhas da planilha, começando da segunda linha
    for linha in sheet.iter_rows(min_row=2):
        
        #Obtem o valor do candidato na coluna 0 da linha atual
        candidato_planilha = linha[0].value
        
        #Verifica se o candidato da planilha é igual ao candidato atual
        if candidato_planilha == candidato:
            
            #Deleta a linha correspondente à posição da célula da primeira coluna
            sheet.delete_rows(linha[0].row)
            
            break
            
    #Salva as alterações no arquivo da planilha
    workbook.save(ARQUIVO_CAMINHO)
    
def atualizar_arquivo(candidato):
    
    #Carrega o arquivo
    workbook = load_workbook(filename = ARQUIVO_CAMINHO)
    
    #Obtém a planilha especifica a ser carregada
    sheet = workbook[NOME_PLANILHA]
    
    #Variavel para indicar se o candidato é novo ou já existe na planilha
    novo_candidato = True
    
       
    #Itera sobre as linhas da planilha, começando da segunda linha
    for row in sheet.iter_rows(min_row=2):
        
        #Obtem o valor do candidato na coluna 0 da linha atual
        candidato_planilha = row[0].value
        
        #Obtem o valor dos votos na coluna 1 da linha atual
        votos = row[1].value
        
        #Verifica se o candidato da planilha é igual ao candidato atual
        if candidato_planilha == candidato:
            
            #Obtem a célula correspondente aos votos do candito        
            celula_votos = sheet.cell(row=row[0].row, column=2)
            
            #Incrementa o número de votos se já existirem votos registrados
            #Caso contrário atribui o número 1
            celula_votos.value = votos + 1 if votos else 1
            
            #O candidato já existe na planilha, portando não é novo
            novo_candidato = False
            
            break
            
    #Se o candidato for novo
    if novo_candidato:
        
        #Cria uma nova linha com o candidato e 0 votos
        nova_linha = (candidato, 0)
        
        #Adiciona a nova linha a planilha
        sheet.append(nova_linha)
        
    #Salva as alterações no arquivo da planilha
    workbook.save(ARQUIVO_CAMINHO)
    
    
def exibir_resultado(contagem):
    
    #Limpa o campo de texto do resultado
    resultado_texto.delete(1.0, tk.END)
    
    #Insere um cabeçalho
    resultado_texto.insert(tk.END, "Resultado da votação:\n")
    
    #Itera sobre os itens do dicionário de contagem, que contem o candidato e o número de votos
    for candidato, votos in contagem.items():
        
        #Imprimindo no campo de texto o candidato e o total de votos
        resultado_texto.insert(tk.END, "Candidato {}: {} voto(s)\n".format(candidato, votos))
    
def carregar_dados():

    #Carrega o arquivo
    workbook = load_workbook(filename = ARQUIVO_CAMINHO)
    
    #Obtém a planilha especifica a ser carregada
    sheet = workbook[NOME_PLANILHA]
    
    #Dicionário para armazenar a contagem de votos por candidato
    contagem = {}
    
    #Itera sobre as linhas da planilha, começando da segunda linha
    for row in sheet.iter_rows(min_row=2, values_only=True):
        
        #Obtem o valor do candidato na coluna 0 da linha atual
        candidato = row[0]
        
        #Obtem o valor dos votos na coluna 1 da linha atual
        votos = row[1]
        
        #Verifica se o candidato não é nulo (vazio)
        if candidato:
            
            #Armazena a contagem de votos para o candidato no dicionário
            contagem[candidato] = votos
            
            #Insere o candidato na lista de opções - Listbox campo da tela
            lista_opcoes.insert(tk.END, candidato)
            
    
    #Chama a função para exibir o resultado no campo de texto
    exibir_resultado(contagem)
        
        
def verificar_vencedor():
    
    #Realiza a contagem dos votos
    contagem = contabilizar_votos()
    
    #Obtém a quantidade máxima de votos entre os candidatos
    votos_maximo = max(contagem.values())
    
    """
    vencedores = [...]: Isso significa que estamos criando uma nova lista chamada vencedores.
    
    candidato for candidato, votos in contagem.items(): Isso é a parte principal da 
    compreensão de lista. Estamos iterando sobre os itens do dicionário contagem, onde 
    cada item é uma tupla (candidato, votos). Aqui, estamos pegando apenas o valor 
    candidato da tupla e o atribuindo a variável candidato.
    
    if votos == votos_maximos: Aqui, estamos filtrando os candidatos apenas se o 
    número de votos (votos) for igual ao número máximo de votos (votos_maximos). Isso 
    significa que queremos incluir apenas os candidatos que têm o maior número de votos.
    
    A expressão candidato é o valor que será adicionado à lista vencedores se a condição 
    if for verdadeira.
    """
    vencedores = [candidato for candidato, votos in contagem.items() if votos == votos_maximo]
    
    if len(vencedores) == 1:
        
        mensagem = "O vencedor é o candidato {} com {} voto(s)".format(vencedores[0], votos_maximo)
        
    else:
        
        #Cria uma mensagem indicando o empate
        mensagem = "Houve um empate entre os seguintes candidatos: {}".format(", ".join(vencedores))
        
    #Limpa o campo de resultado
    resultado_texto.delete(1.0, tk.END)
    
    #Imprimindo no campo de texto o candidato e o total de votos
    resultado_texto.insert(tk.END, mensagem)

#Cria uma janela principal
tela_principal = tk.Tk()
tela_principal.title("Urna Eletrônica")

#Difinindo o estilo dos widgets
estilo = ttk.Style()
estilo.configure("TButton", font=("Arial", 12), padding=10)
estilo.configure("TLabel", font=("Arial", 12), padding=10)

#Frame principal
frame_principal = ttk.Frame(tela_principal, padding=20)
frame_principal.grid(row=0, column=0, sticky = "NSEW")

#Coluna esquerda
frame_esquerda = ttk.Frame(frame_principal, padding=20)
frame_esquerda.grid(row=0, column=0, sticky = "NSEW")

#Configurando o redimensionamento do frame_esquerda
frame_esquerda.columnconfigure(0, weight=1)
frame_esquerda.rowconfigure(0, weight=1)

#Opçõe dos candidatos
lista_opcoes = tk.Listbox(frame_esquerda, font=("Arial 12"))
lista_opcoes.grid(row=0, column=0, sticky = "NSEW")

#Label
numero_candidato_label = ttk.Label(frame_esquerda,
                                  text="Digite o número do candidato",
                                  font=("Arial 12"))
numero_candidato_label.grid(row=1, column=0, sticky = "NSEW")

#Campo de entrada de dados
entrada_texto = tk.Entry(frame_esquerda,
                                  font=("Arial 20"))
entrada_texto.grid(row=2, column=0, sticky = "NSEW")

#Botão Adicionar
adicionar_botao = ttk.Button(frame_esquerda,
                                  text="Adicionar Candidato",
                                  command = adicionar_candidato)
adicionar_botao.grid(row=3, column=0, sticky = "EW")

#Botão Adicionar
remover_botao = ttk.Button(frame_esquerda,
                                  text="Remover Candidato",
                                  command = remover_candidato)
remover_botao.grid(row=4, column=0, sticky = "EW")

#Botão Votar
votar_botao = ttk.Button(frame_esquerda,
                                  text="Votar",
                                  command = votar)
votar_botao.grid(row=5, column=0, sticky = "EW")

#Verificar resultado eleição
verificar_botao = ttk.Button(frame_esquerda,
                                  text="Verificar Vencedor",
                                  command = verificar_vencedor)
verificar_botao.grid(row=6, column=0, sticky = "EW")

#---------------------------------------------------

#Coluna esquerda
frame_direita = ttk.Frame(frame_principal)
frame_direita.grid(row=0, column=1, sticky = "NSEW")

resultado_label = ttk.Label(frame_direita, 
                           text="Resultado da votação",
                           font=("Arial 16"))
resultado_label.pack() #Cria e centraliza um embaixo do outro


#Definindo o tamanho do resultado_texto
resultado_texto = tk.Text(frame_direita, 
                           font=("Arial 16"))
resultado_texto.pack(fill="both", expand=True, padx=10, pady=10) 

#Configurando o redimensionamento do frame_esquerda
tela_principal.columnconfigure(0, weight=1)
tela_principal.rowconfigure(0, weight=1)

#Configurando o redimensionamento do frame_esquerda
frame_direita.columnconfigure(0, weight=1)
frame_direita.rowconfigure(0, weight=1)

#Chamar a função para carregar os dados
carregar_dados()

#Exibo a tela
tela_principal.mainloop()
