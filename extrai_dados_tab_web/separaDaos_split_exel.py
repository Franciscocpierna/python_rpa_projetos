from selenium import webdriver as opcoesSelenium
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import pyautogui as tempoEspera

from openpyxl import load_workbook
import os

#Pegamos o caminho do arquivo no computador
nome_arquivo_tabela = "C:\\python_projetos\\python_rpa_projetos\\extrai_dados_tab_web\\Dados_Tabela.xlsx"
planilhaDadosTabela = load_workbook(nome_arquivo_tabela)

#Selecionamos a sheet de Dados
sheet_selecionada = planilhaDadosTabela['Dados']
#Deleta todas as linhas
sheet_selecionada.delete_rows(1, sheet_selecionada.max_row)

#Abre o navegador da Web do Google Chrome
navegador = opcoesSelenium.Chrome()
navegador.get("https://rpachallengeocr.azurewebsites.net/")



linha = 1

i = 1
#Enquanto o i for menor do que 4
while i < 4:
    
    #Selecionamos a sheet de Dados
    sheet_dados = planilhaDadosTabela['Dados']

    #Copia o XPATH da tabela
    elementoTabela = navegador.find_element(By.XPATH, '//*[@id="tableSandbox"]')

    #Pega Linhas e Colunas
    linhas = elementoTabela.find_elements(By.TAG_NAME, "tr")
    colunas = elementoTabela.find_elements(By.TAG_NAME, "td")

    #Para cada
    for linhaAtual in linhas:

        
        linha = linha + 1
        #Pagamos a ultima linha + 1
        linha = len(sheet_dados['A']) + 1
        
        #Demos o nome da coluna + o numero da linha
        colunaA = "A" + str(linha) #A2
        colunaB = "B" + str(linha) #B2
        colunaC = "C" + str(linha) #C2
        
        #Pegamos o o texto da linha
        texto = linhaAtual.text
        
        #Separamos com o split todas as palavras com critério do espaço entre texto
        texto2 = texto.split(" ")
        
        #Imprimimos os dados da tabela no Exce
        sheet_dados[colunaA] = texto2[0]
        sheet_dados[colunaB] = texto2[1]
        sheet_dados[colunaC] = texto2[2]
        
    i += 1
    #Aguarda 2 segundos para o computador ou site processar as informações
    tempoEspera.sleep(2)
    
    #Encontra o XPATH do botão Next e clica
    navegador.find_element(By.XPATH, '//*[@id="tableSandbox_next"]').click()
    
    #Aguarda 2 segundos para o computador ou site processar as informações
    tempoEspera.sleep(2)
    
        
else:
    
    print("Pronto!")
    
    
#Salva o arquivo com as alterações
planilhaDadosTabela.save(filename=nome_arquivo_tabela)

#Abre o arquivo
os.startfile(nome_arquivo_tabela)