from selenium import webdriver as opcoesSelenium
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import pyautogui as tempoEspera

from openpyxl import load_workbook
import os

#Pegamos o caminho do arquivo no computador
nome_arquivo_tabela = "C:\\python_projetos\\python_rpa_projetos\\extrai_dados_tab_web\\Dados_Tabela.xlsx"
planilhaDadosTabela = load_workbook(nome_arquivo_tabela)

#Selecionamos a sheet de Dados
sheet_selecionada = planilhaDadosTabela['Dados']
#Deleta todas as linhas
sheet_selecionada.delete_rows(1, sheet_selecionada.max_row)

#Abre o navegador da Web do Google Chrome
navegador = opcoesSelenium.Chrome()
navegador.get("https://rpachallengeocr.azurewebsites.net/")



linha = 1

i = 1
#Enquanto o i for menor do que 4
while i < 4:
    
    #Selecionamos a sheet de Dados
    sheet_dados = planilhaDadosTabela['Dados']

    #Copia o XPATH da tabela
    elementoTabela = navegador.find_element(By.XPATH, '//*[@id="tableSandbox"]')

    #Pega Linhas e Colunas
    linhas = elementoTabela.find_elements(By.TAG_NAME, "tr")
    colunas = elementoTabela.find_elements(By.TAG_NAME, "td")

    #Para cada
    for linhaAtual in linhas:

        
        linha = linha + 1
        linha = len(sheet_dados['A']) + 1
        
        colunaA = "A" + str(linha) #A1
        colunaB = "B" + str(linha) #B1
        colunaC = "C" + str(linha) #C1
        # Salva o texto da linha atual da tabela web na próxima linha disponível da coluna A da planilha Excel
        sheet_dados[colunaA] = linhaAtual.text
        
    i += 1
    #Aguarda 2 segundos para o computador ou site processar as informações
    tempoEspera.sleep(2)
    
    #Encontra o XPATH do botão Next e clica
    navegador.find_element(By.XPATH, '//*[@id="tableSandbox_next"]').click()
    
    #Aguarda 2 segundos para o computador ou site processar as informações
    tempoEspera.sleep(2)
    
        
else:
    
    print("Pronto!")
    
    
#Salva o arquivo com as alterações
planilhaDadosTabela.save(filename=nome_arquivo_tabela)

#Abre o arquivo
os.startfile(nome_arquivo_tabela)