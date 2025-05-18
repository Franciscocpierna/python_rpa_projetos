#Mercado Livre

from selenium import webdriver as opcoesSelenium
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import pyautogui as tempoEspera

from openpyxl import load_workbook
import os

#Pegamos o caminho do arquivo no computador
nome_arquivo_tabela = "C:\\python_projetos\\python_rpa_projetos\\extrai_dados_tab_web\\Dados_Tabela.xlsx"
planilhaDadosTabela = load_workbook(nome_arquivo_tabela)

#Selecionamos a sheet de Dados
sheet_selecionada = planilhaDadosTabela['Dados']
#Deleta todas as linhas
sheet_selecionada.delete_rows(1, sheet_selecionada.max_row)

#Abre o navegador da Web do Google Chrome
navegador = opcoesSelenium.Chrome()
navegador.get("https://www.mercadolivre.com.br/")

#Procura o campo Name, digita a palavra que queremos procurar e cola
#navegador.find_element(By.NAME, "as_word").send_keys("carteira")
tempoEspera.sleep(2)
navegador.find_element(By.NAME, "as_word").send_keys("carteira" + Keys.ENTER)
#Aguarda 2 segundos
#tempoEspera.sleep(2)

#Procura o botão com o XPATH e clica no botão para pesquisar
#navegador.find_element(By.XPATH, "/html/body/header/div/form/button").click()
#navegador.find_element(By.XPATH, "/html/body/header/div/div[2]/form/button[2]").click()



#Aguarda 6 segundos
tempoEspera.sleep(6)

dadosProduto = navegador.find_elements(By.CLASS_NAME, "ui-search-layout__item")


linha = 2
for informacoes in dadosProduto:
   
     #Selecionamos a sheet de Dados
    sheet_dados = planilhaDadosTabela['Dados']
    nomeProduto = informacoes.find_element(By.CLASS_NAME, "poly-component__title").text
    
    
    #precoProduto = informacoes.find_element(By.CLASS_NAME, "price-tag-fraction").text
    #precoProduto = informacoes.find_element(By.CLASS_NAME, "poly-component__price").text
    precoProduto = informacoes.find_element(By.CLASS_NAME, "andes-money-amount__fraction").text
    
       
    
    try:
        
        centavosProduto = informacoes.find_element(By.CLASS_NAME, "andes-money-amount__cents").text

    except:
        centavosProduto = "0"
    
    urlProduto = informacoes.find_element(By.TAG_NAME, "a").get_attribute("href")
    
   # print(nomeProduto + " - " + precoProduto + "," + centavosProduto + " - " + urlProduto)
   # print(nomeProduto + " - " + precoProduto)
   
   #Pagamos a ultima linha + 1
    linha = len(sheet_dados['A']) + 1
    
    #Demos o nome da coluna + o numero da linha
    colunaA = "A" + str(linha) #A2
    colunaB = "B" + str(linha) #B2
    colunaC = "C" + str(linha) #C2
    
    #Imprimimos os dados da tabela no Excel
    sheet_dados["A1"] = "Produto"
    sheet_dados["B1"] = "Preço"
    sheet_dados["C1"] = "Imagem"
    
    # precoTexto = precoProduto + "," + centavosProduto
    # precoSemPonto = precoTexto.replace(".", "")
    # precoSemPonto2 = precoSemPonto.replace(",", ".")
    # precoSemPonto2 =float(precoSemPonto2)
   
    #Imprimimos os dados da tabela no Excel
    sheet_dados[colunaA] = nomeProduto
    sheet_dados[colunaB] = float(precoProduto + "." + centavosProduto)
    sheet_dados[colunaC] = urlProduto
    
    #Salva o arquivo com as alterações
    planilhaDadosTabela.save(filename=nome_arquivo_tabela)

    #Abre o arquivo
    os.startfile(nome_arquivo_tabela)

   