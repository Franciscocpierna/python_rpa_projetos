# Importa as bibliotecas necessárias para manipulação de arquivos Excel e operações no sistema operacional
from openpyxl import load_workbook
from openpyxl import Workbook
import os

# Define o caminho do arquivo Excel que será aberto
nome_arquivo = "C:\\python_projetos\\python_rpa_projetos\\python_excel_avancado\\Quebrar.xlsx"

# Carrega o arquivo Excel especificado
planilha_aberta = load_workbook(filename=nome_arquivo)

# Seleciona a aba chamada 'Dados' dentro do arquivo Excel carregado
sheet_selecionada = planilha_aberta['Dados']

# Cria um novo arquivo Excel em branco
criandoNovoArquivoExcel = Workbook()

# Inicializa a variável que armazenará o nome do vendedor atual
nomeNovo = ""

# Calcula o total de linhas na aba 'Dados' (considerando a coluna 'A')
totalLinha = len(sheet_selecionada['A']) + 1

# Itera pelas linhas da aba 'Dados', começando da segunda linha
for linha in range(2, len(sheet_selecionada['A']) + 1):
    
    # Obtém o nome do vendedor na linha atual
    nomeAtual = sheet_selecionada['A%s' % linha].value
    
    # Verifica se o nome do vendedor atual é igual ao nome do vendedor anterior
    if nomeNovo == nomeAtual:
        
        # Calcula a próxima linha disponível na nova planilha para preenchimento
        linhaSheetQuebra = len(selecionaSheetVendasNovaPlanilha['A']) + 1
        celulaColunaA = "A" + str(linhaSheetQuebra)
        celulaColunaB = "B" + str(linhaSheetQuebra)
        celulaColunaC = "C" + str(linhaSheetQuebra)
                    
        # Preenche os dados do vendedor na nova planilha
        selecionaSheetVendasNovaPlanilha[celulaColunaA] = sheet_selecionada['A%s' % linha].value
        selecionaSheetVendasNovaPlanilha[celulaColunaB] = sheet_selecionada['B%s' % linha].value
        selecionaSheetVendasNovaPlanilha[celulaColunaC] = sheet_selecionada['C%s' % linha].value
        
        # Salva o arquivo Excel com os dados atualizados
        criandoNovoArquivoExcel.save(filename=caminhoNovaPlanilha)
        
    else:
        # Atualiza o nome do vendedor atual
        nomeNovo = sheet_selecionada['A%s' % linha].value
        
        # Obtém a aba ativa do novo arquivo Excel
        nova_planilha = criandoNovoArquivoExcel.active
        
        # Renomeia a aba ativa para 'Vendas'
        nova_planilha.title = "Vendas"
        
        # Define o caminho do novo arquivo Excel com base no nome do vendedor
        caminhoNovaPlanilha = "C:\\python_projetos\\python_rpa_projetos\\python_excel_avancado\\" + sheet_selecionada['A%s' % linha].value + ".xlsx"
        
        # Seleciona a aba 'Vendas' no novo arquivo Excel
        selecionaSheetVendasNovaPlanilha = criandoNovoArquivoExcel['Vendas']
        
        # Adiciona os títulos das colunas na nova planilha
        selecionaSheetVendasNovaPlanilha['A1'] = "Vendedor"
        selecionaSheetVendasNovaPlanilha['B1'] = "Produtos"
        selecionaSheetVendasNovaPlanilha['C1'] = "Vendas"
        
        # Preenche as informações do vendedor na segunda linha
        selecionaSheetVendasNovaPlanilha['A2'] = sheet_selecionada['A%s' % linha].value
        selecionaSheetVendasNovaPlanilha['B2'] = sheet_selecionada['B%s' % linha].value
        selecionaSheetVendasNovaPlanilha['C2'] = sheet_selecionada['C%s' % linha].value
        
        # Remove linhas adicionais desnecessárias na nova planilha
        selecionaSheetVendasNovaPlanilha.delete_rows(3, 100)
        
        # Salva o novo arquivo Excel com os dados do vendedor
        criandoNovoArquivoExcel.save(filename=caminhoNovaPlanilha)