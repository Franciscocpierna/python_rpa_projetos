import tkinter as tk
from tkinter import ttk
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import date

class Onibus:
    
    #Método de inicialização da classe
    def __init__(self, capacidade):
        
        #Recebe a capacidade do ônibus como parâmetro
        self.capacidade = capacidade #Armazena a capacidade do ônibus
        self.lugares = [0] * capacidade #Cria uma lista de lugares vazios  do tamanho da capacidade
        
    def reservar_lugar(self, num_lugar, nome, cpf, dia):
        
        #Verifica se o número do luar está dentro do intervalo válido
        if num_lugar < 1 or num_lugar > self.capacidade:
            
            return "Lugar inválido"
        
        #Verifica se o lugar está disponivel (valor igual a 0 na lista de lugares)
        if self.lugares[num_lugar - 1] == 0:
            
            #marca o lugar como reservado (atualiza o valor na lista de lugares para 1)
            self.lugares[num_lugar - 1] = 1
            
            #Chama o método salvar_reserva, passando os parametros
            self.salvar_reserva(num_lugar, nome, cpf, dia)
            
            #Retorna a mensagem informando o status do lugar no onibus
            return f"Lugar {num_lugar} reservado com sucesso."
        
        else:
            
            #Retorna a mensagem informando o status do lugar no onibus
            return f"Lugar {num_lugar} indisponível."
        
    def cancelar_reserva(self, num_lugar):
        
        #Verifica se o número do luar está dentro do intervalo válido
        if num_lugar < 1 or num_lugar > self.capacidade:
            
            return "Lugar inválido"
        
        #Verifica se o lugar está reservado (valor igual a 1 na lista de lugares)
        if self.lugares[num_lugar - 1] == 1:
            
            #marco o lugar como disponivel (atualiza o valor na lista de lugares para 0)
            self.lugares[num_lugar - 1] = 0
            
            #Chama  método excluir_reserva para remover a reserva 
            self.excluir_reserva(num_lugar)
            
            #Retorna a mensagem informando o status do lugar no onibus
            return f"Lugar {num_lugar} reserva cancelada com sucesso."
        
        else:
            
            #Retorna a mensagem informando o status do lugar no onibus
            return f"Lugar {num_lugar} não está reservado."
        
    def excluir_reserva(self, num_lugar):
        
        try:
            
            #Caminho completo do arquivo do excel
            workbook = load_workbook('C:\\python_projetos\\python_rpa_projetos\\tk_reserva_passagem_onibus\\Dados.xlsx')
            sheet = workbook['Reservas']
            
        except FileNotFoundError:
            
            print("Arquivo não encontrado!")
            return
        
        #Obtém a data a partir do campo EntryData
        data = EntryData.get()
        
        #Percorre as linhas da planilha (exceto a primeira linha de cabeçalho)
        #iter_rows() é usada para percorrer as linhas de uma planilha. 
        #Ela aceita alguns parâmetros opcionais, como min_row e max_row,
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            
            # Verifica se o número do lugar e a data correspondem à reserva a ser excluída
            if row[0].value == num_lugar and row[3].value == data:
                
                #Deleta a linhas correspondente da planilha
                sheet.delete_rows(row[0].row)
                
                # Salva as alterações no arquivo Excel
                workbook.save('C:\\python_projetos\\python_rpa_projetos\\tk_reserva_passagem_onibus\\Dados.xlsx')
                
                #Encerro a execução
                break
        
        #Carrega as reservas novamente
        onibus.carregar_reservas(EntryData.get())
        
        #Atualiza o texto do rótulo com a mensagem sucesso
        resultado_label["text"] = "Reserva deletada com sucesso."
        
    def gerar_mapa(self, data):
        
        #Inicializa a variável mapa como uma string vazia
        mapa = ""
        
        #Percorre os lugares do ônibus de dois em dois
        for i in range(0, self.capacidade, 2):
            
            status_esquerda = " "
            status_direita = " "
            
            #Calcula o número do lugar esquerdo e direito
            lugar_esquerda = i + 1
            lugar_direita = i + 2
            
            #Verifica se o lugar esquerdo está ocupado (valor igual a 1 na lista de lugares)
            status_esquerda = "X" if self.lugares[i] == 1 else " "
            
            #Verifica se o lugar direit está ocupado desde que o número não ultrapasse a capacidade do ônibus
            status_direita = "X" if lugar_direita <= self.capacidade and self.lugares[i + 1] == 1 else " "
  
            #Concatena a representação do lugar com seu status ao mapa
            mapa += f"Lugar {lugar_esquerda}: [{status_esquerda}]    Lugar {lugar_direita}: [{status_direita}] \n"
        
        #Retorno Mapa
        return mapa
    
    #Método para salvar a reserva no arquivo do Excel
    def salvar_reserva(self, num_lugar, nome, cpf, dia):
        
        #Cria uma lista com os dados da reserva
        linha = [num_lugar, nome, cpf, dia]
        
        try:
            
            #Caminho completo do arquivo do excel
            workbook = load_workbook('C:\\python_projetos\\python_rpa_projetos\\tk_reserva_passagem_onibus\\\Dados.xlsx')
            sheet = workbook['Reservas']
            
        except FileNotFoundError:
            
            print("Arquivo não encontrado!")
            
        #Adiciona a linha da reserva à planilha
        sheet.append(linha)
        
        #Salva as alterações no arquivo do excel
        workbook.save('C:\\python_projetos\\python_rpa_projetos\\tk_reserva_passagem_onibus\\Dados.xlsx')
        
    def carregar_reservas(self, data):
        
        try:
            
            #Caminho completo do arquivo do excel
            workbook = load_workbook('C:\\python_projetos\\python_rpa_projetos\\tk_reserva_passagem_onibus\\Dados.xlsx')
            sheet = workbook['Reservas']
            
        except FileNotFoundError:
            
            return
        
        #Limpa a lista de lugares
        self.lugares = [0] * self.capacidade
            
        #Percorre as linhas da planilha
        """
        sheet.iter_rows(values_only=True): O método iter_rows() é chamado em sheet 
        para obter um iterador que percorre as linhas da planilha. O argumento 
        values_only=True é usado para obter apenas os valores das células, sem as 
        informações adicionais.

        enumerate(sheet.iter_rows(values_only=True), start=1): A função enumerate() é 
        aplicada ao iterador resultante da chamada anterior. O parâmetro start=1 é usado 
        para especificar o valor inicial do contador i. Isso significa que a contagem 
        das linhas começará em 1.

        for i, linha in enumerate(sheet.iter_rows(values_only=True), start=1): Essa é a 
        estrutura do loop for que irá percorrer cada linha da planilha. Durante cada 
        iteração, a variável i receberá o valor do contador atual e a variável linha 
        receberá os valores das células da linha atual.
        """
        for i, linha in enumerate(sheet.iter_rows(values_only=True), start=1):
            
            #Ignora a primeira linha de cabeçalho
            if i == 1:
                
                continue
                
            #Pega o número do lugar da coluna A
            num_lugar = linha[0]
            
            #Verifica se o número do lugar está dentro do intervalo válido
            if num_lugar and 1 <= num_lugar <= self.capacidade:
                
                #Verifica se a data da reserva correspondente à data fornecida
                if linha[3] == data:
                    
                    #Marca o lugar como reservado (atualiza o valor na lista de lugares para 1)
                    self.lugares[num_lugar - 1] = 1
                    
        #Limpa o texto do campo que exibe o mapa das reserva do onibus
        mapa_text.delete("1.0", tk.END)
        
        #Insere o mapa atualizado no widget mapa_text
        mapa_text.insert(tk.END, onibus.gerar_mapa(data)) 
            

def reservar_lugar():

    #Obtem o número do lugar, nome, CPF e dia a partir dos campos de entrada de texto
    num_lugar = int(lugar_Entry.get())
    nome = nome_Entry.get()
    cpf = cpf_Entry.get()
    dia = dia_Entry.get()
    
    #Chama o método reservar_lugar do objeto onibus para realizar a reserva
    resultado = onibus.reservar_lugar(num_lugar, nome, cpf, dia)
    
    #Atualiza o texto ro rótulo
    resultado_label["text"] = resultado
    
    #Limpa o texto do campo que exibe o mapa das reserva do onibus
    mapa_text.delete("1.0", tk.END)
    
    onibus.carregar_reservas(EntryData.get())
    
def ver_reservas():
    
    #Carrego o arquivo do excel
    workbook = load_workbook('C:\\python_projetos\\python_rpa_projetos\\tk_reserva_passagem_onibus\\Dados.xlsx')
    
    #Seleciono a planilha Reservas
    sheet = workbook['Reservas']
    
    #Cria uma janela secundária para exibir as reservas
    janela_reservas = tk.Toplevel(janela)
    
    #Define o titulo da janela
    janela_reservas.title("Reservas")
    
    # Obtém a largura e altura da tela
    largura_tela = janela_reservas.winfo_screenwidth()
    altura_tela = janela_reservas.winfo_screenheight()

    # Calcula as coordenadas para centralizar a janela
    posicao_x = int(largura_tela / 2 - 1250 / 2)
    posicao_y = int(altura_tela / 2 - 500 / 2)

    # Define a posição da janela para centralizá-la
    janela_reservas.geometry(f"1250x500+{posicao_x}+{posicao_y}")
    
    
    
    #Cria um frame para o filtro
    frame_filtros = tk.Frame(janela_reservas)
    frame_filtros.pack(pady=10)
    
    #Define os rótulos e campos de entrada do filtro
    rotulos_filtro = ["Lugar", "Nome", "CPF", "Data"]
    
    #Cria uma lista vazia
    campos_filtro = []
    
    # Itera sobre os rótulos de filtro
    for rotulo in rotulos_filtro:
        
        #Cria os campos de Informações
        rotulo_entry = tk.Label(frame_filtros,
                               text=rotulo,
                               font=("Arial 14"))
        rotulo_entry.pack(side=tk.LEFT, padx=10)
        
        #Cria os campos de entrada de dados
        campo = tk.Entry(frame_filtros,
                               text=rotulo,
                               font=("Arial 14"))
        campo.pack(side=tk.LEFT)
        
        #Adiciona o campo de entrada à lista de campos_filtro
        campos_filtro.append(campo)
        
    botao_filtrar = tk.Button(frame_filtros,
                               text="Filtrar",
                               font=("Arial 14"),
                               command=lambda: filtrar_reservas(sheet, rotulos_filtro, campos_filtro, treeview))
    botao_filtrar.pack(side=tk.LEFT, padx=10)
    
    # Criação da Treeview para exibir as reservas
    treeview = ttk.Treeview(janela_reservas)  # Cria um objeto Treeview dentro da janela_reservas
    treeview["columns"] = ("Lugar", "Nome", "CPF", "Data")  # Define as colunas da Treeview como "Lugar", "Nome", "CPF" e "Data"
    treeview.heading("Lugar", text="Lugar", anchor=tk.CENTER)  # Define o título da coluna "Lugar" como "Lugar" e centraliza o texto
    treeview.heading("Nome", text="Nome", anchor=tk.CENTER)  # Define o título da coluna "Nome" como "Nome" e centraliza o texto
    treeview.heading("CPF", text="CPF", anchor=tk.CENTER)  # Define o título da coluna "CPF" como "CPF" e centraliza o texto
    treeview.heading("Data", text="Data", anchor=tk.CENTER)  # Define o título da coluna "Data" como "Data" e centraliza o texto
    treeview.column("#0", width=0, stretch=tk.NO)  # Configura a primeira coluna para ter largura 0 e não ser esticada
    
    # Estilo da Treeview
    style = ttk.Style()
    style.configure("Treeview", font=("Arial", 14))
    style.configure("Treeview.Heading", font=("Arial", 12, "bold"))
    
    # Configuração das colunas
    treeview.column("Lugar", width=100, anchor=tk.CENTER)  # Configura a coluna "Lugar" para ter largura de 100 pixels e centralizar o conteúdo
    treeview.column("Nome", width=200, anchor=tk.W)  # Configura a coluna "Nome" para ter largura de 200 pixels e alinhar o conteúdo à esquerda
    treeview.column("CPF", width=150, anchor=tk.CENTER)  # Configura a coluna "CPF" para ter largura de 150 pixels e centralizar o conteúdo
    treeview.column("Data", width=100, anchor=tk.CENTER)  # Configura a coluna "Data" para ter largura de 100 pixels e centralizar o conteúdo

    # Preenche a Treeview com as reservas da planilha
    #sheet.iter_rows retorna um iterador que fornece acesso às linhas da planilha.
    #values_only=True obter apenas os valores das células, ignorando formatação ou estilos.
    for linha in sheet.iter_rows(min_row=2, values_only=True):
        treeview.insert("", tk.END, values=linha)
        
    # Scrollbar vertical
    # Cria uma barra de rolagem vertical para a treeview (widget de exibição em tabela)
    #treeview.yview é um método que retorna informações sobre a visualização atual da treeview na direção vertical.
    scrollbar = ttk.Scrollbar(janela_reservas, orient=tk.VERTICAL, command=treeview.yview)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y) #RIGHT lado direiro
    
    # Configura a barra de rolagem vertical para controlar a rolagem vertical da treeview
    treeview.configure(yscrollcommand=scrollbar.set)

    # Exibe a Treeview
    #fill=tk.BOTH especifica que deve preencher tanto na direção vertical quanto na direção horizontal.
    treeview.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)
    
def filtrar_reservas(sheet, rotulos_filtro, campos_filtro, treeview):
    
    treeview.delete(*treeview.get_children())  # Limpa os dados existentes na Treeview

    # Itera sobre as linhas da planilha
    for linha in sheet.iter_rows(min_row=2, values_only=True):
        
        #zip está sendo usada para percorrer simultaneamente duas listas: rotulos_filtro e campos_filtro. 
        #Isso permite que seja acessado os elementos correspondentes de ambas as listas ao mesmo tempo
        for rotulo, campo in zip(rotulos_filtro, campos_filtro):
            
            if campo.get():
                
                # Verifica se o campo de filtro não está vazio
                #lower - Converter uma string em minúsculas
                if str(linha[rotulos_filtro.index(rotulo)]).lower() != campo.get().lower():
                    
                    # Se o valor na coluna da linha atual não corresponde ao valor do filtro,
                    # interrompe a verificação dos filtros para essa linha
                    break
                    
            else:
                
                # Se o campo de filtro está vazio, continua para o próximo filtro
                continue
        else:
            
            # Se todos os filtros correspondem, insere a linha na Treeview
            treeview.insert("", tk.END, values=linha)
        
        
    
        
#Cria a janela
janela = tk.Tk() 

#Define o titulo da janela
janela.title("Reserva de Passagens")

#Define a geometria da janela com largura x altura
janela.geometry("400x750")

# Obtém a largura e altura da tela
largura_tela = janela.winfo_screenwidth()
altura_tela = janela.winfo_screenheight()

# Calcula as coordenadas para centralizar a janela
posicao_x = int(largura_tela / 2 - 400 / 2)
posicao_y = int(altura_tela / 2 - 750 / 2)

# Define a posição da janela para centralizá-la
janela.geometry(f"400x750+{posicao_x}+{posicao_y}")

#Define a cor do fundo da janela como branco
janela.configure(bg="#FFFFFF")

#Criação do objeto ônibus
onibus = Onibus(20) #Define a capacidade do ônibus aqui

#Label / Titulo
titulo_label = tk.Label(janela, 
                       text="Reserva de Passagens",
                       font=("Arial 16"),
                       bg="#FFFFFF")
titulo_label.pack(pady=10)  #pack - Cria e centraliza

formulario_frame = tk.Frame(janela, bg="#FFFFFF")
formulario_frame.pack(pady=10)

#sticky - (NSEW) - Norte, Sul, Leste e Oeste
lugar_label = tk.Label(formulario_frame,
                      text="Número do lugar: ",
                      font=("Arial 16"),
                      bg="#FFFFFF")
lugar_label.grid(row=0, column=0, sticky="E") #row - Linha, column - Coluna

#Campo de entrada de dados
lugar_Entry = tk.Entry(formulario_frame, font=("Arial 16"))
lugar_Entry.grid(row=0, column=1, sticky="E") #row - Linha, column - Coluna

#-------------------------------------

nome_label = tk.Label(formulario_frame,
                      text="Nome: ",
                      font=("Arial 16"),
                      bg="#FFFFFF")
nome_label.grid(row=1, column=0, sticky="E") #row - Linha, column - Coluna

#Campo de entrada de dados
nome_Entry = tk.Entry(formulario_frame, font=("Arial 16"))
nome_Entry.grid(row=1, column=1, sticky="E") #row - Linha, column - Coluna

#-------------------------------------

cpf_label = tk.Label(formulario_frame,
                      text="CPF: ",
                      font=("Arial 16"),
                      bg="#FFFFFF")
cpf_label.grid(row=2, column=0, sticky="E") #row - Linha, column - Coluna

#Campo de entrada de dados
cpf_Entry = tk.Entry(formulario_frame, font=("Arial 16"))
cpf_Entry.grid(row=2, column=1, sticky="E") #row - Linha, column - Coluna

#-------------------------------------

dia_label = tk.Label(formulario_frame,
                      text="Dia: ",
                      font=("Arial 16"),
                      bg="#FFFFFF")
dia_label.grid(row=3, column=0, sticky="E") #row - Linha, column - Coluna

#Campo de entrada de dados
dia_Entry = tk.Entry(formulario_frame, font=("Arial 16"))
dia_Entry.grid(row=3, column=1, sticky="E") #row - Linha, column - Coluna


#-------------------------------------

#Cria o botão que faz a reserva
reserva_button = tk.Button(janela,
                          text="Reservar",
                          font=("Arial 16"),
                          command=reservar_lugar)
reserva_button.pack(side="top", fill="both", padx=10, pady=2)

#Cria o botão que cancela a reserva
cancelar_button = tk.Button(janela,
                          text="Cancelar Reserva",
                          font=("Arial 16"),
                          command = lambda: onibus.cancelar_reserva(int(lugar_Entry.get())))
cancelar_button.pack(side="top", fill="both", padx=10, pady=2)


#Cria o botão que ver as reservas
ver_reserva_button = tk.Button(janela,
                          text="Ver Reservas",
                          font=("Arial 16"),
                          command = ver_reservas)
ver_reserva_button.pack(side="top", fill="both", padx=10, pady=2)


#Cria o botão que ver as reservas
resultado_label = tk.Label(janela,
                          text="",
                          font=("Arial 16"))
resultado_label.pack()

mapa_text = tk.Text(janela, 
                   width=50,
                   height=10,
                   bg="#FFFFFF",
                   font=("Arial 18"))
mapa_text.pack()

#Pego a data atual do sistema
data_atual = date.today().strftime("%d/%m/%Y")

#Cria o frame da data
data_frame = tk.Frame(janela, bg="#FFFFFF")
data_frame.pack(pady=10)

#Cria o campo de texto
data_label = tk.Label(data_frame, 
                      text="Data",
                      font="Arial 24",
                      bg="#FFFFFF")
data_label.grid(row=0, column=0, sticky="e")

#Campo de entrada de dados
EntryData = tk.Entry(data_frame, font="Arial 24")
EntryData.grid(row=0, column=1, sticky="e")

#Insiro a data atual no campo de entrada de dados
EntryData.insert(tk.END, data_atual)

#Cria o botão que ver as reservas
trazer_reserva_button = tk.Button(data_frame,
                          text="Filtrar Dia",
                          font=("Arial 24"),
                          command=lambda: onibus.carregar_reservas(EntryData.get()))
trazer_reserva_button.grid(row=1, column=0, columnspan=2, sticky="NSEW")

#Abre a interface
janela.mainloop()