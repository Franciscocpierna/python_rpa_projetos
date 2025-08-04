import tkinter as tk
from tkinter import ttk
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import date

class Onibus:
    
    #Método de inicialização da classe
    def __init__(self, capacidade):
        
        #Recebe a capacidade do ônibus como parâmetro
        self.capacidade = capacidade #Armazena a capacidade do ônibus
        self.lugares = [0] * capacidade #Cria uma lista de lugares vazios  do tamanho da capacidade
        
    def reservar_lugar(self, num_lugar, nome, cpf, dia):
        
        #Verifica se o número do luar está dentro do intervalo válido
        if num_lugar < 1 or num_lugar > self.capacidade:
            
            return "Lugar inválido"
        
        #Verifica se o lugar está disponivel (valor igual a 0 na lista de lugares)
        if self.lugares[num_lugar - 1] == 0:
            
            #marca o lugar como reservado (atualiza o valor na lista de lugares para 1)
            self.lugares[num_lugar - 1] = 1
            
            #Chama o método salvar_reserva, passando os parametros
            self.salvar_reserva(num_lugar, nome, cpf, dia)
            
            #Retorna a mensagem informando o status do lugar no onibus
            return f"Lugar {num_lugar} reservado com sucesso."
        
        else:
            
            #Retorna a mensagem informando o status do lugar no onibus
            return f"Lugar {num_lugar} indisponível."
        
    def gerar_mapa(self, data):
        
        #Inicializa a variável mapa como uma string vazia
        mapa = ""
        
        #Percorre os lugares do ônibus de dois em dois
        for i in range(0, self.capacidade, 2):
            
            status_esquerda = " "
            status_direita = " "
            
            #Calcula o número do lugar esquerdo e direito
            lugar_esquerda = i + 1
            lugar_direita = i + 2
            
            #Verifica se o lugar esquerdo está ocupado (valor igual a 1 na lista de lugares)
            status_esquerda = "X" if self.lugares[i] == 1 else " "
            
            #Verifica se o lugar direit está ocupado desde que o número não ultrapasse a capacidade do ônibus
            status_direita = "X" if lugar_direita <= self.capacidade and self.lugares[i + 1] == 1 else " "
  
            #Concatena a representação do lugar com seu status ao mapa
            mapa += f"Lugar {lugar_esquerda}: [{status_esquerda}]    Lugar {lugar_direita}: [{status_direita}] \n"
        
        #Retorno Mapa
        return mapa
    
    #Método para salvar a reserva no arquivo do Excel
    def salvar_reserva(self, num_lugar, nome, cpf, dia):
        
        #Cria uma lista com os dados da reserva
        linha = [num_lugar, nome, cpf, dia]
        
        try:
            
            #Caminho completo do arquivo do excel
            workbook = load_workbook('C:/Users/55119/Desktop/Reserva Onibus/Dados.xlsx')
            sheet = workbook['Reservas']
            
        except FileNotFoundError:
            
            print("Arquivo não encontrado!")
            
        #Adiciona a linha da reserva à planilha
        sheet.append(linha)
        
        #Salva as alterações no arquivo do excel
        workbook.save('C:/Users/55119/Desktop/Reserva Onibus/Dados.xlsx')
        
    def carregar_reservas(self, data):
        
        try:
            
            #Caminho completo do arquivo do excel
            workbook = load_workbook('C:/Users/55119/Desktop/Reserva Onibus/Dados.xlsx')
            sheet = workbook['Reservas']
            
        except FileNotFoundError:
            
            return
        
        #Limpa a lista de lugares
        self.lugares = [0] * self.capacidade
            
        #Percorre as linhas da planilha
        """
        sheet.iter_rows(values_only=True): O método iter_rows() é chamado em sheet 
        para obter um iterador que percorre as linhas da planilha. O argumento 
        values_only=True é usado para obter apenas os valores das células, sem as 
        informações adicionais.

        enumerate(sheet.iter_rows(values_only=True), start=1): A função enumerate() é 
        aplicada ao iterador resultante da chamada anterior. O parâmetro start=1 é usado 
        para especificar o valor inicial do contador i. Isso significa que a contagem 
        das linhas começará em 1.

        for i, linha in enumerate(sheet.iter_rows(values_only=True), start=1): Essa é a 
        estrutura do loop for que irá percorrer cada linha da planilha. Durante cada 
        iteração, a variável i receberá o valor do contador atual e a variável linha 
        receberá os valores das células da linha atual.
        """
        for i, linha in enumerate(sheet.iter_rows(values_only=True), start=1):
            
            #Ignora a primeira linha de cabeçalho
            if i == 1:
                
                continue
                
            #Pega o número do lugar da coluna A
            num_lugar = linha[0]
            
            #Verifica se o número do lugar está dentro do intervalo válido
            if num_lugar and 1 <= num_lugar <= self.capacidade:
                
                #Verifica se a data da reserva correspondente à data fornecida
                if linha[3] == data:
                    
                    #Marca o lugar como reservado (atualiza o valor na lista de lugares para 1)
                    self.lugares[num_lugar - 1] = 1
                    
        #Limpa o texto do campo que exibe o mapa das reserva do onibus
        mapa_text.delete("1.0", tk.END)
        
        #Insere o mapa atualizado no widget mapa_text
        mapa_text.insert(tk.END, onibus.gerar_mapa(data)) 
            

def reservar_lugar():

    #Obtem o número do lugar, nome, CPF e dia a partir dos campos de entrada de texto
    num_lugar = int(lugar_Entry.get())
    nome = nome_Entry.get()
    cpf = cpf_Entry.get()
    dia = dia_Entry.get()
    
    #Chama o método reservar_lugar do objeto onibus para realizar a reserva
    resultado = onibus.reservar_lugar(num_lugar, nome, cpf, dia)
    
    #Atualiza o texto ro rótulo
    resultado_label["text"] = resultado
    
    #Limpa o texto do campo que exibe o mapa das reserva do onibus
    mapa_text.delete("1.0", tk.END)
    
    onibus.carregar_reservas(EntryData.get())
    
    
        
#Cria a janela
janela = tk.Tk() 

#Define o titulo da janela
janela.title("Reserva de Passagens")

#Define a geometria da janela com largura x altura
janela.geometry("400x750")

#Define a cor do fundo da janela como branco
janela.configure(bg="#FFFFFF")

#Criação do objeto ônibus
onibus = Onibus(20) #Define a capacidade do ônibus aqui

#Label / Titulo
titulo_label = tk.Label(janela, 
                       text="Reserva de Passagens",
                       font=("Arial 16"),
                       bg="#FFFFFF")
titulo_label.pack(pady=10)  #pack - Cria e centraliza

formulario_frame = tk.Frame(janela, bg="#FFFFFF")
formulario_frame.pack(pady=10)

#sticky - (NSEW) - Norte, Sul, Leste e Oeste
lugar_label = tk.Label(formulario_frame,
                      text="Número do lugar: ",
                      font=("Arial 16"),
                      bg="#FFFFFF")
lugar_label.grid(row=0, column=0, sticky="E") #row - Linha, column - Coluna

#Campo de entrada de dados
lugar_Entry = tk.Entry(formulario_frame, font=("Arial 16"))
lugar_Entry.grid(row=0, column=1, sticky="E") #row - Linha, column - Coluna

#-------------------------------------

nome_label = tk.Label(formulario_frame,
                      text="Nome: ",
                      font=("Arial 16"),
                      bg="#FFFFFF")
nome_label.grid(row=1, column=0, sticky="E") #row - Linha, column - Coluna

#Campo de entrada de dados
nome_Entry = tk.Entry(formulario_frame, font=("Arial 16"))
nome_Entry.grid(row=1, column=1, sticky="E") #row - Linha, column - Coluna

#-------------------------------------

cpf_label = tk.Label(formulario_frame,
                      text="CPF: ",
                      font=("Arial 16"),
                      bg="#FFFFFF")
cpf_label.grid(row=2, column=0, sticky="E") #row - Linha, column - Coluna

#Campo de entrada de dados
cpf_Entry = tk.Entry(formulario_frame, font=("Arial 16"))
cpf_Entry.grid(row=2, column=1, sticky="E") #row - Linha, column - Coluna

#-------------------------------------

dia_label = tk.Label(formulario_frame,
                      text="Dia: ",
                      font=("Arial 16"),
                      bg="#FFFFFF")
dia_label.grid(row=3, column=0, sticky="E") #row - Linha, column - Coluna

#Campo de entrada de dados
dia_Entry = tk.Entry(formulario_frame, font=("Arial 16"))
dia_Entry.grid(row=3, column=1, sticky="E") #row - Linha, column - Coluna


#-------------------------------------

#Cria o botão que faz a reserva
reserva_button = tk.Button(janela,
                          text="Reservar",
                          font=("Arial 16"),
                          command=reservar_lugar)
reserva_button.pack(side="top", fill="both", padx=10, pady=2)

#Cria o botão que cancela a reserva
cancelar_button = tk.Button(janela,
                          text="Cancelar Reserva",
                          font=("Arial 16"),
                          command=reservar_lugar)
cancelar_button.pack(side="top", fill="both", padx=10, pady=2)


#Cria o botão que ver as reservas
ver_reserva_button = tk.Button(janela,
                          text="Ver Reservas",
                          font=("Arial 16"),
                          command=reservar_lugar)
ver_reserva_button.pack(side="top", fill="both", padx=10, pady=2)


#Cria o botão que ver as reservas
resultado_label = tk.Label(janela,
                          text="",
                          font=("Arial 16"))
resultado_label.pack()

mapa_text = tk.Text(janela, 
                   width=50,
                   height=10,
                   bg="#FFFFFF",
                   font=("Arial 18"))
mapa_text.pack()

#Pego a data atual do sistema
data_atual = date.today().strftime("%d/%m/%Y")

#Cria o frame da data
data_frame = tk.Frame(janela, bg="#FFFFFF")
data_frame.pack(pady=10)

#Cria o campo de texto
data_label = tk.Label(data_frame, 
                      text="Data",
                      font="Arial 24",
                      bg="#FFFFFF")
data_label.grid(row=0, column=0, sticky="e")

#Campo de entrada de dados
EntryData = tk.Entry(data_frame, font="Arial 24")
EntryData.grid(row=0, column=1, sticky="e")

#Insiro a data atual no campo de entrada de dados
EntryData.insert(tk.END, data_atual)

#Cria o botão que ver as reservas
trazer_reserva_button = tk.Button(data_frame,
                          text="Filtrar Dia",
                          font=("Arial 24"),
                          command=lambda: onibus.carregar_reservas(EntryData.get()))
trazer_reserva_button.grid(row=1, column=0, columnspan=2, sticky="NSEW")

#Abre a interface
janela.mainloop()