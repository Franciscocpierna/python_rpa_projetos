from tkinter import *  # Importa a biblioteca tkinter para criar interfaces gráficas
from tkinter import messagebox  # Importa a função messagebox do tkinter para exibir mensagens de diálogo
from datetime import date  # Importa a classe date do módulo datetime
import tkinter as tk
from tkinter import ttk
from openpyxl import Workbook, load_workbook
from datetime import datetime


def tela_cliente():
    
    # Cria a janela de cadastro de cliente
    janela_cadastrar_cliente = tk.Toplevel()
    janela_cadastrar_cliente.title("Clientes")
    janela_cadastrar_cliente.configure(bg="white")
    
    #Dicionário para armazenar os dados dos clientes
    Clientes = {}
    
    # Cria um Frame para o filtro
    filtro_frame = tk.Frame(janela_cadastrar_cliente, bg="white")
    filtro_frame.pack(pady = 10)

    # Cria um rótulo para o campo de nome
    filtro_label = tk.Label(filtro_frame,
                       text = "Filtrar por nome:",
                       font="Arial 14",
                       bg="white")
    filtro_label.grid(row = 0, column = 0, padx = 5, pady = 3)
    
    # Cria uma entrada de texto para o nome
    filtro_entry = Entry(filtro_frame, font="Arial 14")
    filtro_entry.grid(row = 0, column = 1, padx = 5, pady = 3)
    
    # Cria um Frame para a treeview e a barra de rolagem
    frame_treeview = tk.Frame(janela_cadastrar_cliente, bg="white")
    frame_treeview.pack(fill = tk.BOTH, expand = True)
    
    #Cria a treeview
    clientes_treeview = ttk.Treeview(frame_treeview, height=7) #Define a altura deseja da Treeview
    
    #Define o estilo da fonte para todas as colunas
    style = ttk.Style()
    style.configure("Treeview", font=("Arial", 14), rowheight=28) #Altera o valor de rowheight conforme necessário
    
    #Define as colunas da treeview
    clientes_treeview['columns'] = ('CPF', 'Sexo', 'Nome', 'Aniversário', 'Email')

    # Define os cabeçalhos das colunas
    clientes_treeview.heading('#0', text='', anchor='w')  # Cabeçalho vazio na primeira coluna
    clientes_treeview.column('#0', width=0, stretch='NO')  # Define a largura da primeira coluna como zero
    
    clientes_treeview.heading('CPF', text='CPF', anchor='w')  # Define o cabeçalho da coluna CPF
    clientes_treeview.column('CPF', anchor='w')  # Define a âncora da coluna CPF como 'w' (oeste) 
    
    clientes_treeview.heading('Sexo', text='Sexo', anchor='w')  # Define o cabeçalho da coluna Sexo
    clientes_treeview.column('Sexo', anchor='w')  # Define a âncora da coluna Sexo como 'w' (oeste)

    clientes_treeview.heading('Nome', text='Nome', anchor='w')  # Define o cabeçalho da coluna Nome
    clientes_treeview.column('Nome', anchor='w')  # Define a âncora da coluna Nome como 'w' (oeste)

    clientes_treeview.heading('Aniversário', text='Aniversário', anchor='w')  # Define o cabeçalho da coluna Aniversário
    clientes_treeview.column('Aniversário', anchor='w')  # Define a âncora da coluna Aniversário como 'w' (oeste)

    clientes_treeview.heading('Email', text='Email', anchor='w')  # Define o cabeçalho da coluna Email
    clientes_treeview.column('Email', anchor='w')  # Define a âncora da coluna Email como 'w' (oeste)
    
    clientes_treeview.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)  # Empacota e posiciona a treeview

    # Cria uma barra de rolagem vertical
    scrollbar = ttk.Scrollbar(frame_treeview, orient="vertical", command=clientes_treeview.yview)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    
    # Configura a barra de rolagem na treeview
    clientes_treeview.configure(yscrollcommand=scrollbar.set)
    
    #Frame para os campos de cadastro/alteração
    cadastro_frame = tk.Frame(janela_cadastrar_cliente)
    cadastro_frame.pack(pady=10)
    
    cpf_label = tk.Label(cadastro_frame, text="CPF", font="Arial 14")
    cpf_label.grid(row=0, column=0, padx=5, pady=3)
    cpf_entry = tk.Entry(cadastro_frame, font="Arial 14", width=30) # Definindo largura de 30 caracteres (aproximadamente 200 pixels) 
    cpf_entry.grid(row=0, column=1, padx=5, pady=3)
    
    sexo_label = tk.Label(cadastro_frame, text="Sexo", font="Arial 14")
    sexo_label.grid(row=1, column=0, padx=5, pady=3)
    sexo_combobox = ttk.Combobox(cadastro_frame, font="Arial 14", values=["Feminino", "Masculino"], width=27) # Definindo largura de 30 caracteres (aproximadamente 200 pixels) 
    sexo_combobox.grid(row=1, column=1, padx=5, pady=3)
    
    nome_label = tk.Label(cadastro_frame, text="Nome", font="Arial 14")
    nome_label.grid(row=2, column=0, padx=5, pady=3)
    nome_entry = tk.Entry(cadastro_frame, font="Arial 14", width=30) # Definindo largura de 30 caracteres (aproximadamente 200 pixels) 
    nome_entry.grid(row=2, column=1, padx=5, pady=3)
    
    aniversario_label = tk.Label(cadastro_frame, text="Aniversário", font="Arial 14")
    aniversario_label.grid(row=3, column=0, padx=5, pady=3)
    aniversario_entry = tk.Entry(cadastro_frame, font="Arial 14", width=30) # Definindo largura de 30 caracteres (aproximadamente 200 pixels) 
    aniversario_entry.grid(row=3, column=1, padx=5, pady=3)
    
    email_label = tk.Label(cadastro_frame, text="Email", font="Arial 14")
    email_label.grid(row=4, column=0, padx=5, pady=3)
    email_entry = tk.Entry(cadastro_frame, font="Arial 14", width=30) # Definindo largura de 30 caracteres (aproximadamente 200 pixels) 
    email_entry.grid(row=4, column=1, padx=5, pady=3)
    
    def filtrar_clientes():
        
        # Obtém o texto do filtro e converte para minúsculas
        filtro = filtro_entry.get().lower()
        
        #Limpa os dados da treeview removendo todos os itens
        clientes_treeview.delete(*clientes_treeview.get_children())
        
        # Itera sobre os clientes armazenados no dicionário
        for cliente in Clientes.values():
            
            # Obtém o nome do cliente atual e converte para minúsculas
            nome = cliente['Nome'].lower()
            
            #Verifica se o filtro está contido no nome do cliente
            if filtro in nome:
                
                nome_planilha = cliente['Nome']
                cpf = cliente['CPF']
                sexo = cliente['Sexo']
                aniversario = cliente['Aniversario']
                email = cliente['Email']
                
                # Insere os valores do cliente na treeview
                clientes_treeview.insert('', 'end', values=(cpf, sexo, nome_planilha, aniversario, email))
              
    #Associa o evento KeyRelease ao campo filtro, para acionar a função filtrar_clientes
    filtro_entry.bind("<KeyRelease>", lambda event: filtrar_clientes())
    
    def carregar_clientes_do_excel():
        
        #Limpa os dados da treeview removendo todos os itens
        clientes_treeview.delete(*clientes_treeview.get_children())
        
        #Caminho do arquivo Excel
        caminho_arquivo = r"C:\python_projetos\python_rpa_projetos\tk_controle_pedido\Base+Dados.xlsx"
        
        #Nome da planilha
        planilha_nome = "Clientes"
        
        try:
            
            # Carrega o arquivo Excel
            workbook = load_workbook(filename=caminho_arquivo)
            sheet = workbook[planilha_nome]
            
            # Itera sobre as linhas da planilha (excluindo a primeira linha de cabeçalho)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                
                # Obtém os valores de cada coluna da linha
                cpf, sexo, nome, aniversario, email = row
                
                # Adiciona o cliente ao dicionário de Clientes
                Clientes[nome] = {'CPF': cpf, 'Sexo': sexo, 'Nome': nome, 'Aniversario': aniversario, 'Email': email}
                
                # Insere os valores do cliente na treeview
                clientes_treeview.insert('', 'end', values=(cpf, sexo, nome, aniversario, email))
              
            
        except FileNotFoundError:
            
            # Exibe uma mensagem de aviso se o arquivo não for encontrado
            messagebox.showwarning("Carregar Clientes", "Arquivo Clientes.xlsx não encontrado!", parent=janela_cadastrar_cliente)
            
    #Preencher a treeview com os clientes existentes
    carregar_clientes_do_excel()
    
    
    def preencher_campos(event):
        
        #Obtém o item selecionado na treeview
        item_selecionado = clientes_treeview.focus()
        
        if item_selecionado:
            
            #Obtém os valores das colunas do item selecionado
            cpf = clientes_treeview.item(item_selecionado)['values'][0]
            sexo = clientes_treeview.item(item_selecionado)['values'][1]
            nome = clientes_treeview.item(item_selecionado)['values'][2]
            aniversario = clientes_treeview.item(item_selecionado)['values'][3]
            email = clientes_treeview.item(item_selecionado)['values'][4]
            
            #Preenche os campos de entrada com os valores do cliente selecionado
            cpf_entry.delete(0, tk.END)
            cpf_entry.insert(tk.END, cpf)
            
            sexo_combobox.set(sexo)
            
            nome_entry.delete(0, tk.END)
            nome_entry.insert(tk.END, nome)
            
            aniversario_entry.delete(0, tk.END)
            aniversario_entry.insert(tk.END, aniversario)
            
            email_entry.delete(0, tk.END)
            email_entry.insert(tk.END, email)
    
    clientes_treeview.bind("<Double-1>", preencher_campos)
            
    def cadastrar_cliente():
        
        # Obtém os valores dos campos de entrada
        cpf = cpf_entry.get()
        sexo = sexo_combobox.get()
        nome = nome_entry.get()
        aniversario = aniversario_entry.get()
        email = email_entry.get()
        
        if cpf and sexo and nome and aniversario and email:
            
            if nome not in Clientes:
                
                # Adiciona o cliente ao dicionário de Clientes
                Clientes[nome] = {'CPF': cpf, 'Sexo': sexo, 'Nome': nome, 'Aniversario': aniversario, 'Email': email}
                
                # Insere os valores do cliente na treeview
                clientes_treeview.insert('', 'end', values=(cpf, sexo, nome, aniversario, email))
                
                # Exibe uma mensagem de sucesso
                messagebox.showinfo("Cadastro de Cliente", "Cliente cadastrado com sucesso!", parent=janela_cadastrar_cliente)
                
                limpa_campos()
                
                salvar_clientes_excel()
            
        else:
            
            messagebox.showerror("Cadastro de Cliente", "Preencha todos os campos!", parent=janela_cadastrar_cliente)
    
    # Cria o botão de cadastrar cliente
    cadastrar_botao = tk.Button(janela_cadastrar_cliente,
                               text="Cadastrar",
                               command=cadastrar_cliente,
                               font="Arial 14")
    cadastrar_botao.pack(side=tk.LEFT, padx=5, pady=3)
    
    def alterar_cliente():
        
        # Obtém os valores dos campos de entrada
        cpf = cpf_entry.get()
        sexo = sexo_combobox.get()
        nome = nome_entry.get()
        aniversario = aniversario_entry.get()
        email = email_entry.get()
        
        if cpf and sexo and nome and aniversario and email:
            
            #Caminho do arquivo Excel
            caminho_arquivo = r"C:\python_projetos\python_rpa_projetos\tk_controle_pedido\Base+Dados.xlsx"

            #Nome da planilha
            planilha_nome = "Clientes"

            try:

                # Carrega o arquivo Excel
                workbook = load_workbook(filename=caminho_arquivo)
                sheet = workbook[planilha_nome]

                # Itera sobre as linhas da planilha (excluindo a primeira linha de cabeçalho)
                for row in sheet.iter_rows(min_row=2):

                    if row[0].value == cpf:
                        
                        #Atualiza os valores do cliente na planilha do Excel
                        row[1].value = sexo
                        row[2].value = nome
                        row[3].value = aniversario
                        row[4].value = email
                        
                        break
                        
                workbook.save(caminho_arquivo)
                
                
                
                # Exibe uma mensagem de aviso se o arquivo não for encontrado
                messagebox.showinfo("Alterar Clientes", "Cliente alterado com sucesso!", parent=janela_cadastrar_cliente)
                
                #Limpa os campos de entrada
                limpa_campos()
                
                #Preenche o Treeview com os clientes existentes e alterados
                carregar_clientes_do_excel()

                    

            except FileNotFoundError:

                # Exibe uma mensagem de aviso se o arquivo não for encontrado
                messagebox.showwarning("Alterar Cliente", "Base Dados.xlsx não encontrado!", parent=janela_cadastrar_cliente)

        
            
        else:
            
            # Exibe uma mensagem de aviso se o arquivo não for encontrado
            messagebox.showwarning("Alterar Cliente", "Preencha o campo CPF!", parent=janela_cadastrar_cliente)

    
    # Cria o botão de alterar cliente
    alterar_botao = tk.Button(janela_cadastrar_cliente,
                               text="Alterar",
                               command=alterar_cliente,
                               font="Arial 14")
    alterar_botao.pack(side=tk.LEFT, padx=5, pady=3)
    
    def excluir_cliente():
        
        #Obtém o valor do campo CPF
        cpf = cpf_entry.get()
        
        if cpf:
            
            #Caminho do arquivo Excel
            caminho_arquivo = r"C:\python_projetos\python_rpa_projetos\tk_controle_pedido\Base+Dados.xlsx"
            #Nome da planilha
            planilha_nome = "Clientes"

            try:

                # Carrega o arquivo Excel
                workbook = load_workbook(filename=caminho_arquivo)
                sheet = workbook[planilha_nome]

                # Itera sobre as linhas da planilha (excluindo a primeira linha de cabeçalho)
                for row in sheet.iter_rows(min_row=2):

                    if row[0].value == cpf:
                        
                        #Exclui a linha correspondente ao CPF do cliente na planilha do Excel
                        sheet.delete_rows(row[0].row)
                        
                        break
                        
                workbook.save(caminho_arquivo)
                
                
                
                # Exibe uma mensagem de sucesso
                messagebox.showinfo("Excluir Cliente", "Cliente excluido com sucesso!", parent=janela_cadastrar_cliente)
                
                #Limpa os campos de entrada
                limpa_campos()
                
                #Preenche o Treeview com os clientes existentes e alterados
                carregar_clientes_do_excel()

                    

            except FileNotFoundError:

                # Exibe uma mensagem de aviso se o arquivo não for encontrado
                messagebox.showwarning("Excluir Cliente", "Base Dados.xlsx não encontrado!", parent=janela_cadastrar_cliente)

        
        
        else:
            
            # Exibe uma mensagem de aviso se o arquivo não for encontrado
            messagebox.showwarning("Excluir Cliente", "Preencha o campo CPF!", parent=janela_cadastrar_cliente)

    
    # Cria o botão de alterar cliente
    excluir_botao = tk.Button(janela_cadastrar_cliente,
                               text="Excluir",
                               command=excluir_cliente,
                               font="Arial 14")
    excluir_botao.pack(side=tk.LEFT, padx=5, pady=3)
    
    def limpa_campos():
        
        cpf_entry.delete(0, tk.END)
        sexo_combobox.delete(0, tk.END)
        nome_entry.delete(0, tk.END)
        aniversario_entry.delete(0, tk.END)
        email_entry.delete(0, tk.END)
                
    def salvar_clientes_excel():
        
        #Caminho do arquivo Excel
        caminho_arquivo = r"C:\python_projetos\python_rpa_projetos\tk_controle_pedido\Base+Dados.xlsx"
        
        #Nome da planilha
        planilha_nome = "Clientes"
        
        try:
            
            #Tenta carregar o arquivo existente
            workbook = load_workbook(filename=caminho_arquivo)
            
        except FileNotFoundError:
            
            print("Arquivo não existe")
            
        #Verifica se a planilha Clientes já existe no arquivo
        if planilha_nome not in workbook.sheetnames:
            
            #Cria a planilha Clientes
            workbook.create_sheet(planilha_nome)
            
        #Obtém a referência para a planilha
        sheet = workbook[planilha_nome]
            
        #Percorre os clientes no dicionário
        for cliente in Clientes.values():
            
            cpf = cliente['CPF']
            sexo = cliente['Sexo']
            nome = cliente['Nome']
            aniversario = cliente['Aniversario']
            email = cliente['Email']
            
            #Verifica se o cliente já existe na planilha
            cliente_existe = False
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
            
                if row[0] == cpf:
                    
                    # Se o cliente já existe, marca como encontrado e interrompe o loop
                    cliente_existe = True
                    
                    break
                    
            # Se o cliente não existe, adiciona na planilha
            if not cliente_existe:

                # Obtém a próxima linha disponível na planilha
                max_row = sheet.max_row + 1
                
                #Insere os dados do cliente nas colunas correspondentes
                sheet.cell(row=max_row, column=1, value=cpf)
                sheet.cell(row=max_row, column=2, value=sexo)
                sheet.cell(row=max_row, column=3, value=nome)
                sheet.cell(row=max_row, column=4, value=aniversario)
                sheet.cell(row=max_row, column=5, value=email)
                
        
        # Salva as alterações no arquivo
        workbook.save(filename=caminho_arquivo)
            
    
    # Atualiza a interface gráfica para processar todos os eventos pendentes
    janela_cadastrar_cliente.update_idletasks()

    # Obtém a largura atual da janela
    largura = janela_cadastrar_cliente.winfo_width()

    # Obtém a altura atual da janela
    altura = janela_cadastrar_cliente.winfo_height()

    # Calcula a posição x para centralizar a janela na tela
    posicao_x = (janela_cadastrar_cliente.winfo_screenwidth() // 2) - (largura // 2)

    # Calcula a posição y para centralizar a janela na tela
    posicao_y = (janela_cadastrar_cliente.winfo_screenheight() // 2) - (altura // 2)

    # Define a geometria da janela com a largura, altura, posição x e posição y
    janela_cadastrar_cliente.geometry('{}x{}+{}+{}'.format(largura, altura, posicao_x, posicao_y))    
    

    
#------------------------------------------------------------------------------------
#------------------------------------------------------------------------------------

def tela_produtos():
    
    # Cria a janela de cadastro de cliente
    janela_cadastrar_produtos = tk.Toplevel()
    janela_cadastrar_produtos.title("Produtos")
    janela_cadastrar_produtos.configure(bg="white")
    
    #Dicionário para armazenar os dados dos clientes
    Produtos = {}
    
    # Cria um Frame para o filtro
    filtro_frame = tk.Frame(janela_cadastrar_produtos, bg="white")
    filtro_frame.pack(pady = 10)

    # Cria um rótulo para o campo de nome
    filtro_label = tk.Label(filtro_frame,
                       text = "Filtrar por produto:",
                       font="Arial 14",
                       bg="white")
    filtro_label.grid(row = 0, column = 0, padx = 5, pady = 3)
    
    # Cria uma entrada de texto para o nome
    filtro_entry = Entry(filtro_frame, font="Arial 14")
    filtro_entry.grid(row = 0, column = 1, padx = 5, pady = 3)
    
    # Cria um Frame para a treeview e a barra de rolagem
    frame_treeview = tk.Frame(janela_cadastrar_produtos, bg="white")
    frame_treeview.pack(fill = tk.BOTH, expand = True)
    
    #Cria a treeview
    produtos_treeview = ttk.Treeview(frame_treeview, height=7) #Define a altura deseja da Treeview
    
    #Define o estilo da fonte para todas as colunas
    style = ttk.Style()
    style.configure("Treeview", font=("Arial", 14), rowheight=28) #Altera o valor de rowheight conforme necessário
    
    #Define as colunas da treeview
    produtos_treeview['columns'] = ('Codigo', 'Categoria', 'Nome', 'Preco')

    # Define os cabeçalhos das colunas
    produtos_treeview.heading('#0', text='', anchor='w')  # Cabeçalho vazio na primeira coluna
    produtos_treeview.column('#0', width=0, stretch='NO')  # Define a largura da primeira coluna como zero
    
    produtos_treeview.heading('Codigo', text='Codigo', anchor='w')  # Define o cabeçalho da coluna Codigo
    produtos_treeview.column('Codigo', anchor='w')  # Define a âncora da coluna Codigo como 'w' (oeste) 
    
    produtos_treeview.heading('Categoria', text='Categoria', anchor='w')  # Define o cabeçalho da coluna Sexo
    produtos_treeview.column('Categoria', anchor='w')  # Define a âncora da coluna Sexo como 'w' (oeste)

    produtos_treeview.heading('Nome', text='Nome', anchor='w')  # Define o cabeçalho da coluna Nome
    produtos_treeview.column('Nome', anchor='w')  # Define a âncora da coluna Nome como 'w' (oeste)

    produtos_treeview.heading('Preco', text='Preco', anchor='w')  # Define o cabeçalho da coluna Aniversário
    produtos_treeview.column('Preco', anchor='w')  # Define a âncora da coluna Aniversário como 'w' (oeste)

    
    produtos_treeview.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)  # Empacota e posiciona a treeview

    # Cria uma barra de rolagem vertical
    scrollbar = ttk.Scrollbar(frame_treeview, orient="vertical", command=produtos_treeview.yview)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    
    # Configura a barra de rolagem na treeview
    produtos_treeview.configure(yscrollcommand=scrollbar.set)
    
    #Frame para os campos de cadastro/alteração
    cadastro_frame = tk.Frame(janela_cadastrar_produtos)
    cadastro_frame.pack(pady=10)
    
    codigo_label = tk.Label(cadastro_frame, text="Código do Produto", font="Arial 14")
    codigo_label.grid(row=0, column=0, padx=5, pady=3)
    codigo_entry = tk.Entry(cadastro_frame, font="Arial 14", width=30) # Definindo largura de 30 caracteres (aproximadamente 200 pixels) 
    codigo_entry.grid(row=0, column=1, padx=5, pady=3)
    
    categoria_label = tk.Label(cadastro_frame, text="Categoria Produto", font="Arial 14")
    categoria_label.grid(row=1, column=0, padx=5, pady=3)
    categoria_combobox = ttk.Combobox(cadastro_frame, font="Arial 14", values=["Alimentos", "Bebidas"], width=27) # Definindo largura de 30 caracteres (aproximadamente 200 pixels) 
    categoria_combobox.grid(row=1, column=1, padx=5, pady=3)
    
    nome_label = tk.Label(cadastro_frame, text="Nome", font="Arial 14")
    nome_label.grid(row=2, column=0, padx=5, pady=3)
    nome_entry = tk.Entry(cadastro_frame, font="Arial 14", width=30) # Definindo largura de 30 caracteres (aproximadamente 200 pixels) 
    nome_entry.grid(row=2, column=1, padx=5, pady=3)
    
    preco_label = tk.Label(cadastro_frame, text="Preço", font="Arial 14")
    preco_label.grid(row=3, column=0, padx=5, pady=3)
    preco_entry = tk.Entry(cadastro_frame, font="Arial 14", width=30) # Definindo largura de 30 caracteres (aproximadamente 200 pixels) 
    preco_entry.grid(row=3, column=1, padx=5, pady=3)
    
       
    def filtrar_produtos():
        
        # Obtém o texto do filtro e converte para minúsculas
        filtro = filtro_entry.get().lower()
        
        #Limpa os dados da treeview removendo todos os itens
        produtos_treeview.delete(*produtos_treeview.get_children())
        
        # Itera sobre os produtos armazenados no dicionário
        for produto in Produtos.values():
            
            # Obtém o nome do cliente atual e converte para minúsculas
            nome = produto['Nome'].lower()
            
            #Verifica se o filtro está contido no nome do cliente
            if filtro in nome:
                
                codigo = produto['Codigo']
                categoria = produto['Categoria']
                nome_planilha = produto['Nome']
                preco = produto['Preço']
                
                # Insere os valores do produto na treeview
                produtos_treeview.insert('', 'end', values=(codigo, categoria, nome_planilha, preco))
              
    #Associa o evento KeyRelease ao campo filtro, para acionar a função filtrar_produtos
    filtro_entry.bind("<KeyRelease>", lambda event: filtrar_produtos())
    
    def carregar_produtos_do_excel():
        
        #Limpa os dados da treeview removendo todos os itens
        produtos_treeview.delete(*produtos_treeview.get_children())
        
        #Caminho do arquivo Excel
        caminho_arquivo = r"C:\python_projetos\python_rpa_projetos\tk_controle_pedido\Base+Dados.xlsx"
        
        #Nome da planilha
        planilha_nome = "Produtos"
        
        try:
            
            # Carrega o arquivo Excel
            workbook = load_workbook(filename=caminho_arquivo)
            sheet = workbook[planilha_nome]
            
            # Itera sobre as linhas da planilha (excluindo a primeira linha de cabeçalho)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                
                # Obtém os valores de cada coluna da linha
                codigo, categoria, nome, preco = row
                
                # Adiciona o cliente ao dicionário de Clientes
                Produtos[codigo] = {'Codigo': codigo, 'Categoria': categoria, 'Nome': nome, 'Preço': preco}
                
                # Insere os valores do produto na treeview
                produtos_treeview.insert('', 'end', values=(codigo, categoria, nome, preco))              
            
        except FileNotFoundError:
            
            # Exibe uma mensagem de aviso se o arquivo não for encontrado
            messagebox.showwarning("Carregar Produtos", "Base Dados.xlsx não encontrado!", parent=janela_cadastrar_produtos)
            
    #Preencher a treeview com os clientes existentes
    carregar_produtos_do_excel()
    
    
    def preencher_campos(event):
        
        #Obtém o item selecionado na treeview
        item_selecionado = produtos_treeview.focus()
        
        if item_selecionado:
            
            #Obtém os valores das colunas do item selecionado
            codigo = produtos_treeview.item(item_selecionado)['values'][0]
            categoria = produtos_treeview.item(item_selecionado)['values'][1]
            nome = produtos_treeview.item(item_selecionado)['values'][2]
            preco = produtos_treeview.item(item_selecionado)['values'][3]
            
            #Preenche os campos de entrada com os valores do cliente selecionado
            codigo_entry.delete(0, tk.END)
            codigo_entry.insert(tk.END, codigo)
            
            categoria_combobox.set(categoria)
            
            nome_entry.delete(0, tk.END)
            nome_entry.insert(tk.END, nome)
            
            preco_entry.delete(0, tk.END)
            preco_entry.insert(tk.END, preco)
            
                
    produtos_treeview.bind("<Double-1>", preencher_campos)
            
    def cadastrar_produto():
        
        # Obtém os valores dos campos de entrada
        codigo = codigo_entry.get()
        categoria = categoria_combobox.get()
        nome = nome_entry.get()
        preco = preco_entry.get()
        
        if codigo and categoria and nome and preco:
            
            if codigo not in Produtos:
                
                # Adiciona o cliente ao dicionário de Clientes
                Produtos[codigo] = {'Codigo': codigo, 'Categoria': categoria, 'Nome': nome, 'Preço': preco}
                
                # Insere os valores do produto na treeview
                produtos_treeview.insert('', 'end', values=(codigo, categoria, nome, preco))             
            
                # Exibe uma mensagem de sucesso
                messagebox.showinfo("Cadastro de Produto", "Produto cadastrado com sucesso!", parent=janela_cadastrar_produtos)
                
                limpa_campos()
                
                salvar_produtos_excel()
            
        else:
            
            messagebox.showerror("Cadastro de Produto", "Preencha todos os campos!", parent=janela_cadastrar_produtos)
    
    # Cria o botão de cadastrar cliente
    cadastrar_botao = tk.Button(janela_cadastrar_produtos,
                               text="Cadastrar",
                               command=cadastrar_produto,
                               font="Arial 14")
    cadastrar_botao.pack(side=tk.LEFT, padx=5, pady=3)
    
    def alterar_produto():
        
        # Obtém os valores dos campos de entrada
        codigo = codigo_entry.get()
        categoria = categoria_combobox.get()
        nome = nome_entry.get()
        preco = preco_entry.get()
        
        if codigo and categoria and nome and preco:
            
            #Caminho do arquivo Excel
            caminho_arquivo = r"C:\python_projetos\python_rpa_projetos\tk_controle_pedido\Base+Dados.xlsx"

            #Nome da planilha
            planilha_nome = "Produtos"

            try:

                # Carrega o arquivo Excel
                workbook = load_workbook(filename=caminho_arquivo)
                sheet = workbook[planilha_nome]

                # Itera sobre as linhas da planilha (excluindo a primeira linha de cabeçalho)
                for row in sheet.iter_rows(min_row=2):

                    if row[0].value == codigo:
                        
                        #Atualiza os valores do cliente na planilha do Excel
                        row[1].value = categoria
                        row[2].value = nome
                        row[3].value = preco
                        
                        break
                        
                workbook.save(caminho_arquivo)
                
                
                
                # Exibe uma mensagem de aviso se o arquivo não for encontrado
                messagebox.showinfo("Alterar Produtos", "Produto alterado com sucesso!", parent=janela_cadastrar_produtos)
                
                #Limpa os campos de entrada
                limpa_campos()
                
                #Preenche o Treeview com os clientes existentes e alterados
                carregar_produtos_do_excel()

                    

            except FileNotFoundError:

                # Exibe uma mensagem de aviso se o arquivo não for encontrado
                messagebox.showwarning("Alterar Produto", "Base Dados.xlsx não encontrado!", parent=janela_cadastrar_produtos)

        
            
        else:
            
            # Exibe uma mensagem de aviso se o arquivo não for encontrado
            messagebox.showwarning("Alterar Produto", "Preencha o campo Codigo!", parent=janela_cadastrar_produtos)

    
    # Cria o botão de alterar cliente
    alterar_botao = tk.Button(janela_cadastrar_produtos,
                               text="Alterar",
                               command=alterar_produto,
                               font="Arial 14")
    alterar_botao.pack(side=tk.LEFT, padx=5, pady=3)
    
    def excluir_produto():
        
        #Obtém o valor do campo CPF
        codigo = codigo_entry.get()
        
        if codigo:
            
            #Caminho do arquivo Excel
            caminho_arquivo = r"C:\python_projetos\python_rpa_projetos\tk_controle_pedido\Base+Dados.xlsx"

            #Nome da planilha
            planilha_nome = "Produtos"

            try:

                # Carrega o arquivo Excel
                workbook = load_workbook(filename=caminho_arquivo)
                sheet = workbook[planilha_nome]

                # Itera sobre as linhas da planilha (excluindo a primeira linha de cabeçalho)
                for row in sheet.iter_rows(min_row=2):

                    if row[0].value == codigo:
                        
                        #Exclui a linha correspondente ao CPF do cliente na planilha do Excel
                        sheet.delete_rows(row[0].row)
                        
                        break
                        
                workbook.save(caminho_arquivo)
                
                
                
                # Exibe uma mensagem de sucesso
                messagebox.showinfo("Excluir Produto", "Produto excluido com sucesso!", parent=janela_cadastrar_produtos)
                
                #Limpa os campos de entrada
                limpa_campos()
                
                #Preenche o Treeview com os clientes existentes e alterados
                carregar_produtos_do_excel()

                    

            except FileNotFoundError:

                # Exibe uma mensagem de aviso se o arquivo não for encontrado
                messagebox.showwarning("Excluir Produto", "Base Dados.xlsx não encontrado!", parent=janela_cadastrar_produtos)

        
        
        else:
            
            # Exibe uma mensagem de aviso se o arquivo não for encontrado
            messagebox.showwarning("Excluir Produto", "Preencha o campo Codigo!", parent=janela_cadastrar_produtos)

    
    # Cria o botão de alterar cliente
    excluir_botao = tk.Button(janela_cadastrar_produtos,
                               text="Excluir",
                               command=excluir_produto,
                               font="Arial 14")
    excluir_botao.pack(side=tk.LEFT, padx=5, pady=3)
    
    def limpa_campos():
        
        codigo_entry.delete(0, tk.END)
        categoria_combobox.delete(0, tk.END)
        nome_entry.delete(0, tk.END)
        preco_entry.delete(0, tk.END)
                
    def salvar_produtos_excel():
        
        #Caminho do arquivo Excel
        caminho_arquivo = r"C:\python_projetos\python_rpa_projetos\tk_controle_pedido\Base+Dados.xlsx"
        
        #Nome da planilha
        planilha_nome = "Produtos"
        
        try:
            
            #Tenta carregar o arquivo existente
            workbook = load_workbook(filename=caminho_arquivo)
            
        except FileNotFoundError:
            
            print("Arquivo não existe")
            
        #Verifica se a planilha Produtos já existe no arquivo
        if planilha_nome not in workbook.sheetnames:
            
            #Cria a planilha Produtos
            workbook.create_sheet(planilha_nome)
            
        #Obtém a referência para a planilha
        sheet = workbook[planilha_nome]
            
        #Percorre os Produtos no dicionário
        for produto in Produtos.values():
            
            codigo = produto['Codigo']
            categoria = produto['Categoria']
            nome = produto['Nome']
            preco = produto['Preço']
            
            #Verifica se o cliente já existe na planilha
            produto_existe = False
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
            
                if row[0] == codigo:
                    
                    # Se o produto já existe, marca como encontrado e interrompe o loop
                    produto_existe = True
                    
                    break
                    
            # Se o produto não existe, adiciona na planilha
            if not produto_existe:

                # Obtém a próxima linha disponível na planilha
                max_row = sheet.max_row + 1
                
                #Insere os dados do cliente nas colunas correspondentes
                sheet.cell(row=max_row, column=1, value=codigo)
                sheet.cell(row=max_row, column=2, value=categoria)
                sheet.cell(row=max_row, column=3, value=nome)
                sheet.cell(row=max_row, column=4, value=preco)
                
        
        # Salva as alterações no arquivo
        workbook.save(filename=caminho_arquivo)
            
    
    # Atualiza a interface gráfica para processar todos os eventos pendentes
    janela_cadastrar_produtos.update_idletasks()

    # Obtém a largura atual da janela
    largura = janela_cadastrar_produtos.winfo_width()

    # Obtém a altura atual da janela
    altura = janela_cadastrar_produtos.winfo_height()

    # Calcula a posição x para centralizar a janela na tela
    posicao_x = (janela_cadastrar_produtos.winfo_screenwidth() // 2) - (largura // 2)

    # Calcula a posição y para centralizar a janela na tela
    posicao_y = (janela_cadastrar_produtos.winfo_screenheight() // 2) - (altura // 2)

    # Define a geometria da janela com a largura, altura, posição x e posição y
    janela_cadastrar_produtos.geometry('{}x{}+{}+{}'.format(largura, altura, posicao_x, posicao_y))


def selecionar_cliente():
    
    janela_selecionar_cliente = tk.Toplevel(root)
    
    # Define o título da janela
    janela_selecionar_cliente.title("Selecionar Clientes")

    # Configura o fundo da janela como branco
    janela_selecionar_cliente.configure(bg="white")
    
    #Lista para armazenar os dados dos clientes
    clientes = []
    
    #Caminho do arquivo Excel
    caminho_arquivo = r"C:\python_projetos\python_rpa_projetos\tk_controle_pedido\Base+Dados.xlsx"
        
    #Nome da planilha
    planilha_nome = "Clientes"
        
    try:
            
        # Carrega o arquivo Excel
        workbook = load_workbook(filename=caminho_arquivo)
        
        # Acessa a planilha "Clientes"
        sheet = workbook[planilha_nome]

        # Itera pelas linhas da planilha a partir da segunda linha (ignorando o cabeçalho)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            
            # Extrai os dados da linha (CPF, sexo, nome, aniversario, email)
            cpf, sexo, nome, aniversario, email = row
            
            # Cria um dicionário com os dados do cliente e adiciona à lista de clientes
            cliente = {'CPF': cpf, 'Sexo': sexo, 'Nome': nome, 'Aniversário': aniversario, 'Email': email}
            clientes.append(cliente)
        
    except FileNotFoundError:
        
        # Se o arquivo não for encontrado, exibe uma mensagem de aviso
        messagebox.showwarning("Carregar Clientes", "Arquivo Base Dados.xlsx não encontrado!")
        
        
    def filtrar_clientes():
        
        #Obtém o texto digitado no campo de filtro
        #lower - converte as letras para minúsculas
        filtro = filtro_entry.get().lower()
        
        #Limpa a Treeview
        clientes_treeview.delete(*clientes_treeview.get_children())
        
        # Filtra os clientes com base no nome digitado
        """
        clientes_filtrados: É a nova lista que irá armazenar os clientes filtrados, ou seja, os 
        clientes que atendem ao critério definido.

        for cliente in clientes: É a parte do loop da list comprehension. Ele itera através de 
        cada cliente presente na lista de clientes original.

        if filtro in cliente["Nome"].lower(): É a condição de filtro que determina se o cliente 
        deve ser incluído na lista de clientes filtrados. Neste caso, estamos verificando se o conteúdo 
        da variável filtro está presente no nome do cliente, ignorando maiúsculas e minúsculas (comparação em 
        letras minúsculas).
        
        filtro: É a variável que contém o termo que desejamos buscar no nome do cliente.
        
        cliente["Nome"]: Acessa o valor associado à chave "Nome" no dicionário do cliente 
        atual (nome do cliente).
        
        .lower(): Converte o nome do cliente para letras minúsculas. Isso é feito para garantir 
        que a busca não seja sensível a maiúsculas e minúsculas, tornando a filtragem não 
        case sensitive.

        """
        clientes_filtrados = [cliente for cliente in clientes if filtro in cliente["Nome"].lower()]
        
        #Adiciona os clientes filtrados a Treeview
        for cliente in clientes_filtrados:
            
                clientes_treeview.insert("", "end", values=(cliente["CPF"], cliente["Sexo"], cliente["Nome"], cliente["Aniversário"], cliente["Email"] ))
        
        
    
    if len(clientes) == 0:
        
        mensagem_label = tk.Label(janela_selecionar_cliente, 
                                  text = "Nenhum cliente cadastrado.",
                                  font = "Arial 12",
                                  bg="#FFFFFF")
        mensagem_label.pack(pady=10)
        
    else:
        
        # Se há clientes cadastrados, exibe a lista de clientes
        clientes_label = tk.Label(janela_selecionar_cliente, 
                                  text = "Clientes Cadastrados",
                                  font = "Arial 12",
                                  bg="#FFFFFF")
        clientes_label.pack(pady=10)
        
        # Cria um Frame para o campo de filtro e o botão "Abrir Menu de Produtos"
        frame_filtro = ttk.Frame(janela_selecionar_cliente)
        frame_filtro.pack(pady=10)

        # Cria um campo de entrada para o filtro
        filtro_label = tk.Label(frame_filtro, text="Filtrar por Nome:", font=("Arial", 12), bg="#FFFFFF")
        filtro_label.pack(side="left")
        filtro_entry = tk.Entry(frame_filtro, font=("Arial", 12))
        filtro_entry.pack(side="left", padx=5)
        filtrar_button = tk.Button(frame_filtro, text="Filtrar", font=("Arial", 12), command=filtrar_clientes)
        filtrar_button.pack(side="left")
        
        #Cria uma Treeview para exibir os dados dos clientes
        clientes_treeview = ttk.Treeview(janela_selecionar_cliente, columns=("CPF", "Sexo", "Nome", "Aniversário", "Email"), show="headings")
        clientes_treeview.heading("CPF", text="CPF")
        clientes_treeview.heading("Sexo", text="Sexo")
        clientes_treeview.heading("Nome", text="Nome")
        clientes_treeview.heading("Aniversário", text="Aniversário")
        clientes_treeview.heading("Email", text="Email")
        
        #Cria um scroll vertical
        scrollbar = ttk.Scrollbar(janela_selecionar_cliente, orient="vertical",command=clientes_treeview.yview)
        scrollbar.pack(side='right', fill='y')
        
        #Configura a Treeview para atualizar o Scroolbaer sempre que
        #os dados da Treeview forem alterados
        clientes_treeview.configure(yscrollcommand=scrollbar.set)
        clientes_treeview.pack()
        
        #Adiciona os dados dos clientes à Treeview
        for cliente in clientes:
            
            clientes_treeview.insert("", "end", values=(cliente["CPF"], cliente["Sexo"], cliente["Nome"], cliente["Aniversário"], cliente["Email"] ))
        
        clientes_treeview.column("Nome", width=200) #Ajusta a largura da coluna conforme desejado
    
        
        def selecionar():
            
            #Obtém o item selecionado na Treeview
            item_selecionado = clientes_treeview.focus()
            
            if item_selecionado:
                
                #Obtém os valores do item selecionado
                valores = clientes_treeview.item(item_selecionado)["values"]
                
                cliente_selecionado = {
                    "CPF": valores[0],
                    "Sexo": valores[1],
                    "Nome": valores[2],
                    "Aniversário": valores[3],
                    "Email": valores[4]
                }
                
                exibir_menu_produtos(cliente_selecionado)
                
        
        selecionar_botao = tk.Button(janela_selecionar_cliente, 
                                     text="Abrir Menu de Produtos", 
                                     font=("Arial", 12),
                                     command=selecionar)
        selecionar_botao.pack(pady=10)
    
    # Atualiza a interface gráfica para processar todos os eventos pendentes
    janela_selecionar_cliente.update_idletasks()

    # Obtém a largura atual da janela
    largura = janela_selecionar_cliente.winfo_width()

    # Obtém a altura atual da janela
    altura = janela_selecionar_cliente.winfo_height()

    # Calcula a posição x para centralizar a janela na tela
    posicao_x = (janela_selecionar_cliente.winfo_screenwidth() // 2) - (largura // 2)

    # Calcula a posição y para centralizar a janela na tela
    posicao_y = (janela_selecionar_cliente.winfo_screenheight() // 2) - (altura // 2)

    # Define a geometria da janela com a largura, altura, posição x e posição y
    janela_selecionar_cliente.geometry('{}x{}+{}+{}'.format(largura, altura, posicao_x, posicao_y))
    

def exibir_menu_produtos(cliente):
    
    
    janela_menu_produtos = tk.Toplevel(root)
    
    # Define o título da janela
    janela_menu_produtos.title("Menu de Produtos")

    # Configura o fundo da janela como branco
    janela_menu_produtos.configure(bg="white")
    
    
    #Criando um label para exibir as informações do cliente
    if "CPF" in cliente and "Nome" in cliente:
        
        cliente_label = tk.Label(janela_menu_produtos,
                                text="CPF: {}  -  Nome: {}".format(cliente["CPF"], cliente["Nome"]),
                                font="Arial 16",
                                bg="#FFFFFF")
        cliente_label.pack(pady=10)
        
    else:
        
        cliente_label = tk.Label(janela_menu_produtos,
                                text="Cliente Desconhecido",
                                font="Arial 16",
                                bg="#FFFFFF")
        cliente_label.pack(pady=10)
        
    #Cria uma lista vazia para o carrinho
    carrinho = []
    
    #Cria uma lista vazia para os produtos
    produtos = []
    
    def carregar_produtos_do_excel():
        
        #Limpa a lista de produtos antes de carregar novos produtos
        produtos.clear()
        
        #Caminho do arquivo Excel
        caminho_arquivo = r"C:\python_projetos\python_rpa_projetos\tk_controle_pedido\Base+Dados.xlsx"

        #Nome da planilha
        planilha_nome = "Produtos"
        
        try:
            
            # Carrega o arquivo Excel
            workbook = load_workbook(filename=caminho_arquivo)
            
            #Acessa a planilha "Produtos" dentro do arquivo de excel
            sheet = workbook[planilha_nome]
            
            # Itera através das linhas da planilha, começando a partir da segunda linha (min_row=2)
            for row in sheet.iter_rows(min_row=2, values_only=True):

                # Extrai as informações de cada coluna da planilha para variáveis
                codigo_produto, categoria, nome, preco = row

                # Cria um dicionário contendo as informações do produto e o adiciona à lista de produtos
                produto = {"Codigo": codigo_produto, "Categoria": categoria, "Nome": nome, "Preço": float(preco)}
                produtos.append(produto)
                
                #Chama a função para atualizar a lista de produtos no programa
                atualizar_lista_produtos()
            
        except FileNotFoundError:
            
            # Exibe uma mensagem de aviso se o arquivo não for encontrado
            messagebox.showwarning("Carregar Clientes", "Arquivo Clientes.xlsx não encontrado!", parent=janela_cadastrar_cliente)
            
    
    
    #Cria um campo label para exibir as informações
    info_label = tk.Label(janela_menu_produtos,
                         text="",
                         font="Arial 12",
                         bg="#FFFFFF")
    info_label.pack(pady=10)
    
    frame_principal = tk.Frame(janela_menu_produtos, bg="#FFFFFF")
    frame_principal.pack(pady=10)
    
    #Cria um subframe chamado "menu_frame" dentro do "frame_principal"
    menu_frame = tk.Frame(frame_principal, bg="#FFFFFF")
    menu_frame.pack(side=tk.LEFT, padx=10)
    
    #Cria um campo label para exibir as informações
    produtos_label = tk.Label(janela_menu_produtos,
                         text="Produtos Disponíveis",
                         font="Arial 12",
                         bg="#FFFFFF")
    produtos_label.pack(pady=10)
    
    #Cria uma listbox chamada "menu_listbox" dentro do "menu_frame" e
    #Associa à listbox "menu_listbox"
    menu_listbox = tk.Listbox(menu_frame, font="Arial 12")
    menu_listbox.pack(side=tk.LEFT)
    
    
    """
    cria uma barra de rolagem vertical para a listbox menu_listbox, permitindo que o usuário 
    possa navegar pelos itens da lista caso ela não caiba completamente na janela. 
    
    Vamos explicar cada linha:

    scrollbar_menu = tk.Scrollbar(menu_frame): Cria uma scrollbar vertical chamada scrollbar_menu, 
    associada ao frame menu_frame.

    scrollbar_menu.pack(side=tk.RIGHT, fill=tk.Y): Posiciona a scrollbar na lateral direita do 
    menu_frame (side=tk.RIGHT) e preenche o espaço vertical do frame (fill=tk.Y). Isso fará com 
    que a scrollbar seja exibida no lado direito e tenha a altura correspondente ao espaço disponível no frame.

    menu_listbox.config(yscrollcommand=scrollbar_menu.set): Configura a menu_listbox para usar a scrollbar 
    scrollbar_menu como sua barra de rolagem vertical. Isso significa que quando o usuário rolar a barra 
    de rolagem, a listbox será atualizada para exibir os itens correspondentes.

    scrollbar_menu.config(command=menu_listbox.yview): Configura a barra de rolagem para chamar o 
    método yview() da listbox quando a posição da barra de rolagem for alterada. Isso permite que a 
    barra de rolagem controle a visualização da listbox conforme o usuário a desloca para cima ou para baixo.
    """
    scrollbar_menu = tk.Scrollbar(menu_frame)
    scrollbar_menu.pack(side=tk.RIGHT, fill=tk.Y)
    menu_listbox.config(yscrollcommand=scrollbar_menu.set)
    scrollbar_menu.config(command=menu_listbox.yview)
    
    # Função para atualizar o carrinho na interface gráfica
    def atualizar_carrinho():

        # Limpa a listbox do carrinho antes de atualizá-la
        carrinho_listbox.delete(0, tk.END)

        # Itera através dos itens no carrinho
        for item in carrinho:
            
            # Insere cada item na listbox do carrinho na forma "Nome do Produto - R$ Preço"
            carrinho_listbox.insert(tk.END, "{} - R$ {:.2f}".format(item["Nome"], item["Preço"]))

        # Atualiza o rótulo do total do carrinho com o valor total calculado
        total_label.config(text="Total: R$ {:.2f}".format(calcular_total()))


    # Função para calcular o valor total dos itens no carrinho
    def calcular_total():

        # Calcula o valor total somando os preços de cada item no carrinho
        total = sum(item["Preço"] for item in carrinho)
        
        return total
    
    def adicionar_carrinho():
        
        #Obtém o índice do item selecionado na listbox de produtos
        indice = menu_listbox.curselection()
        
        # Verifica se foi selecionado algum item na listbox
        if indice:
            
            # Obtém o produto correspondente ao índice selecionado
            produto = produtos[indice[0]]

            # Obtém a quantidade digitada pelo usuário no Entry
            quantidade = quantidade_entry.get()

            # Verifica se a quantidade é um número inteiro
            if quantidade.isnumeric():
                
                # Converte a quantidade para inteiro
                quantidade = int(quantidade)

                # Adiciona o produto várias vezes ao carrinho de acordo com a quantidade especificada
                carrinho.extend([produto] * quantidade)

                # Remove o produto da lista de produtos
                produtos.pop(indice[0])

                # Exibe uma mensagem indicando que o(s) produto(s) foi(ram) adicionado(s) ao carrinho
                info_label.config(text="Produto(s) adicionado(s) ao carrinho.")

                # Limpa o Entry de quantidade
                quantidade_entry.delete(0, tk.END)

                # Define o foco para o Entry de quantidade para uma nova digitação
                quantidade_entry.focus()

                # Chama as funções para atualizar a listbox de carrinho e a listbox de produtos
                atualizar_carrinho()
                atualizar_lista_produtos()
                
            else:
                
                # Exibe uma mensagem de erro caso a quantidade digitada não seja um número inteiro
                info_label.config(text="Digite uma quantidade válida (número inteiro).")
                
    
    # Cria um botão "adicionar_button" dentro do "menu_frame" para adicionar produtos ao carrinho, associando-o à função "adicionar_carrinho"
    adicionar_botao = tk.Button(menu_frame,
                               text="Adicionar ao Carrinho",
                               font="Aria 14",
                               command = adicionar_carrinho)
    adicionar_botao.pack(pady=5)
    
    # Função para remover um produto do carrinho
    def remover_carrinho():
        
        # Obtém o índice do item selecionado na listbox do carrinho
        indice = carrinho_listbox.curselection()

        # Verifica se foi selecionado algum item na listbox do carrinho
        if indice:
            
            # Obtém o produto correspondente ao índice selecionado na listbox do carrinho
            produto = carrinho[indice[0]]

            # Remove o produto da lista do carrinho
            carrinho.remove(produto)

            # Verifica se o produto já está na lista de produtos
            if produto not in produtos:
                
                # Caso não esteja, adiciona o produto de volta à lista de produtos
                produtos.append(produto)

            # Exibe uma mensagem indicando que o produto foi removido do carrinho
            info_label.config(text="Produto removido do carrinho.")

            # Atualiza a listbox do carrinho e a listbox de produtos
            atualizar_carrinho()
            atualizar_lista_produtos()  
            
        else:
            
            # Exibe uma mensagem de erro caso nenhum produto tenha sido selecionado na listbox do carrinho
            info_label.config(text="Selecione um produto para remover do carrinho.")
            
            
    def finalizar_pedido(cliente):
        
        # Verifica se há produtos no carrinho
        if len(carrinho) == 0:
            
            messagebox.showwarning("Finalizar Pedido", "O carrinho está vazio. Adicione produtos antes de finalizar o pedido.")
            
            return

        # Obtém o CPF e Nome do cliente selecionado
        cpf_cliente = cliente["CPF"]
        nome_cliente = cliente["Nome"]

        # Abre o arquivo Excel
        caminho_arquivo = r"C:\python_projetos\python_rpa_projetos\tk_controle_pedido\Base+Dados.xlsx"
        workbook = load_workbook(filename=caminho_arquivo)
        
        # Obtém a planilha "Pedidos"
        planilha_pedidos = workbook["Pedidos"]
        
        # Encontra a próxima linha vazia na planilha
        proxima_linha = planilha_pedidos.max_row + 1

        # Salva os dados dos produtos no carrinho
        for produto in carrinho:
            
            codigo_produto = produto["Codigo"]
            categoria = produto["Categoria"]
            nome_produto = produto["Nome"]
            preco = produto["Preço"]

            # Salva os dados do cliente
            planilha_pedidos.cell(row=proxima_linha, column=1, value=cpf_cliente)
            planilha_pedidos.cell(row=proxima_linha, column=2, value=nome_cliente)

            # Salva os dados do produto na planilha
            planilha_pedidos.cell(row=proxima_linha, column=3, value=codigo_produto)
            planilha_pedidos.cell(row=proxima_linha, column=4, value=categoria)
            planilha_pedidos.cell(row=proxima_linha, column=5, value=nome_produto)
            planilha_pedidos.cell(row=proxima_linha, column=6, value=preco)

            # Incrementa o número da próxima linha
            #proxima_linha = proxima_linha + 1
            proxima_linha += 1
            
        # Salva as alterações no arquivo Excel
        workbook.save(caminho_arquivo)

        # Exibe uma mensagem de sucesso
        messagebox.showinfo("Finalizar Pedido", "Pedido salvo com sucesso.")
        
        # Limpa o carrinho e atualiza a interface
        carrinho.clear()
        atualizar_carrinho()
        atualizar_lista_produtos()
        quantidade_entry.focus()  # Retorna o foco para o campo de quantidade
        
        janela_menu_produtos.destroy()
    
    
    # Cria um subframe chamado "quantidade_frame" dentro do "menu_frame" para a entrada da quantidade
    quantidade_frame = tk.Frame(menu_frame)
    quantidade_frame.pack(pady=5)
    
    #Cria um campo label para exibir as informações
    quantidade_label = tk.Label(quantidade_frame,
                         text="Quantidade",
                         font="Arial 12",
                         bg="#FFFFFF")
    quantidade_label.pack(side=tk.LEFT)
    
    # Cria um Entry chamado "quantidade_entry" dentro do "quantidade_frame" para digitar a quantidade de produtos a adicionar ao carrinho
    quantidade_entry = tk.Entry(quantidade_frame,
                         font="Arial 12",
                         width=5)
    quantidade_entry.pack(side=tk.LEFT)
    quantidade_entry.focus() #Define o foco inicial para o Entry
    
    #----------------------------------------------------------------------
    
    # Cria um subframe chamado "carrinho_frame" dentro do "frame_principal" para exibir os itens no carrinho
    carrinho_frame = tk.Frame(frame_principal, bg="#FFFFFF")
    carrinho_frame.pack(side=tk.LEFT, padx=10)
    
    # Cria um label para mostrar o texto "Carrinho:" dentro do "carrinho_frame"
    carrinho_label = tk.Label(carrinho_frame,
                         text="Carrinho",
                         font="Arial 12",
                         bg="#FFFFFF")
    carrinho_label.pack(pady=10)
    
    
    # Cria uma listbox chamada "carrinho_listbox" dentro do 
    #"carrinho_frame" para exibir os itens do carrinho
    carrinho_listbox = tk.Listbox(carrinho_frame, font="Arial 12")
    carrinho_listbox.pack(side=tk.LEFT)
    
    # Cria uma scrollbar chamada "scrollbar_carrinho" dentro do 
    #"carrinho_frame" e a associa à listbox "carrinho_listbox"
    scrollbar_carrinho = tk.Scrollbar(carrinho_frame)
    scrollbar_carrinho.pack(side=tk.RIGHT, fill=tk.Y)
    carrinho_listbox.config(yscrollcommand=scrollbar_carrinho.set)
    scrollbar_carrinho.config(command=carrinho_listbox.yview)
    
    # Cria um label "total_label" dentro do "carrinho_frame" 
    #para exibir o valor total do carrinho
    total_label = tk.Label(carrinho_frame,
                         text="Total: R$ 0.00",
                         font="Arial 12",
                         bg="#FFFFFF")
    total_label.pack(pady=10)
    
    
    # Cria um botão "remover_carrinho_botao" dentro do "carrinho_frame" para remover os produtos do carrinho
    remover_carrinho_botao = tk.Button(carrinho_frame,
                               text="Remover do Carrinho",
                               font="Aria 14",
                               command = remover_carrinho)
    remover_carrinho_botao.pack(pady=5)
    
    
    
    # Cria um botão "finalizar_botao" dentro da janela "janela_menu_produtos" para finalizar o pedido, associando-o à função lambda que recebe o cliente como parâmetro
    finalizar_botao = tk.Button(janela_menu_produtos,
                               text="Finalizar Pedido",
                               font="Aria 14",
                               command = lambda: finalizar_pedido(cliente))
    finalizar_botao.pack(pady=5)
    
    def atualizar_lista_produtos():
        
        menu_listbox.delete(0, tk.END)
        for produto in produtos:
            
            menu_listbox.insert(tk.END, "{} - R$ {:.2f}".format(produto["Nome"], produto["Preço"]))
            
        
    #Chama a função para carregar os produtos
    carregar_produtos_do_excel()
    atualizar_lista_produtos()
    
    
    # Atualiza a interface gráfica para processar todos os eventos pendentes
    janela_menu_produtos.update_idletasks()

    # Obtém a largura atual da janela
    largura = janela_menu_produtos.winfo_width()

    # Obtém a altura atual da janela
    altura = janela_menu_produtos.winfo_height()

    # Calcula a posição x para centralizar a janela na tela
    posicao_x = (janela_menu_produtos.winfo_screenwidth() // 2) - (largura // 2)

    # Calcula a posição y para centralizar a janela na tela
    posicao_y = (janela_menu_produtos.winfo_screenheight() // 2) - (altura // 2)

    # Define a geometria da janela com a largura, altura, posição x e posição y
    janela_menu_produtos.geometry('{}x{}+{}+{}'.format(largura, altura, posicao_x, posicao_y))
    

def ler_pedidos_excel():
    
    # Cria a janela para exibir os pedidos
    janela_pedidos = tk.Toplevel(root)
    janela_pedidos.title("Pedidos")
    janela_pedidos.configure(bg="white")

    # Define o estilo para o Treeview
    estilo_treeview = ttk.Style()
    estilo_treeview.configure("Treeview", font=("Arial", 14))

    # Cria uma Treeview para exibir os pedidos de cada cliente
    treeview_pedidos = ttk.Treeview(janela_pedidos, columns=("CPF", "Pedido", "Total"), show="tree")
    treeview_pedidos.heading("#0", text="Cliente", anchor="w")
    treeview_pedidos.heading("CPF", text="CPF", anchor="w")
    treeview_pedidos.heading("Pedido", text="Pedido", anchor="w")
    treeview_pedidos.heading("Total", text="Total", anchor="w")
    
    # Aumenta a largura das colunas
    treeview_pedidos.column("#0", width=400, minwidth=400, anchor="w")
    treeview_pedidos.column("CPF", width=120, minwidth=120, anchor="w")
    treeview_pedidos.column("Pedido", width=400, minwidth=400, anchor="w")
    treeview_pedidos.column("Total", width=100, minwidth=100, anchor="w")
    
    treeview_pedidos.pack(expand=True, fill="both")
    
    #Caminho do arquivo Excel
    caminho_arquivo = r"C:\python_projetos\python_rpa_projetos\tk_controle_pedido\Base+Dados.xlsx"
        
    #Nome da planilha
    planilha_nome = "Pedidos"
        
    try:
            
        # Carrega o arquivo Excel
        workbook = load_workbook(filename=caminho_arquivo)
        planilha_pedidos = workbook[planilha_nome]
            
        # Dicionário para armazenar os pedidos de cada cliente
        pedidos_por_cliente = {}
        
        # Agrupa os pedidos por cliente        
        for row in planilha_pedidos.iter_rows(min_row=2, values_only=True):
            
            cpf, nome, codigo_produto, categoria, nome_produto, preco = row
            pedido = f"{codigo_produto} - Produto: {nome_produto}"

            # Verifica se o cliente já possui pedidos registrados
            if cpf in pedidos_por_cliente:
                
                pedidos_por_cliente[cpf].append((pedido, preco))
            
            else:
                
                pedidos_por_cliente[cpf] = [(pedido, preco)]
                
                
        # Exibe os pedidos na Treeview
        for cpf, pedidos in pedidos_por_cliente.items():
            
            # Variável para armazenar o nome do cliente atual
            nome_cliente = None
            
            # Variável para calcular o total do valor dos pedidos do cliente atual
            total_pedido = 0

            # Itera sobre as linhas da planilha de pedidos
            for row in planilha_pedidos.iter_rows(min_row=2, values_only=True):
                
                # Recebe os valores de cliente_cpf e cliente_nome a partir da linha atual da planilha
                cliente_cpf, cliente_nome, *_ = row
                
                # Verifica se o cliente_cpf da linha atual corresponde ao cpf do cliente atual do loop externo
                if cliente_cpf == cpf:
                    
                    # Se o CPF for correspondente, armazena o nome do cliente atual
                    nome_cliente = cliente_nome
                    
                    # Encerra o loop, pois já encontrou o cliente correspondente
                    break
                    
            # Insere o item do cliente
            cliente_item_id = treeview_pedidos.insert("", "end", text=f"{nome_cliente} - {cpf}", values=("", "", ""))

            # Insere os pedidos do cliente (caso existam)
            if cliente_item_id:
                
                for pedido in pedidos:
                    
                    produto, preco = pedido
                    
                    # Insere o pedido na Treeview e calcula o total do pedido
                    treeview_pedidos.insert(cliente_item_id, "end", text=produto, values=("", f"R${preco:.2f}"))
                    total_pedido += preco

                # Insere o total do pedido do cliente no final
                treeview_pedidos.item(cliente_item_id, values=("", "Total:", f"R${total_pedido:.2f}"))

    except FileNotFoundError:
        messagebox.showwarning("Exibir Pedidos", "Arquivo Base Dados.xlsx não encontrado!")
     
    #Fechar a janela de pedidos
    fechar_botao = tk.Button(janela_pedidos,
                            text="Fechar",
                            font="Arial 14",
                            command=janela_pedidos.destroy)
    fechar_botao.pack(pady=10)
    
    def salvar_pedidos_encerrados():
    
        # Obtém o item selecionado na Treeview
        item_selecionado = treeview_pedidos.selection()

        if item_selecionado:
            
            # Obtém os valores do item selecionado
            cliente_info = treeview_pedidos.item(item_selecionado, "text")
            #Tulio Souza - 849.353.278-49
            
            """
            Essa linha de código utiliza o método split() na string cliente_info para dividir essa 
            string em três partes, com base no separador " - ". Em seguida, ela atribui cada parte a 
            três variáveis: codigo_produto, cliente_nome e cliente_cpf.

            cliente_info: Essa é uma string que contém informações sobre um cliente. Ela está no formato: 
            "Nome do Cliente - CPF do Cliente".

            " - ": Esse é o separador utilizado para dividir a string cliente_info.
            

            split(" - ", 2): Esse método é usado para dividir a string cliente_info 
            em três partes utilizando " - " como separador. O argumento 2 é opcional e especifica o
            número máximo de divisões a serem feitas. Nesse caso, ele significa que serão feitas no 
            máximo duas divisões, resultando em três partes.

            Após a divisão, os valores são atribuídos às variáveis da seguinte forma:

            codigo_produto: Essa variável irá receber a primeira parte da divisão, que 
            corresponde ao "Nome do Cliente".

            cliente_nome: Essa variável irá receber a segunda parte da divisão, que corresponde 
            ao "CPF do Cliente".

            cliente_cpf: Essa variável irá receber a terceira parte da divisão, que estará vazia nesse 
            caso, já que existem apenas duas partes.

            Em resumo, essa linha de código pega as informações armazenadas em cliente_info, que 
            estão no formato "Nome do Cliente - CPF do Cliente", e separa o nome e o CPF, armazenando-os 
            em variáveis diferentes: codigo_produto e cliente_nome, respectivamente. A terceira variável, 
            cliente_cpf, ficará vazia nesse caso específico. Os nomes das variáveis podem não estar bem 
            escolhidos, considerando o contexto, mas isso depende de como cliente_info está 
            estruturado no Treeview.
            """
            cliente_nome, cliente_cpf = cliente_info.split(" - ", 1)
            
                        
            # Remove o número antes do nome do produto
            """
            Essa linha de código realiza duas operações em sequência na variável cliente_nome:

            cliente_nome.split(" - ", 1): Nessa etapa, a string cliente_nome é dividida em duas 
            partes usando o separador " - " apenas uma vez. O método split() é utilizado para realizar 
            essa divisão, e o segundo argumento 1 especifica o número máximo de divisões a serem feitas. 
            Nesse caso, ele indica que apenas uma divisão deve ocorrer, resultando em duas partes.

            [0]: Após a divisão, esse trecho de código seleciona o primeiro elemento da lista resultante 
            da operação de divisão. Como a divisão foi feita apenas uma vez, a lista terá dois elementos, e 
            o [0] seleciona o primeiro elemento, que corresponde à primeira parte da divisão, ou seja, o 
            "Nome do Cliente".

            Portanto, o resultado dessa linha de código é atribuir à variável cliente_nome somente o "Nome 
            do Cliente", eliminando qualquer informação adicional após o primeiro " - " encontrado na string 
            original.
            
            """
            cliente_nome = cliente_nome.split(" - ", 1)[0]   
            
            #Pega os pedidos do ní filho que é os pedidos do cliente selecionado
            pedidos  = treeview_pedidos.get_children(item_selecionado)
            
            data_atual = date.today().strftime("%d/%m/%Y")
            
            #Caminho do arquivo Excel
            caminho_arquivo = r"C:\python_projetos\python_rpa_projetos\tk_controle_pedido\Base+Dados.xlsx"
            
            # Carrega o arquivo Excel
            workbook = load_workbook(filename=caminho_arquivo)
            
            #Nome da planilha
            planilha_pedidos = workbook["Pedidos"]
            planilha_encerrados = workbook["Encerrados"]
            
            #Encontra a próxima linha vazia da planilha "Encerrados"
            proxima_linha_encerrados = planilha_encerrados.max_row + 1

            for pedido_id in pedidos:
                
                produto_str = treeview_pedidos.item(pedido_id, "text")
                preco_str = treeview_pedidos.item(pedido_id, "values")[1]
                preco = float(preco_str.replace("R$", "").replace(",", "."))
                
                #Remove o número antes do nome do produto
                produto = produto_str.split(": ", 1)[1]
                
                planilha_encerrados.cell(row=proxima_linha_encerrados, column=1, value=cliente_cpf)
                planilha_encerrados.cell(row=proxima_linha_encerrados, column=2, value=cliente_nome)
                planilha_encerrados.cell(row=proxima_linha_encerrados, column=3, value=produto)
                planilha_encerrados.cell(row=proxima_linha_encerrados, column=4, value=preco)
                planilha_encerrados.cell(row=proxima_linha_encerrados, column=5, value=data_atual)                
                proxima_linha_encerrados += 1
                
                #Remove o pedido da Treview
                treeview_pedidos.delete(pedido_id)
                
            # Exclui os pedidos da planilha "Pedidos"
            # Cria uma lista para armazenar os números das linhas que serão excluídas da planilha
            linhas_a_excluir = []

            # Itera sobre as linhas da planilha de pedidos começando a partir da segunda linha (min_row=2)
            # A função enumerate retorna tanto o número da linha (row_num) quanto os valores da linha (row)
            for row_num, row in enumerate(planilha_pedidos.iter_rows(min_row=2, values_only=True), start=2):
                
                # Verifica se o valor da primeira coluna da linha (CPF) é igual ao CPF do cliente fornecido
                # A função str() é usada para garantir que ambos os valores sejam comparados como strings
                if str(row[0]) == str(cliente_cpf):
                    
                    # Se o CPF da linha atual for igual ao CPF do cliente fornecido, adiciona o número da linha à lista de linhas a excluir
                    linhas_a_excluir.append(row_num)
                    
            #Remove as linhas da planilha de pedidos
            for numero_linha in reversed(linhas_a_excluir):
                
                planilha_pedidos.delete_rows(numero_linha)
                
            
            # Salva as alterações no arquivo Excel
            workbook.save(caminho_arquivo)

            # Exibe uma mensagem de sucesso
            messagebox.showinfo("Pedido Encerrado", f"O pedido do cliente {cliente_nome} foi salvo na planilha 'Encerrados'.")
         
                
                
    
    #Cria um botão para encerrar o pedido selecionado
    encerrar_pedidos_botao = tk.Button(janela_pedidos,
                            text="Encerrar Pedidos",
                            font="Arial 14",
                            command = salvar_pedidos_encerrados)
    encerrar_pedidos_botao.pack(pady=10)
                
    
    # Atualiza a interface gráfica para processar todos os eventos pendentes
    janela_pedidos.update_idletasks()

    # Obtém a largura atual da janela
    largura = janela_pedidos.winfo_width()

    # Obtém a altura atual da janela
    altura = janela_pedidos.winfo_height()

    # Calcula a posição x para centralizar a janela na tela
    posicao_x = (janela_pedidos.winfo_screenwidth() // 2) - (largura // 2)

    # Calcula a posição y para centralizar a janela na tela
    posicao_y = (janela_pedidos.winfo_screenheight() // 2) - (altura // 2)

    # Define a geometria da janela com a largura, altura, posição x e posição y
    janela_pedidos.geometry('{}x{}+{}+{}'.format(largura, altura, posicao_x, posicao_y))
    

# Variável global para armazenar os dados carregados do arquivo Excel
dados_encerrados = []
    
def carregar_dados_encerrados():
    
    # Define as variáveis de filtro como globais para que possam ser acessadas dentro da função
    global dados_encerrados, entry_cpf, entry_cliente, entry_codigo_produto, entry_produto, entry_preco, entry_data_inicio, entry_data_fim

    
    # Cria a janela para exibir os pedidos
    janela_encerrados = tk.Toplevel(root)
    janela_encerrados.title("Pedidos Encerrados")
    janela_encerrados.configure(bg="white")
    
    #Frame para o titulo da Treeview
    frame_titulo = tk.Frame(janela_encerrados, bg="white")
    frame_titulo.pack(pady=10)
    
    titulo = tk.Label(frame_titulo,
                     text="Pedidos Encerrados",
                     font="Arial 16",
                     bg="white")
    titulo.pack()

    # Define o estilo para o Treeview
    estilo_treeview = ttk.Style()
    estilo_treeview.configure("Treeview", font=("Arial", 14))

    # Cria uma Treeview para exibir os pedidos de cada cliente
    treeview_encerrados = ttk.Treeview(janela_encerrados, columns=("CPF", "Cliente", "Produto", "Preço", "Data"), show="headings")
    treeview_encerrados.heading("#0", text="Cliente", anchor="w")
    treeview_encerrados.heading("CPF", text="CPF", anchor="w")
    treeview_encerrados.heading("Cliente", text="Cliente", anchor="w")
    treeview_encerrados.heading("Produto", text="Produto", anchor="w")
    treeview_encerrados.heading("Preço", text="Preço", anchor="w")
    treeview_encerrados.heading("Data", text="Data", anchor="w")
    
    # Aumenta a largura das colunas
    treeview_encerrados.column("#0", width=300, minwidth=300, anchor="w")
    treeview_encerrados.column("CPF", width=120, minwidth=120, anchor="w")
    treeview_encerrados.column("Cliente", width=200, minwidth=200, anchor="w")
    treeview_encerrados.column("Produto", width=200, minwidth=200, anchor="w")
    treeview_encerrados.column("Preço", width=100, minwidth=100, anchor="w")
    treeview_encerrados.column("Data", width=100, minwidth=100, anchor="w")
    
    scrollbar = ttk.Scrollbar(janela_encerrados, orient="vertical", command=treeview_encerrados.yview)
    scrollbar.pack(side="right", fill="y")
    
    #Configura a Treeview para atualizar o Scrollbar
    treeview_encerrados.configure(yscrollcommand=scrollbar.set)
    
    treeview_encerrados.pack(expand=True, fill="both")
    
    #Caminho do arquivo Excel
    caminho_arquivo = r"C:\python_projetos\python_rpa_projetos\tk_controle_pedido\Base+Dados.xlsx"
        
    #Nome da planilha
    planilha_nome = "Encerrados"
        
    try:
        
        workbook = load_workbook(filename=caminho_arquivo)
        planilha_encerrados = workbook[planilha_nome]

        # Carrega os dados da planilha "Encerrados" na Treeview
        # Obtém os dados da planilha "Encerrados" a partir da segunda linha (min_row=2) e apenas os 
        # valores das células (values_only=True)
        dados_encerrados = list(planilha_encerrados.iter_rows(min_row=2, values_only=True))

        # Itera sobre os dados carregados da planilha
        for row in dados_encerrados:
            
            # Desempacota os valores da linha em variáveis individuais: cpf, cliente, produto, preco, data
            cpf, cliente, produto, preco, data = row

            # Insere os dados na Treeview
            # O parâmetro "" indica que os valores serão inseridos no nível raiz da Treeview
            # O parâmetro "end" indica que os valores serão adicionados no final da Treeview
            # O parâmetro text é utilizado para definir o texto que aparecerá na primeira coluna da 
            # Treeview (coluna "#0")
            # Os demais parâmetros (values) são utilizados para preencher as demais colunas da Treeview 
            # com os valores da linha
            treeview_encerrados.insert("", "end", text=cliente, values=(cpf, cliente, produto, f"R${preco:.2f}", data))

    except FileNotFoundError:
        
        messagebox.showwarning("Carregar Dados Encerrados", "Arquivo Base Dados.xlsx não encontrado!")
    
    #Criação dos campos de filtro
    frame_filtros = tk.Frame(janela_encerrados, bg="#FFFFFF")
    frame_filtros.pack(pady=10)
    
    #Lista com os rótulos dos campos
    labels = ["CPF", "Cliente", "Produto", "Preço", "Data (DD/MM/AAAA):"]
    entries = []
    
    # Criação dos campos de filtro individualmente e adicionando-os à lista 'entries'
    label_cpf = tk.Label(frame_filtros, text=labels[0], font=("Arial", 12), bg="white")
    label_cpf.grid(row=0, column=0, pady=5, padx=5, sticky="w")
    entry_cpf = tk.Entry(frame_filtros, font=("Arial", 12))
    entry_cpf.grid(row=0, column=1, pady=5, padx=5, sticky="ew")
    entries.append(entry_cpf)
    
    label_cliente = tk.Label(frame_filtros, text=labels[1], font=("Arial", 12), bg="white")
    label_cliente.grid(row=1, column=0, pady=5, padx=5, sticky="w")
    entry_cliente = tk.Entry(frame_filtros, font=("Arial", 12))
    entry_cliente.grid(row=1, column=1, pady=5, padx=5, sticky="ew")
    entries.append(entry_cliente)
    
    label_produto = tk.Label(frame_filtros, text=labels[2], font=("Arial", 12), bg="white")
    label_produto.grid(row=2, column=0, pady=5, padx=5, sticky="w")
    entry_produto = tk.Entry(frame_filtros, font=("Arial", 12))
    entry_produto.grid(row=2, column=1, pady=5, padx=5, sticky="ew")
    entries.append(entry_produto)
    
    label_preco = tk.Label(frame_filtros, text=labels[3], font=("Arial", 12), bg="white")
    label_preco.grid(row=3, column=0, pady=5, padx=5, sticky="w")
    entry_preco = tk.Entry(frame_filtros, font=("Arial", 12))
    entry_preco.grid(row=3, column=1, pady=5, padx=5, sticky="ew")
    entries.append(entry_preco)
    
    label_data = tk.Label(frame_filtros, text=labels[4], font=("Arial", 12), bg="white")
    label_data.grid(row=4, column=0, pady=5, padx=5, sticky="w")
    entry_data_inicio = tk.Entry(frame_filtros, font=("Arial", 12))
    entry_data_inicio.grid(row=4, column=1, pady=5, padx=5, sticky="ew")
    entries.append(entry_data_inicio)
    
    label_data_fim = tk.Label(frame_filtros, text="Data Fim (DD/MM/AAAA):", font=("Arial", 12), bg="white")
    label_data_fim.grid(row=5, column=0, pady=5, padx=5, sticky="w")
    entry_data_fim = tk.Entry(frame_filtros, font=("Arial", 12))
    entry_data_fim.grid(row=5, column=1, pady=5, padx=5, sticky="ew")
    
    # Função para executar o filtro
    def filtrar_dados():
        
        # Define as variáveis de filtro como globais para que possam ser acessadas dentro da função
        global entry_cpf, entry_cliente, entry_codigo_produto, entry_produto, entry_preco, entry_data_inicio, entry_data_fim

        # Obtém os valores dos campos de filtro (Entry) para cada critério de filtragem
        cpf_filtro = entry_cpf.get()
        cliente_filtro = entry_cliente.get()
        produto_filtro = entry_produto.get()
        preco_filtro = entry_preco.get()
        data_inicio_filtro = entry_data_inicio.get()
        data_fim_filtro = entry_data_fim.get()
        
        treeview_encerrados.delete(*treeview_encerrados.get_children())
        
        try:
            
            # Itera sobre os dados encerrados carregados anteriormente
            for row in dados_encerrados:
                
                cpf, cliente, produto, preco, data = row
                preco_str = f"R${preco:.2f}"
                data_atual = datetime.strptime(data, "%d/%m/%Y")

                # Aplica o filtro nos dados carregados e exibe os resultados na Treeview
                # Verifica se cada critério de filtro é atendido para cada linha de dados
                if (not cpf_filtro or cpf_filtro in str(cpf)) and \
                        (not cliente_filtro or cliente_filtro.lower() in cliente.lower()) and \
                        (not produto_filtro or produto_filtro.lower() in produto.lower()) and \
                        (not preco_filtro or preco_filtro == preco_str) and \
                        (not data_inicio_filtro or data_atual >= datetime.strptime(data_inicio_filtro, "%d/%m/%Y")) and \
                        (not data_fim_filtro or data_atual <= datetime.strptime(data_fim_filtro, "%d/%m/%Y")):
                    
                    # Insere os dados filtrados na Treeview
                    # O parâmetro "" indica que os valores serão inseridos no nível raiz da Treeview
                    # O parâmetro "end" indica que os valores serão adicionados ao final da Treeview
                    # O parâmetro text é utilizado para definir o texto que aparecerá na primeira coluna da Treeview (coluna "#0")
                    # Os demais parâmetros (values) são utilizados para preencher as demais colunas da Treeview com os valores da linha
                    treeview_encerrados.insert("", "end", text=cliente, values=(cpf, cliente, produto, preco_str, data))

        except ValueError:
            
            # Exibe uma mensagem de erro caso ocorra um problema de formatação de data no filtro
            messagebox.showerror("Erro no Filtro", "O formato da data deve ser DD/MM/AAAA.")
    
    #Botão para filtrar os dados
    botao_filtrar = tk.Button(janela_encerrados, 
                             text="Filtrar", 
                             font=("Arial", 14),
                             width=50,
                            command=filtrar_dados)
    botao_filtrar.pack(side=LEFT, pady=10, padx=5)
    
    #Botão para Sair os dados
    botao_sair = tk.Button(janela_encerrados, 
                             text="Sair", 
                             font=("Arial", 14),
                             width=50,
                            command=janela_encerrados.destroy)
    botao_sair.pack(side=LEFT, pady=10, padx=5)
    
    # Atualiza a interface gráfica para processar todos os eventos pendentes
    janela_encerrados.update_idletasks()

    # Obtém a largura atual da janela
    largura = janela_encerrados.winfo_width()

    # Obtém a altura atual da janela
    altura = janela_encerrados.winfo_height()

    # Calcula a posição x para centralizar a janela na tela
    posicao_x = (janela_encerrados.winfo_screenwidth() // 2) - (largura // 2)

    # Calcula a posição y para centralizar a janela na tela
    posicao_y = (janela_encerrados.winfo_screenheight() // 2) - (altura // 2)

    # Define a geometria da janela com a largura, altura, posição x e posição y
    janela_encerrados.geometry('{}x{}+{}+{}'.format(largura, altura, posicao_x, posicao_y))
    
def sair():
    
    root.destroy()

# Cria uma instância da classe Tk
root = Tk()

# Define o título da janela
root.title("Sistema de Pedidos")

# Configura o fundo da janela como branco
root.configure(bg="white")

# Define a janela em modo de tela cheia
root.attributes("-fullscreen", True)

# Cria uma lista vazia para armazenar os clientes
clientes = []

# Cria um rótulo com uma mensagem de boas-vindas
mensagem_label = Label(root, 
                       text="Bem vindo ao Sistema de Pedidos!",
                      font="Arial 16",
                      bg="white")
mensagem_label.pack(pady = 50)

botao_frame = Frame(root, bg="white")
botao_frame.pack()

# Cria um botão para cadastrar um cliente
"""
Os parâmetros ipadx e ipady definem o espaçamento interno horizontal e vertical do conteúdo do 
botão, respectivamente. Eles controlam o preenchimento interno do botão, ou seja, o espaço adicional 
entre o texto do botão e as bordas do próprio botão.

O valor ipadx=20 define um espaçamento interno horizontal de 20 pixels, enquanto ipady=10 define 
um espaçamento interno vertical de 10 pixels. Isso permite ajustar o tamanho do botão e o espaçamento 
interno para obter uma aparência visualmente agradável.
"""
cadastrar_cliente_button = Button(botao_frame,
                                 text="1. Cadastrar Cliente",
                                 font="Arial 16",
                                 command = tela_cliente,
                                 width = 40)
cadastrar_cliente_button.pack(side = TOP, padx=50, pady=10, ipadx=20, ipady=10)


# Cria um botão para cadastrar produto
cadastrar_produto_button = Button(botao_frame,
                                 text="2. Cadastrar Produto",
                                 font="Arial 16",
                                 command = tela_produtos,
                                 width = 40)
cadastrar_produto_button.pack(side = TOP, padx=50, pady=10, ipadx=20, ipady=10)

# Cria um botão para selecionar um cliente
selecionar_cliente_button = Button(botao_frame,
                                 text="3. Selecionar Cliente",
                                 font="Arial 16",
                                 command = selecionar_cliente,
                                 width = 40)
selecionar_cliente_button.pack(side = TOP, padx=50, pady=10, ipadx=20, ipady=10)

# Cria um botão para visualizar os pedidos de todos os clientes
visualizar_pedidos_button = Button(botao_frame,
                                 text="4. Visualizar Pedidos de Todos os Clientes",
                                 font="Arial 16",
                                 command = ler_pedidos_excel,
                                 width = 40)
visualizar_pedidos_button.pack(side = TOP, padx=50, pady=10, ipadx=20, ipady=10)


# Cria um botão para encerra os pedidos de todos os clientes
pedidos_encerrados_button = Button(botao_frame,
                                 text="5. Pedidos Encerrados",
                                 font="Arial 16",
                                 command = carregar_dados_encerrados,
                                 width = 40)
pedidos_encerrados_button.pack(side = TOP, padx=50, pady=10, ipadx=20, ipady=10)

# Cria um botão para sair do sistema
sair_button = Button(botao_frame,
                                 text="6. Sair",
                                 font="Arial 16",
                                 command = sair,
                                 width = 40)
sair_button.pack(side = TOP, padx=50, pady=10, ipadx=20, ipady=10)

# Inicia o loop principal da aplicação
root.mainloop()

