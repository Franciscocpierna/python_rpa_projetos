# Importa tudo (*) do módulo tkinter. Tkinter é uma 
# biblioteca padrão do Python usada para criar interfaces gráficas.
from tkinter import *

# Importa o módulo tkinter e o renomeia como 'tk', uma 
# convenção comum para facilitar a referência a este módulo.
import tkinter as tk

# Do módulo tkinter, importa classes específicas. Tk é a 
# classe base para a criação de janelas, Frame é um contêiner para outros widgets,
# Label é um widget que exibe texto ou imagens, Entry é um 
# widget para entrada de texto de linha única,
# Button é um widget de botão e messagebox é um módulo para caixas de diálogo padrão.
from tkinter import Tk, Frame, Label, Entry, Button, messagebox

# Importa o módulo ttk do tkinter, que oferece acesso a 
# widgets com um visual melhorado e aprimorado.
from tkinter import ttk

# Importa a biblioteca openpyxl, usada para ler e escrever em arquivos do Excel.
import openpyxl

# Importa a biblioteca pandas e a renomeia como 'pd'. Pandas é uma
# biblioteca popular para análise de dados.
import pandas as pd

# Do módulo openpyxl, importa especificamente as classes Workbook e load_workbook.
# Workbook é usado para criar novos arquivos do Excel, e load_workbook é
# usado para abrir arquivos existentes.
from openpyxl import Workbook, load_workbook

# Importa as classes date e datetime do módulo datetime. 
# 'date' é usada para trabalhar com datas (ano, mês, dia), enquanto
# 'datetime' é para datas e horas.
from datetime import date, datetime

# Declaração de três variáveis globais. As variáveis globais são acessíveis em todo o código, não apenas no escopo onde foram declaradas.
entry_data_pesquisa = None  # Variável para armazenar um widget de entrada de dados (para pesquisa).
window_painel = None  # Variável para armazenar uma janela do painel (janela principal ou um painel específico).
registros_vagas = {}  # Dicionário para armazenar registros de vagas (informações sobre vagas de estacionamento).


# Definição da função 'obter_vagas_do_excel_filtro' com um parâmetro opcional 'evento'
# 'evento=None' significa que a função pode ser chamada sem passar o argumento 'evento';
# se 'evento' não for fornecido, ele terá o valor padrão de None.
def obter_vagas_do_excel_filtro(evento=None):
    
    # Utiliza a palavra-chave 'global' para declarar que a função
    # irá utilizar a variável 'registros_vagas' definida no escopo global.
    # Isso é necessário para modificar a variável global dentro da função.
    global registros_vagas  

    # Carrega a planilha de Excel chamada "Dados.xlsx". Esta planilha
    # deve estar no mesmo diretório que o script Python, 
    # ou o caminho deve ser fornecido. A função 'load_workbook' é usada
    # para abrir uma planilha existente.
    planilha = openpyxl.load_workbook("Dados.xlsx")
    
    # Seleciona uma aba específica da planilha carregada, neste caso, a aba chamada 'Reservas'.
    # As abas são como 'folhas' ou 'páginas' em uma planilha Excel, e cada uma pode conter dados diferentes.
    aba = planilha["Reservas"]

    # Acessa o valor atual do widget 'entrada_data_pesquisa', que é 
    # presumivelmente um campo de entrada (Entry) na interface gráfica.
    # O método '.get()' é usado para obter o texto atualmente inserido 
    # nesse campo. Este valor é usado como a data para a busca,
    # indicando que a função irá buscar registros com base nesta data.
    data_para_busca = entrada_data_pesquisa.get()

    # Reinicializa a variável 'registros_vagas' como um dicionário vazio. 
    # Esta ação limpa quaisquer dados anteriores
    # armazenados em 'registros_vagas', preparando a variável para 
    # armazenar novos dados de vagas baseados na busca atual.
    registros_vagas = {}

    
    # Este loop 'for' itera sobre os números de 1 a 4, incluindo
    # ambos. Cada número representa uma 'seção' no estacionamento.
    for secao in range(1, 5):

        # Aqui, o código converte o número da 'seção' em um caractere ASCII correspondente. 
        # O número 64 é somado ao número da seção, e o resultado é convertido para um caractere.
        # Por exemplo, 64 + 1 = 65, que é o código ASCII para 'A'. Portanto, 
        # 'secao' 1 se torna 'A', 2 se torna 'B', e assim por diante.
        # Então, um novo dicionário vazio é criado para cada seção e atribuído ao 
        # dicionário 'registros_vagas' com a chave sendo a letra da seção.
        registros_vagas[chr(64 + secao)] = {}

        # Este loop 'for' itera sobre os números de 1 a 5, representando 
        # lugares dentro de cada seção do estacionamento.
        for lugar in range(1, 6):

            # Aqui, o número do 'lugar' é convertido para uma string e usado 
            # como chave no subdicionário da seção correspondente.
            # O valor associado a essa chave é inicializado como "--    ", 
            # que pode indicar que o lugar está disponível ou não ocupado.
            # Assim, para cada lugar em cada seção, um registro é criado com o
            # status inicial "--    ".
            registros_vagas[chr(64 + secao)][str(lugar)] = "--    "


    # Este loop itera sobre cada linha na aba 'Reservas' da
    # planilha de Excel, começando da segunda linha.
    # 'min_row=2' indica que o loop começa na segunda linha, 
    # pulando o cabeçalho da tabela.
    # 'values_only=True' significa que o loop retornará apenas 
    # os valores das células, e não os objetos de célula completos.
    for linha in aba.iter_rows(min_row=2, values_only=True):

        # Extrai a data de entrada da terceira coluna (índice 2, 
        # já que a indexação começa do zero) da linha atual.
        # O valor é convertido em uma string para facilitar comparações e manipulações.
        data_entrada = str(linha[2])

        # Extrai a data de saída da quinta coluna (índice 4) da linha 
        # atual, também convertendo para string.
        data_saida = str(linha[4])

        # Extrai o status da décima coluna (índice 9) da linha. O 
        # status pode indicar o estado atual da reserva (por exemplo, 'Reservado').
        status = str(linha[9])

        # Verifica se a data de entrada na linha atual é igual à data 
        # especificada para a busca ('data_para_busca')
        # e se o status da reserva é 'Reservado'.
        # Essa condição é usada para filtrar as reservas que são 
        # relevantes para a data especificada pelo usuário.
        if data_entrada == data_para_busca and status == "Reservado":

            # Obtém a seção da oitava coluna (índice 7) da linha atual, convertendo o valor para string.
            # A seção pode ser um identificador de parte do estacionamento (por exemplo, um setor ou área).
            secao = str(linha[7])  # Coluna 8 (índice 7) para 'secao'

            # Obtém o número do lugar da sétima coluna (índice 6) da linha, convertendo para string.
            # O número do lugar é um identificador específico para a vaga dentro da seção.
            numero_lugar = str(linha[6])  # Coluna 7 (índice 6) para 'numero_lugar'

            # Atualiza o valor correspondente no dicionário 'registros_vagas' para a seção e o lugar especificados.
            # O valor 'X    ' pode indicar que a vaga está ocupada ou reservada para a data de busca.
            # O espaço após o 'X' pode ser usado para alinhamento na interface de usuário.
            registros_vagas[secao][numero_lugar] = "X    "

    # Ao final do loop, a função retorna o dicionário 'registros_vagas' atualizado.
    # Este dicionário contém as informações sobre quais vagas estão reservadas para a data especificada.
    return registros_vagas
    

# Define a função 'exibir_painel_vagas'. Esta função não tem
# parâmetros e é usada para exibir uma janela de painel de estacionamento.
def exibir_painel_vagas():

    # Declara variáveis como globais para que as alterações feitas nessas variáveis dentro da função sejam refletidas globalmente.
    # Essas variáveis são usadas para armazenar widgets e outras informações importantes para a interface gráfica.
    global entrada_data_pesquisa, janela_painel, registros_vagas, frame_reserva, widgets_entrada
    
    # Cria uma nova janela (tipo Toplevel, que é uma janela que pode
    # existir independentemente da janela principal) para exibir o painel de estacionamento.
    # 'Toplevel()' cria uma nova janela na interface gráfica.
    janela_painel = Toplevel()
    
    # Define o título da janela criada como "Painel de Estacionamento".
    janela_painel.title("Painel de Estacionamento")

    # Cria um Frame, que é um contêiner para agrupar e organizar outros widgets. Este Frame é colocado na janela de painel.
    # O argumento 'bg="#FFFFFF"' define a cor de fundo do Frame como branco.
    frame_busca = Frame(janela_painel, bg="#FFFFFF")
    
    # O método 'pack' com os argumentos 'fill="both"' e 'expand=True' 
    # faz com que o Frame expanda para preencher todo o espaço disponível na janela.
    frame_busca.pack(fill="both", expand=True)

    # Cria um widget de entrada de texto (Entry) para inserção da data de pesquisa.
    # Este widget é configurado com uma fonte Arial tamanho 20 e alinhamento centralizado (justify=CENTER).
    entrada_data_pesquisa = Entry(frame_busca, font="Arial 20", justify=CENTER)
    
    # Posiciona o widget de entrada na grade do Frame de busca. 'padx' e 
    # 'pady' adicionam um preenchimento em torno do widget.
    entrada_data_pesquisa.grid(row=0, column=0, padx=5, pady=10)

    # Obtém a data atual usando a função 'now()' do módulo 'datetime' e
    # formata como dia/mês/ano.
    data_atual = datetime.now().strftime("%d/%m/%Y")
    
    # Insere a data atual no campo de entrada de texto, posicionando-a
    # no início do campo (índice 0).
    entrada_data_pesquisa.insert(0, data_atual)


    # Dentro da função 'exibir_painel_vagas', define uma função
    # interna chamada 'buscar_vagas'.
    # Esta função interna é usada para buscar vagas atualizadas do
    # arquivo Excel e atualizar a visualização das vagas no painel.
    def buscar_vagas():

        # Declara a variável 'registros_vagas' como global. 
        # Isso permite que a função modifique a variável 'registros_vagas'
        # definida no escopo global.
        global registros_vagas

        # Chama a função 'obter_vagas_do_excel_filtro' e armazena o 
        # retorno, que é um dicionário atualizado de vagas, em 'registros_vagas'.
        # Esta função busca e processa informações de
        # vagas de estacionamento de um arquivo Excel.
        registros_vagas = obter_vagas_do_excel_filtro()

        # Chama a função 'atualizar_vagas_no_painel' e passa 'registros_vagas' como argumento.
        # Esta função é responsável por atualizar a interface gráfica com
        # as informações mais recentes das vagas.
        atualizar_vagas_no_painel(registros_vagas)
        

    # Cria um botão na interface gráfica chamado "Buscar" e o vincula à função 'buscar_vagas'.
    # Quando o botão é clicado, a função 'buscar_vagas' é executada.
    botao_buscar = Button(frame_busca, text="Buscar", command=buscar_vagas, font="Arial 16")
    
    # Posiciona o botão no layout da interface usando o método 'grid'. 
    # 'padx' e 'pady' adicionam um preenchimento externo em torno do botão para
    # separá-lo de outros elementos da interface.
    botao_buscar.grid(row=0, column=1, padx=5, pady=5)
    
    
    # Cria um novo Frame na janela do painel de estacionamento para
    # conter widgets relacionados à reserva de vagas.
    # Este Frame tem um fundo branco ('bg="#FFFFFF"').
    frame_reserva = Frame(janela_painel, bg="#FFFFFF")
    
    # Empacota o frame na janela usando 'pack', preenchendo todo o 
    # espaço disponível ('fill="both"') e permitindo que ele expanda ('expand=True').
    # 'pady=20' adiciona um preenchimento vertical de 20 pixels.
    frame_reserva.pack(fill="both", expand=True, pady=20)
    
    
    # Configurando as colunas no frame 'frame_reserva' para terem pesos iguais.
    # Isso garante que as colunas se ajustem dinamicamente ao tamanho do frame.
    frame_reserva.grid_columnconfigure(0, weight=1)  # Configura a primeira coluna para expandir proporcionalmente.
    frame_reserva.grid_columnconfigure(1, weight=1)  # Configura a segunda coluna para expandir proporcionalmente.
    
    
    # Definindo a função 'atualizar_vagas_no_painel' que 
    # aceita 'registros_vagas' como argumento.
    # Esta função é responsável por atualizar a interface gráfica do
    # painel de estacionamento com as informações mais recentes das vagas.
    def atualizar_vagas_no_painel(registros_vagas):
        
        # Este loop itera sobre todos os widgets (elementos da
        # interface gráfica) que estão presentes na janela 'janela_painel'.
        for widget in janela_painel.winfo_children():
            
            # 'winfo_children()' retorna uma lista de todos os widgets filhos da 'janela_painel'.
            # A função verifica se o widget atual não é parte dos frames 'frame_busca' ou 'frame_reserva'.
            # Esses frames contêm elementos importantes da interface que não devem ser removidos.
            if widget not in (frame_busca, frame_reserva):
                
                # Se o widget atual não faz parte dos frames essenciais 
                # (busca ou reserva), ele é removido da janela.
                # 'widget.destroy()' é um método que remove o widget da interface gráfica.
                widget.destroy()


        # Criando um rótulo (Label) para servir como cabeçalho, mostrando a estrutura da seção de vagas.
        # Este rótulo funcionará como um cabeçalho para exibir a organização das vagas do estacionamento.
        label_cabecalho = Label(janela_painel, 
                                text="Seção/vaga 1     2      3     4      5", # A string "Seção/vaga 1     2      3     4      5" indica as seções e números das vagas,
                                font="Arial 16") # 'Arial 16' é o tipo e tamanho da fonte. 

        # Adicionando o rótulo ao layout da janela 'janela_painel' usando o método 'pack'.
        # 'pack' é um gerenciador de geometria no Tkinter que organiza os
        # widgets em blocos antes de colocá-los na janela pai.
        label_cabecalho.pack()

        # Neste loop, iteramos sobre cada seção e suas vagas no dicionário 'registros_vagas'.
        # 'registros_vagas' é um dicionário onde cada chave é uma seção do
        # estacionamento e o valor é um dicionário das vagas.
        for secao, v in registros_vagas.items():
            
            # Criando um novo quadro (Frame) para cada seção dentro da janela 'janela_painel'.
            # Um Frame é um container que pode conter outros widgets.
            frame_secao = Frame(janela_painel)

            # Criando um rótulo para a seção atual e adicionando ao quadro da seção.
            # O texto do rótulo é o nome da seção, alinhado à esquerda com
            # 19 caracteres de espaço ('{:<19}'.format(secao)).
            label_secao = Label(frame_secao, 
                                text="{:<19}".format(secao), 
                                font="Arial 16")
            
            # Empacotando o rótulo no lado esquerdo do quadro da seção.
            label_secao.pack(side=LEFT)

            # Agora, iteramos sobre cada vaga na seção.
            # Esta iteração é feita para cada seção individualmente, processando 
            # as vagas de número 1 a 5.
            for lugar in range(1, 6):
                
                # Obtemos o status da vaga.
                # 'status' recebe o valor correspondente à vaga atual na
                # seção, que pode indicar se a vaga está ocupada ou livre.
                status = v[str(lugar)]

                # Criamos um rótulo para a vaga e o adicionamos ao quadro da seção.
                # Cada vaga é representada por um rótulo que mostra seu status atual.
                label_lugar = Label(frame_secao, text=status, font="Arial 16")
                
                # O método 'pack' com 'side=LEFT' organiza os rótulos das 
                # vagas em linha, lado a lado, dentro do quadro da seção.
                label_lugar.pack(side=LEFT)

            # Finalmente, adicionamos o quadro da seção completo ao layout
            # da janela 'janela_painel'.
            # Após configurar todos os rótulos de vagas dentro de um 
            # quadro de seção, o quadro inteiro é adicionado à janela principal.
            frame_secao.pack()
                
    
    
    

def reservas():
    
    print("Tela Reservas...")

def tela_clientes():
    
    print("Exibir Tela Clientes...")




# Cria uma instância da classe Tk
janela = tk.Tk()

# Define o título da janela
janela.title("Sistema de Pedidos")

# Configura o fundo da janela como branco
janela.configure(bg="white")

# Define a janela em modo de tela cheia
janela.attributes("-fullscreen", True)

# Importa as classes Tk e Label da biblioteca tkinter.
from tkinter import Tk, Label

# Importa as classes ImageTk e Image da biblioteca Python Imaging Library (PIL).
from PIL import ImageTk, Image

# Inicia um bloco try-except para capturar possíveis exceções
# durante a execução do código.
try:
                                           
    # Usa a função open do módulo Image para carregar a imagem 'estacionamento.jpg' do disco.
    imagem_carregada = Image.open("estacionamento.jpg")

    # Converte a imagem para um formato compatível com o Tkinter, criando um objeto PhotoImage.
    imagem = ImageTk.PhotoImage(imagem_carregada)

    # Cria um objeto Label na janela 'janela'. Este Label será usado para exibir a imagem.
    rotulo_imagem = Label(janela)

    # Armazena uma referência da imagem no rótulo. Isso é necessário 
    # porque o Tkinter pode descartar a imagem se não houver uma referência forte para ela.
    rotulo_imagem.image = imagem

    # Configura o rótulo para exibir a imagem.
    rotulo_imagem.configure(image=imagem)

    # Adiciona o rótulo à janela, usando o gerenciador de layout pack.
    rotulo_imagem.pack()

# Se ocorrer uma exceção durante a execução do bloco try, ela é 
# capturada aqui. A exceção é impressa na saída padrão.
except Exception as e:
    print(f"Erro ao carregar a imagem: {e}")

    
# Cria um rótulo (Label) na janela 'janela' com uma mensagem de boas-vindas.
# Define a fonte do texto como Arial de tamanho 16 e o fundo do rótulo como branco.
mensagem_label = tk.Label(janela, 
                          text="Bem-vindo ao Controle de Vagas Estacionamento", 
                          font=("Arial", 16), bg="white")

# Adiciona o rótulo à janela, usando o gerenciador de layout pack.
# A opção 'pady' adiciona um espaçamento vertical de 50 pixels acima e abaixo do rótulo.
mensagem_label.pack(pady=50)

    
# Cria um frame (container) na janela 'janela' para organizar os botões.
# O fundo do frame é definido como branco.
button_frame = tk.Frame(janela, bg="white")

# Adiciona o frame à janela, usando o gerenciador de layout pack.
button_frame.pack()


# Cria um botão no 'button_frame'.
# Define o texto do botão, a fonte, o comando a ser executado quando o botão é clicado,
# e a largura do botão.
# O comando 'exibir_painel_vagas' deve ser uma função definida em outro lugar no código.
botao_painel = tk.Button(button_frame, 
                         text="Exibir Painel de Vagas", 
                         font=("Arial", 12),
                         command=exibir_painel_vagas, 
                         width=40)

# Adiciona o botão ao frame, usando o gerenciador de layout pack.
# A opção 'side=tk.TOP' posiciona o botão no topo do frame.
# 'padx' e 'pady' adicionam espaçamentos horizontal e vertical, respectivamente.
# 'ipadx' e 'ipady' adicionam espaçamentos internos no botão, horizontal e verticalmente.
botao_painel.pack(side=tk.TOP, padx=50, pady=10, ipadx=20, ipady=10)

# Cria um botão para a tela de clientes no 'button_frame'.
# O texto do botão é "Clientes", e a fonte definida é Arial tamanho 12.
# O comando 'tela_clientes' é chamado quando o botão é clicado, e a largura do botão é definida como 40.
botao_clientes = tk.Button(button_frame, 
                           text="Clientes", 
                           font=("Arial", 12),
                           command=tela_clientes,
                           width=40)

# Adiciona o botão 'botao_clientes' ao frame, alinhando-o no topo.
# Define espaçamentos horizontal e vertical de 50 e 10 pixels, respectivamente.
# 'ipadx' e 'ipady' adicionam espaçamento interno no botão, horizontal e verticalmente.
botao_clientes.pack(side=tk.TOP, padx=50, pady=10, ipadx=20, ipady=10)

# Cria um botão para a tela de reservas no mesmo frame.
# O texto do botão é "Reservas", a fonte é Arial tamanho 12, e a largura é 40.
# O comando 'reservas' é associado a este botão para ser executado quando clicado.
botao_Reservas = tk.Button(button_frame, 
                           text="Reservas", 
                           font=("Arial", 12), 
                           command=reservas, 
                           width=40)

# Posiciona o botão 'botao_Reservas' no frame, alinhando-o no topo.
# Define os mesmos espaçamentos horizontal e vertical que os outros botões.
botao_Reservas.pack(side=tk.TOP, padx=50, pady=10, ipadx=20, ipady=10)


    
# Cria um botão para sair do sistema.
# O texto do botão é "Sair", a fonte é Arial tamanho 12, e a largura é definida como 40.
# O comando 'janela.destroy' é associado a este botão, o que fechará a janela quando clicado.
sair_button = tk.Button(button_frame, 
                        text="Sair",                    
                        font=("Arial", 12), 
                        command=janela.destroy, 
                        width=40)

# Posiciona o botão 'sair_button' no frame, alinhando-o no topo.
# Mantém o mesmo padrão de espaçamentos dos botões anteriores.
# 'padx' e 'pady' adicionam espaçamentos horizontal e vertical, respectivamente.
# 'ipadx' e 'ipady' adicionam espaçamentos internos no botão, horizontal e verticalmente.
sair_button.pack(side=tk.TOP, padx=50, pady=10, ipadx=20, ipady=10)
    
# Inicia o loop principal da aplicação
janela.mainloop()
    
