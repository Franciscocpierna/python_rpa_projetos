# Importa tudo (*) do módulo tkinter. Tkinter é uma 
# biblioteca padrão do Python usada para criar interfaces gráficas.
from tkinter import *

# Importa o módulo tkinter e o renomeia como 'tk', uma 
# convenção comum para facilitar a referência a este módulo.
import tkinter as tk

# Do módulo tkinter, importa classes específicas. Tk é a 
# classe base para a criação de janelas, Frame é um contêiner para outros widgets,
# Label é um widget que exibe texto ou imagens, Entry é um 
# widget para entrada de texto de linha única,
# Button é um widget de botão e messagebox é um módulo para caixas de diálogo padrão.
from tkinter import Tk, Frame, Label, Entry, Button, messagebox

# Importa o módulo ttk do tkinter, que oferece acesso a 
# widgets com um visual melhorado e aprimorado.
from tkinter import ttk

# Importa a biblioteca openpyxl, usada para ler e escrever em arquivos do Excel.
import openpyxl

# Importa a biblioteca pandas e a renomeia como 'pd'. Pandas é uma
# biblioteca popular para análise de dados.
import pandas as pd

# Do módulo openpyxl, importa especificamente as classes Workbook e load_workbook.
# Workbook é usado para criar novos arquivos do Excel, e load_workbook é
# usado para abrir arquivos existentes.
from openpyxl import Workbook, load_workbook

# Importa as classes date e datetime do módulo datetime. 
# 'date' é usada para trabalhar com datas (ano, mês, dia), enquanto
# 'datetime' é para datas e horas.
from datetime import date, datetime

# Declaração de três variáveis globais. As variáveis globais são acessíveis em todo o código, não apenas no escopo onde foram declaradas.
entry_data_pesquisa = None  # Variável para armazenar um widget de entrada de dados (para pesquisa).
window_painel = None  # Variável para armazenar uma janela do painel (janela principal ou um painel específico).
registros_vagas = {}  # Dicionário para armazenar registros de vagas (informações sobre vagas de estacionamento).


# Definição da função 'obter_vagas_do_excel_filtro' com um parâmetro opcional 'evento'
# 'evento=None' significa que a função pode ser chamada sem passar o argumento 'evento';
# se 'evento' não for fornecido, ele terá o valor padrão de None.
def obter_vagas_do_excel_filtro(evento=None):
    
    # Utiliza a palavra-chave 'global' para declarar que a função
    # irá utilizar a variável 'registros_vagas' definida no escopo global.
    # Isso é necessário para modificar a variável global dentro da função.
    global registros_vagas  

    # Carrega a planilha de Excel chamada "Dados.xlsx". Esta planilha
    # deve estar no mesmo diretório que o script Python, 
    # ou o caminho deve ser fornecido. A função 'load_workbook' é usada
    # para abrir uma planilha existente.
    planilha = openpyxl.load_workbook("Dados.xlsx")
    
    # Seleciona uma aba específica da planilha carregada, neste caso, a aba chamada 'Reservas'.
    # As abas são como 'folhas' ou 'páginas' em uma planilha Excel, e cada uma pode conter dados diferentes.
    aba = planilha["Reservas"]

    # Acessa o valor atual do widget 'entrada_data_pesquisa', que é 
    # presumivelmente um campo de entrada (Entry) na interface gráfica.
    # O método '.get()' é usado para obter o texto atualmente inserido 
    # nesse campo. Este valor é usado como a data para a busca,
    # indicando que a função irá buscar registros com base nesta data.
    data_para_busca = entrada_data_pesquisa.get()

    # Reinicializa a variável 'registros_vagas' como um dicionário vazio. 
    # Esta ação limpa quaisquer dados anteriores
    # armazenados em 'registros_vagas', preparando a variável para 
    # armazenar novos dados de vagas baseados na busca atual.
    registros_vagas = {}

    
    # Este loop 'for' itera sobre os números de 1 a 4, incluindo
    # ambos. Cada número representa uma 'seção' no estacionamento.
    for secao in range(1, 5):

        # Aqui, o código converte o número da 'seção' em um caractere ASCII correspondente. 
        # O número 64 é somado ao número da seção, e o resultado é convertido para um caractere.
        # Por exemplo, 64 + 1 = 65, que é o código ASCII para 'A'. Portanto, 
        # 'secao' 1 se torna 'A', 2 se torna 'B', e assim por diante.
        # Então, um novo dicionário vazio é criado para cada seção e atribuído ao 
        # dicionário 'registros_vagas' com a chave sendo a letra da seção.
        registros_vagas[chr(64 + secao)] = {}

        # Este loop 'for' itera sobre os números de 1 a 5, representando 
        # lugares dentro de cada seção do estacionamento.
        for lugar in range(1, 6):

            # Aqui, o número do 'lugar' é convertido para uma string e usado 
            # como chave no subdicionário da seção correspondente.
            # O valor associado a essa chave é inicializado como "--    ", 
            # que pode indicar que o lugar está disponível ou não ocupado.
            # Assim, para cada lugar em cada seção, um registro é criado com o
            # status inicial "--    ".
            registros_vagas[chr(64 + secao)][str(lugar)] = "--    "


    # Este loop itera sobre cada linha na aba 'Reservas' da
    # planilha de Excel, começando da segunda linha.
    # 'min_row=2' indica que o loop começa na segunda linha, 
    # pulando o cabeçalho da tabela.
    # 'values_only=True' significa que o loop retornará apenas 
    # os valores das células, e não os objetos de célula completos.
    for linha in aba.iter_rows(min_row=2, values_only=True):

        # Extrai a data de entrada da terceira coluna (índice 2, 
        # já que a indexação começa do zero) da linha atual.
        # O valor é convertido em uma string para facilitar comparações e manipulações.
        data_entrada = str(linha[2])

        # Extrai a data de saída da quinta coluna (índice 4) da linha 
        # atual, também convertendo para string.
        data_saida = str(linha[4])

        # Extrai o status da décima coluna (índice 9) da linha. O 
        # status pode indicar o estado atual da reserva (por exemplo, 'Reservado').
        status = str(linha[9])

        # Verifica se a data de entrada na linha atual é igual à data 
        # especificada para a busca ('data_para_busca')
        # e se o status da reserva é 'Reservado'.
        # Essa condição é usada para filtrar as reservas que são 
        # relevantes para a data especificada pelo usuário.
        if data_entrada == data_para_busca and status == "Reservado":

            # Obtém a seção da oitava coluna (índice 7) da linha atual, convertendo o valor para string.
            # A seção pode ser um identificador de parte do estacionamento (por exemplo, um setor ou área).
            secao = str(linha[7])  # Coluna 8 (índice 7) para 'secao'

            # Obtém o número do lugar da sétima coluna (índice 6) da linha, convertendo para string.
            # O número do lugar é um identificador específico para a vaga dentro da seção.
            numero_lugar = str(linha[6])  # Coluna 7 (índice 6) para 'numero_lugar'

            # Atualiza o valor correspondente no dicionário 'registros_vagas' para a seção e o lugar especificados.
            # O valor 'X    ' pode indicar que a vaga está ocupada ou reservada para a data de busca.
            # O espaço após o 'X' pode ser usado para alinhamento na interface de usuário.
            registros_vagas[secao][numero_lugar] = "X    "

    # Ao final do loop, a função retorna o dicionário 'registros_vagas' atualizado.
    # Este dicionário contém as informações sobre quais vagas estão reservadas para a data especificada.
    return registros_vagas


# Define uma função chamada 'obter_hora_atual'. Esta função 
# recebe um argumento chamado 'widgets',
# que é esperado ser um dicionário contendo widgets da 
# interface gráfica, especificamente um widget de entrada de texto.
def obter_hora_atual(widgets):
    
    # Utiliza a função 'now()' do módulo 'datetime' para obter o momento atual.
    # Em seguida, a função 'strftime' formata esse momento no formato de hora e minuto ("%H:%M").
    # '%H:%M' significa hora e minuto no formato 24 horas, por exemplo, "15:30" para 3:30 PM.
    # O resultado formatado é armazenado na variável 'hora_atual'.
    hora_atual = datetime.now().strftime("%H:%M")
    
    # Acessa o widget de entrada de texto para a 'Hora de Entrada' no dicionário 'widgets'.
    # O método 'delete' é chamado no widget, com os parâmetros 0 e END, 
    # que efetivamente limpa todo o texto presente.
    # 0 é o índice inicial do texto no widget, e END é uma constante 
    # que representa o final do texto.
    widgets["Hora de Entrada"].delete(0, END)
    
    # A função 'insert' é usada para inserir a 'hora_atual' no widget de entrada de texto para a 'Hora de Entrada'.
    # O primeiro argumento, 0, especifica que a inserção deve ocorrer a partir do início do widget (índice 0).
    # Assim, a hora atual formatada é inserida no campo, mostrando a hora atual ao usuário.
    widgets["Hora de Entrada"].insert(0, hora_atual)
    

# Define a função 'exibir_painel_vagas'. Esta função não tem
# parâmetros e é usada para exibir uma janela de painel de estacionamento.
def exibir_painel_vagas():

    # Declara variáveis como globais para que as alterações feitas nessas variáveis dentro da função sejam refletidas globalmente.
    # Essas variáveis são usadas para armazenar widgets e outras informações importantes para a interface gráfica.
    global entrada_data_pesquisa, janela_painel, registros_vagas, frame_reserva, widgets_entrada
    
    # Cria uma nova janela (tipo Toplevel, que é uma janela que pode
    # existir independentemente da janela principal) para exibir o painel de estacionamento.
    # 'Toplevel()' cria uma nova janela na interface gráfica.
    janela_painel = Toplevel()
    
    # Define o título da janela criada como "Painel de Estacionamento".
    janela_painel.title("Painel de Estacionamento")

    # Cria um Frame, que é um contêiner para agrupar e organizar outros widgets. Este Frame é colocado na janela de painel.
    # O argumento 'bg="#FFFFFF"' define a cor de fundo do Frame como branco.
    frame_busca = Frame(janela_painel, bg="#FFFFFF")
    
    # O método 'pack' com os argumentos 'fill="both"' e 'expand=True' 
    # faz com que o Frame expanda para preencher todo o espaço disponível na janela.
    frame_busca.pack(fill="both", expand=True)

    # Cria um widget de entrada de texto (Entry) para inserção da data de pesquisa.
    # Este widget é configurado com uma fonte Arial tamanho 20 e alinhamento centralizado (justify=CENTER).
    entrada_data_pesquisa = Entry(frame_busca, font="Arial 20", justify=CENTER)
    
    # Posiciona o widget de entrada na grade do Frame de busca. 'padx' e 
    # 'pady' adicionam um preenchimento em torno do widget.
    entrada_data_pesquisa.grid(row=0, column=0, padx=5, pady=10)

    # Obtém a data atual usando a função 'now()' do módulo 'datetime' e
    # formata como dia/mês/ano.
    data_atual = datetime.now().strftime("%d/%m/%Y")
    
    # Insere a data atual no campo de entrada de texto, posicionando-a
    # no início do campo (índice 0).
    entrada_data_pesquisa.insert(0, data_atual)


    # Dentro da função 'exibir_painel_vagas', define uma função
    # interna chamada 'buscar_vagas'.
    # Esta função interna é usada para buscar vagas atualizadas do
    # arquivo Excel e atualizar a visualização das vagas no painel.
    def buscar_vagas():

        # Declara a variável 'registros_vagas' como global. 
        # Isso permite que a função modifique a variável 'registros_vagas'
        # definida no escopo global.
        global registros_vagas

        # Chama a função 'obter_vagas_do_excel_filtro' e armazena o 
        # retorno, que é um dicionário atualizado de vagas, em 'registros_vagas'.
        # Esta função busca e processa informações de
        # vagas de estacionamento de um arquivo Excel.
        registros_vagas = obter_vagas_do_excel_filtro()

        # Chama a função 'atualizar_vagas_no_painel' e passa 'registros_vagas' como argumento.
        # Esta função é responsável por atualizar a interface gráfica com
        # as informações mais recentes das vagas.
        atualizar_vagas_no_painel(registros_vagas)
        

    # Cria um botão na interface gráfica chamado "Buscar" e o vincula à função 'buscar_vagas'.
    # Quando o botão é clicado, a função 'buscar_vagas' é executada.
    botao_buscar = Button(frame_busca, text="Buscar", command=buscar_vagas, font="Arial 16")
    
    # Posiciona o botão no layout da interface usando o método 'grid'. 
    # 'padx' e 'pady' adicionam um preenchimento externo em torno do botão para
    # separá-lo de outros elementos da interface.
    botao_buscar.grid(row=0, column=1, padx=5, pady=5)
    
    
    # Cria um novo Frame na janela do painel de estacionamento para
    # conter widgets relacionados à reserva de vagas.
    # Este Frame tem um fundo branco ('bg="#FFFFFF"').
    frame_reserva = Frame(janela_painel, bg="#FFFFFF")
    
    # Empacota o frame na janela usando 'pack', preenchendo todo o 
    # espaço disponível ('fill="both"') e permitindo que ele expanda ('expand=True').
    # 'pady=20' adiciona um preenchimento vertical de 20 pixels.
    frame_reserva.pack(fill="both", expand=True, pady=20)
    
    
    # Define uma lista chamada 'colunas' com os nomes das colunas para a reserva de vagas.
    # Esses nomes correspondem a campos de dados a serem preenchidos ou exibidos na interface.
    colunas = ["CPF", "Nome", "Data de Entrada", "Hora de Entrada", "Número da Vaga", "Seção"]

    # Inicializa um dicionário chamado 'widgets_entrada'. 
    # Este dicionário será usado para armazenar e acessar 
    # widgets de entrada de dados associados a cada coluna.
    widgets_entrada = {}

    # Este comando lê os dados da planilha Excel localizada no 
    # caminho especificado em 'caminho_arquivo'.
    # 'planilha_nome' especifica o nome da aba na planilha 
    # Excel que contém os dados relevantes.
    # Os dados lidos são armazenados em um DataFrame do pandas chamado 'df'.
    caminho_arquivo = "Dados.xlsx"
    planilha_nome = "Clientes"
    df = pd.read_excel(caminho_arquivo, sheet_name=planilha_nome)

    # Cria variáveis 'StringVar' do tkinter, que são usadas para
    # armazenar e manipular strings em interfaces gráficas.
    # 'combo_cpf_var' e 'combo_nome_var' armazenarão os valores 
    # selecionados nos comboboxes de CPF e Nome, respectivamente.
    combo_cpf_var = StringVar()
    combo_nome_var = StringVar()

    # Define uma função chamada 'cpf_selecionado'. Esta função é 
    # chamada quando um evento ocorre (por exemplo, um item é selecionado em uma combobox).
    # 'event' é um parâmetro que representa o evento que disparou a função, 
    # mas não é usado explicitamente no corpo da função.
    def cpf_selecionado(event):

        # Obtém o valor atual selecionado no combobox de CPF, que está 
        # vinculado à variável 'combo_cpf_var'.
        cpf_atual = combo_cpf_var.get()

        # Filtra o DataFrame 'df' para incluir apenas as linhas onde o valor
        # na coluna 'CPF' é igual ao CPF atualmente selecionado.
        # 'df['CPF'] == cpf_atual' cria uma série de valores booleanos, que é 
        # usada para filtrar o DataFrame.
        # O resultado é um novo DataFrame chamado 'df_filtrado'.
        df_filtrado = df[df['CPF'] == cpf_atual]

        # Verifica se 'df_filtrado' contém algum dado (não está vazio).
        if not df_filtrado.empty:  

            # Obtém o primeiro valor na coluna 'Nome' de 'df_filtrado', que
            # corresponde ao nome associado ao CPF selecionado.
            # '.values[0]' é usado para acessar o primeiro valor da série de 'Nome'.
            nome_correspondente = df_filtrado['Nome'].values[0]

            # Atualiza a variável 'combo_nome_var' para refletir o 
            # nome correspondente ao CPF selecionado.
            # Isso atualiza o combobox de Nome na interface gráfica para 
            # mostrar o nome correspondente.
            combo_nome_var.set(nome_correspondente)

        else:

            # Se 'df_filtrado' estiver vazio, significa que não 
            # foram encontrados registros com o CPF selecionado.
            # Nesse caso, imprime uma mensagem de erro indicando que
            # o CPF não foi encontrado.
            print(f"CPF {cpf_atual} não encontrado.")
            
            
    # Percorre a lista de colunas através de um loop for
    # Inicia um loop 'for' que percorre cada item na lista 'colunas', 
    # utilizando a função 'enumerate' para obter tanto o índice (i) quanto o
    # valor (coluna) de cada item na lista.
    for i, coluna in enumerate(colunas):

        # Cria um rótulo (Label) para cada coluna dentro do 'frame_reserva'.
        # 'text=coluna' define o texto do rótulo como o nome da coluna.
        # 'font="Arial 14"' define a fonte do texto do rótulo.
        # 'bg="#FFFFFF"' define a cor de fundo do rótulo como branco.
        rotulo = Label(frame_reserva, text=coluna, font="Arial 14", bg="#FFFFFF")
        
        # Posiciona o rótulo no 'frame_reserva' usando o método 'grid'. 
        # 'row=i' coloca o rótulo na linha correspondente ao índice da coluna.
        # 'padx=5' e 'pady=5' adicionam um preenchimento externo ao redor do rótulo.
        rotulo.grid(row=i, column=0, padx=5, pady=5)

        # Verifica se a coluna atual é "CPF".
        if coluna == "CPF":

            # Cria uma lista dos valores da coluna 'CPF' do 
            # DataFrame 'df' e armazena em 'valores_combo'.
            # A função 'tolist()' converte a série de pandas em uma lista Python.
            valores_combo = df['CPF'].tolist()

            # Cria um Combobox (caixa de combinação) para a coluna "CPF" no 'frame_reserva'.
            # 'textvariable=combo_cpf_var' vincula o combobox à variável 'combo_cpf_var'.
            # 'values=valores_combo' define os valores que aparecem na lista dropdown do combobox.
            # 'font="Arial 14"' define a fonte do texto dentro do combobox.
            entrada = ttk.Combobox(frame_reserva, 
                                   textvariable=combo_cpf_var, 
                                   values=valores_combo, 
                                   font="Arial 14")

            # Vincula os eventos de seleção ('<<ComboboxSelected>>') e
            # perda de foco ('<FocusOut>') do combobox à função 'cpf_selecionado'.
            # Isso significa que a função 'cpf_selecionado' será chamada 
            # quando um item for selecionado no combobox ou quando ele perder o foco.
            entrada.bind("<<ComboboxSelected>>", cpf_selecionado)
            entrada.bind("<FocusOut>", cpf_selecionado)
            
        elif coluna == "Nome":
            
            # Cria uma lista dos valores da coluna 'CPF' do 
            # DataFrame 'df' e armazena em 'valores_combo'.
            # A função 'tolist()' converte a série de pandas em uma lista Python.
            valores_combo = df['Nome'].tolist()

            # Cria um Combobox (caixa de combinação) para a coluna "CPF" no 'frame_reserva'.
            # 'textvariable=combo_cpf_var' vincula o combobox à variável 'combo_cpf_var'.
            # 'values=valores_combo' define os valores que aparecem na lista dropdown do combobox.
            # 'font="Arial 14"' define a fonte do texto dentro do combobox.
            entrada = ttk.Combobox(frame_reserva, 
                                   textvariable=combo_nome_var, 
                                   values=valores_combo, 
                                   font="Arial 14")

        # Verifica se a coluna atual é "Nome".
        elif coluna == "Nome":

            # Cria uma lista dos valores da coluna 'Nome' do 
            # DataFrame 'df' e armazena em 'valores_combo'.
            # A função 'tolist()' converte a série de pandas em uma lista Python.
            valores_combo = df['Nome'].tolist()

            # Cria um Combobox (caixa de combinação) para a coluna "Nome" no 'frame_reserva'.
            # 'textvariable=combo_nome_var' vincula o combobox à variável 'combo_nome_var'.
            # 'values=valores_combo' define os valores que aparecem na lista dropdown do combobox.
            # 'font="Arial 14"' define a fonte do texto dentro do combobox.
            entrada = ttk.Combobox(frame_reserva, 
                                   textvariable=combo_nome_var, 
                                   values=valores_combo, 
                                   font="Arial 14")

        # Verifica se a coluna atual é "Número da Vaga".
        elif coluna == "Número da Vaga":

            # Cria uma lista de strings representando números de
            # 1 a 5, usando compreensão de lista.
            # A função 'str(num)' converte o número inteiro em uma string.
            # Esta lista será usada como os valores para o combobox da coluna "Número da Vaga".
            valores_combo = [str(num) for num in range(1, 6)]

            # Cria um Combobox para a coluna "Número da Vaga" no 'frame_reserva'.
            # 'values=valores_combo' define os valores que aparecem na 
            # lista dropdown do combobox.
            # 'font="Arial 14"' define a fonte do texto dentro do combobox.
            # 'state="readonly"' configura o combobox para ser somente leitura, 
            # o que significa que o usuário só pode selecionar os valores da lista, sem editar.
            entrada = ttk.Combobox(frame_reserva, 
                                   values=valores_combo, 
                                   font="Arial 14", 
                                   state="readonly")


        # Verifica se a coluna atual é a coluna "Seção"
        elif coluna == "Seção":

            # Cria uma lista estática com os valores "A", "B", "C", "D", que 
            # representam as seções disponíveis.
            # Esses valores são usados para preencher o combobox para a coluna "Seção".
            valores_combo = ["A", "B", "C", "D"]

            # Cria um Combobox para a coluna "Seção" no 'frame_reserva'.
            # 'values=valores_combo' define os valores que aparecem na lista dropdown do combobox.
            # 'font="Arial 14"' define a fonte do texto dentro do combobox.
            # 'state="readonly"' configura o combobox para ser somente leitura, o que significa 
            # que o usuário só pode selecionar os valores da lista, sem editar.
            entrada = ttk.Combobox(frame_reserva, values=valores_combo, font="Arial 14", state="readonly")

        # Caso a coluna atual não seja nenhuma das anteriores (CPF, 
        # Nome, Número da Vaga ou Seção).
        else:

            # Cria um campo de entrada (Entry) para colunas que não têm um combobox específico.
            # 'font="Arial 14"' define a fonte do texto dentro do campo de entrada.
            entrada = Entry(frame_reserva, font="Arial 14")

        # Posiciona a entrada (seja um combobox ou um campo de entrada normal) na grade do 'frame_reserva'.
        # 'row=i' coloca a entrada na linha correspondente ao índice da coluna.
        # 'padx=5' e 'pady=5' adicionam um preenchimento externo ao redor da entrada.
        entrada.grid(row=i, column=1, padx=5, pady=5)

        # Adiciona a entrada criada (combobox ou campo de entrada) ao dicionário 'widgets_entrada'.
        # Usa o nome da coluna como chave e a entrada como valor.
        # Isso permite que as entradas sejam facilmente acessadas mais tarde 
        # usando o nome da coluna como referência.
        widgets_entrada[coluna] = entrada
        
    
    # Configurando as colunas no frame 'frame_reserva' para terem pesos iguais.
    # Isso garante que as colunas se ajustem dinamicamente ao tamanho do frame.
    frame_reserva.grid_columnconfigure(0, weight=1)  # Configura a primeira coluna para expandir proporcionalmente.
    frame_reserva.grid_columnconfigure(1, weight=1)  # Configura a segunda coluna para expandir proporcionalmente.
    
    
    # Cria um botão para salvar a reserva no frame de reserva.
    # 'text="Salvar Reserva"' define o texto exibido no botão.
    # 'command=salvar_reserva' vincula o botão à função 'salvar_reserva', que será chamada quando o botão for clicado.
    # 'width=30' define a largura do botão.
    botao_salvar_reserva = Button(frame_reserva, 
                                  text="Salvar Reserva", 
                                  command=salvar_reserva, 
                                  width=30)
    
    # Posiciona o botão no frame usando o método 'grid'.
    # 'row=i+1' coloca o botão na linha seguinte após os widgets de entrada.
    # 'sticky="EW"' faz com que o botão se expanda horizontalmente.
    # 'padx=5' e 'pady=10' adicionam preenchimento ao redor do botão.
    botao_salvar_reserva.grid(row=i+1, 
                              column=0, 
                              sticky="EW", 
                              padx=5, 
                              pady=10)
    
    
    # Cria outro botão para escolher automaticamente a vaga.
    # 'text="Escolher Automaticamente"' define o texto exibido no botão.
    # 'command=escolher_vaga_automaticamente' vincula o botão à função 'escolher_vaga_automaticamente'.
    # 'width=20' define a largura do botão.
    botao_salvar_automaticamente = Button(frame_reserva, 
                                          text="Escolher Automaticamente", 
                                          command=escolher_vaga_automaticamente, 
                                          width=20)
    
    # Posicionamento do botão 'botao_salvar_automaticamente' no frame de reserva.
    botao_salvar_automaticamente.grid(row=i+1,  # 'row=i+1' coloca o botão na linha seguinte aos widgets de entrada.
                                      column=1,  # 'column=1' coloca o botão na segunda coluna do grid.
                                      sticky="EW",  # 'sticky="EW"' faz com que o botão se expanda horizontalmente.
                                      padx=5,  # 'padx=5' adiciona um preenchimento horizontal de 5 pixels.
                                      pady=10)  # 'pady=10' adiciona um preenchimento vertical de 10 pixels.
    
    
    # Limpa o campo de entrada associado à "Data de Entrada".
    widgets_entrada["Data de Entrada"].delete(0, END)  # Remove todos os caracteres existentes desde o índice 0 até o final.

    # Insere a data atual no campo de entrada "Data de Entrada".
    data_atual = datetime.now().strftime("%d/%m/%Y")  # Obtém a data atual e a formata como 'dia/mês/ano'.
    widgets_entrada["Data de Entrada"].insert(0, data_atual)  # Insere a data formatada no índice 0 do campo de entrada.

    # Chama a função 'obter_hora_atual' para atualizar o campo de entrada "Hora de Entrada" com a hora atual.
    obter_hora_atual(widgets_entrada)  # Passa o dicionário 'widgets_entrada' como argumento para a função.
    
    
    # Definindo a função 'atualizar_vagas_no_painel' que 
    # aceita 'registros_vagas' como argumento.
    # Esta função é responsável por atualizar a interface gráfica do
    # painel de estacionamento com as informações mais recentes das vagas.
    def atualizar_vagas_no_painel(registros_vagas):
        
        # Este loop itera sobre todos os widgets (elementos da
        # interface gráfica) que estão presentes na janela 'janela_painel'.
        for widget in janela_painel.winfo_children():
            
            # 'winfo_children()' retorna uma lista de todos os widgets filhos da 'janela_painel'.
            # A função verifica se o widget atual não é parte dos frames 'frame_busca' ou 'frame_reserva'.
            # Esses frames contêm elementos importantes da interface que não devem ser removidos.
            if widget not in (frame_busca, frame_reserva):
                
                # Se o widget atual não faz parte dos frames essenciais 
                # (busca ou reserva), ele é removido da janela.
                # 'widget.destroy()' é um método que remove o widget da interface gráfica.
                widget.destroy()


        # Criando um rótulo (Label) para servir como cabeçalho, mostrando a estrutura da seção de vagas.
        # Este rótulo funcionará como um cabeçalho para exibir a organização das vagas do estacionamento.
        label_cabecalho = Label(janela_painel, 
                                text="Seção/vaga 1     2      3     4      5", # A string "Seção/vaga 1     2      3     4      5" indica as seções e números das vagas,
                                font="Arial 16") # 'Arial 16' é o tipo e tamanho da fonte. 

        # Adicionando o rótulo ao layout da janela 'janela_painel' usando o método 'pack'.
        # 'pack' é um gerenciador de geometria no Tkinter que organiza os
        # widgets em blocos antes de colocá-los na janela pai.
        label_cabecalho.pack()

        # Neste loop, iteramos sobre cada seção e suas vagas no dicionário 'registros_vagas'.
        # 'registros_vagas' é um dicionário onde cada chave é uma seção do
        # estacionamento e o valor é um dicionário das vagas.
        for secao, v in registros_vagas.items():
            
            # Criando um novo quadro (Frame) para cada seção dentro da janela 'janela_painel'.
            # Um Frame é um container que pode conter outros widgets.
            frame_secao = Frame(janela_painel)

            # Criando um rótulo para a seção atual e adicionando ao quadro da seção.
            # O texto do rótulo é o nome da seção, alinhado à esquerda com
            # 19 caracteres de espaço ('{:<19}'.format(secao)).
            label_secao = Label(frame_secao, 
                                text="{:<19}".format(secao), 
                                font="Arial 16")
            
            # Empacotando o rótulo no lado esquerdo do quadro da seção.
            label_secao.pack(side=LEFT)

            # Agora, iteramos sobre cada vaga na seção.
            # Esta iteração é feita para cada seção individualmente, processando 
            # as vagas de número 1 a 5.
            for lugar in range(1, 6):
                
                # Obtemos o status da vaga.
                # 'status' recebe o valor correspondente à vaga atual na
                # seção, que pode indicar se a vaga está ocupada ou livre.
                status = v[str(lugar)]

                # Criamos um rótulo para a vaga e o adicionamos ao quadro da seção.
                # Cada vaga é representada por um rótulo que mostra seu status atual.
                label_lugar = Label(frame_secao, text=status, font="Arial 16")
                
                # O método 'pack' com 'side=LEFT' organiza os rótulos das 
                # vagas em linha, lado a lado, dentro do quadro da seção.
                label_lugar.pack(side=LEFT)

            # Finalmente, adicionamos o quadro da seção completo ao layout
            # da janela 'janela_painel'.
            # Após configurar todos os rótulos de vagas dentro de um 
            # quadro de seção, o quadro inteiro é adicionado à janela principal.
            frame_secao.pack()
            
            
    # Obtém as vagas atualizadas do arquivo Excel.
    # Chama a função 'obter_vagas_do_excel_filtro' para atualizar as 
    # informações sobre as vagas de estacionamento.
    registros_vagas = obter_vagas_do_excel_filtro()

    # Cria o cabeçalho do painel com os números das vagas de 1 a 5.
    # Um rótulo é criado para servir como cabeçalho, indicando os números das vagas.
    label_cabecalho = Label(janela_painel, 
                            text="Seção/vaga 1     2      3     4      5", 
                            font="Arial 16")
    
    # O método 'pack' é utilizado para adicionar este rótulo à janela do painel.
    label_cabecalho.pack()
                
    
    # Verifica se o dicionário 'registros_vagas' não é None. 
    # Esta verificação é importante para garantir que haja dados a serem exibidos.
    if registros_vagas is not None:

        # Itera sobre cada par chave-valor (seção-vagas) no dicionário 'registros_vagas'.
        # Para cada seção no dicionário, o loop a seguir irá criar uma representação visual.
        for secao, v in registros_vagas.items():

            # Cria um novo quadro (Frame) na janela 'janela_painel' para cada seção de vagas.
            # Um quadro é uma área retangular que pode conter outros widgets.
            frame_secao = Frame(janela_painel)

            # Cria um rótulo (Label) com o nome da seção e o adiciona ao quadro da seção.
            # O texto do rótulo é o nome da seção, formatado para ocupar 19 caracteres e alinhado à esquerda.
            label_secao = Label(frame_secao, 
                                text="{:<19}".format(secao), 
                                font="Arial 16")
            
            # O rótulo é posicionado à esquerda dentro do quadro da seção.
            label_secao.pack(side=LEFT)


            # Itera sobre as vagas na seção atual.
            # Este loop passa por cada vaga numerada de 1 a 5 dentro da
            # seção específica do estacionamento.
            for lugar in range(1, 6):

                # Obtém o status da vaga (ocupado ou livre).
                # O status de cada vaga (indicado como '-' para livre ou 'X' 
                # para ocupado) é obtido do dicionário 'registros_vagas'.
                status = v[str(lugar)]

                # Cria um rótulo (Label) para cada vaga com o status da vaga e o
                # adiciona ao quadro da seção.
                # Um rótulo é criado para representar visualmente o status da 
                # vaga na interface gráfica.
                label_lugar = Label(frame_secao, text=status, font="Arial 16")
                
                # Este rótulo é adicionado ao quadro que representa a seção, 
                # alinhado à esquerda, ao lado dos outros rótulos de vagas.
                label_lugar.pack(side=LEFT)

            # Adiciona o quadro da seção (completo com rótulos de 
            # seção e vaga) à janela 'janela_painel'.
            # Após adicionar rótulos para todas as vagas na seção, o 
            # quadro completo da seção é adicionado à janela principal do painel.
            frame_secao.pack()
     
    
    # Chama a função 'buscar_vagas' para buscar as vagas mais recentes
    # do arquivo Excel e atualizar o painel.
    # Esta função lê as informações de vagas de um 
    # arquivo e atualiza a interface gráfica com esses dados.
    buscar_vagas()

    # Atualiza a interface gráfica para processar todos os eventos pendentes.
    # Este método garante que todas as ações pendentes na fila de eventos
    # da interface gráfica sejam processadas e atualizadas.
    janela_painel.update_idletasks()

    # Obtém a largura atual da janela do painel.
    # Esta função retorna a largura atual da janela 'janela_painel' em pixels.
    largura = janela_painel.winfo_width()

    # Obtém a altura atual da janela do painel.
    # Similarmente, esta função retorna a altura atual da janela 'janela_painel' em pixels.
    altura = janela_painel.winfo_height()

    # Calcula a posição x para centralizar a janela na tela.
    # A posição x é calculada para que a janela apareça centralizada 
    # horizontalmente na tela do usuário.
    posicao_x = (janela_painel.winfo_screenwidth() // 2) - (largura // 2)

    # Calcula a posição y para centralizar a janela na tela.
    # Similarmente, a posição y é calculada para centralizar a janela verticalmente na tela.
    posicao_y = (janela_painel.winfo_screenheight() // 2) - (altura // 2)

    # Define a geometria da janela com a largura, altura, posição x e posição y.
    # Esta linha ajusta a localização e o tamanho da janela 'janela_painel',
    # usando os valores calculados para posicao_x e posicao_y.
    # A janela será posicionada de modo que seu centro fique alinhado com o centro da tela.
    janela_painel.geometry('{}x{}+{}+{}'.format(largura, altura, posicao_x, posicao_y))
    
    
# Função para salvar a reserva
def salvar_reserva():
    
    # Define variáveis globais usadas na função
    global widgets_entrada, frame_cadastro, vagas

    # Obtenha os dados da reserva a partir dos widgets de entrada
    # Estas linhas recuperam os valores inseridos pelo usuário nos campos de entrada da interface gráfica.
    cpf_cliente = widgets_entrada["CPF"].get()  # Obtém o CPF do cliente
    nome_cliente = widgets_entrada["Nome"].get()  # Obtém o nome do cliente
    data_entrada = widgets_entrada["Data de Entrada"].get()  # Obtém a data de entrada
    hora_entrada = widgets_entrada["Hora de Entrada"].get()  # Obtém a hora de entrada
    numero_vaga = widgets_entrada["Número da Vaga"].get()  # Obtém o número da vaga
    bloco = widgets_entrada["Seção"].get()  # Obtém a seção ou bloco da vaga
    status = "Reservado"  # Define o status da reserva como 'Reservado'

    # Abre o arquivo Excel e a planilha "Reservas"
    # Esta linha carrega o arquivo 'Dados.xlsx' que contém os dados de reservas.
    workbook = openpyxl.load_workbook("Dados.xlsx")
    planilha = workbook["Reservas"]  # Seleciona a aba 'Reservas' do arquivo Excel

    # Variável para rastrear se a vaga de estacionamento já está reservada
    # Esta variável é usada para indicar se a vaga específica já foi reservada.
    vaga_encontrada = False

    # Converte os dados da reserva para o formato correto para comparação
    # Esta linha converte a string da data de entrada no formato 'dd/mm/aaaa' para um objeto datetime.
    data_entrada_datetime = datetime.strptime(data_entrada, "%d/%m/%Y")


    # Itera sobre todas as linhas da planilha "Reservas", começando 
    # da segunda linha (a primeira linha geralmente contém cabeçalhos)
    for linha in planilha.iter_rows(min_row=2, values_only=True):

        # Converte a data de entrada (localizada na terceira coluna, índice 2) de
        # string para um objeto datetime
        # Isso é feito para permitir comparações de data de forma mais eficiente.
        data_entrada_excel = datetime.strptime(str(linha[2]), "%d/%m/%Y")

        # Inicializa a variável 'data_saida_excel' como None. 
        # Esta variável armazenará a data de saída após a conversão de
        # string para datetime, se aplicável.
        data_saida_excel = None

        # Verifica se a célula correspondente à data de saída (quinta coluna, índice 4) não está vazia
        # A função 'strip()' remove espaços em branco do início e do fim da
        # string, ajudando a garantir que a célula contenha dados válidos.
        if linha[4] and str(linha[4]).strip():  

            try:
                
                # Tenta converter a string da data de saída para um objeto datetime.
                # Este passo é necessário para realizar comparações de datas posteriormente.
                data_saida_excel = datetime.strptime(str(linha[4]), "%d/%m/%Y")

            except ValueError:
                
                # Se a conversão falhar (por exemplo, se a string não 
                # estiver no formato de data esperado),
                # ignora a exceção e continua o loop. Isso é útil para
                # evitar interrupção do programa devido a formatos de data incorretos.
                pass


        # Acessa e converte para string o valor da sétima
        #coluna (índice 6) da linha atual, que representa o número da vaga reservada.
        vaga_reservada = str(linha[6])

        # Acessa e converte para string o valor da oitava coluna (índice 7) 
        # da linha atual, que representa o bloco onde a vaga está localizada.
        bloco_reservado = str(linha[7])

        # Compara a data de entrada na reserva com a data desejada pelo
        # usuário e verifica se a vaga e o bloco correspondem aos desejados.
        # A condição '(not data_saida_excel)' assegura que a reserva ainda 
        #está válida (ou seja, a data de saída ainda não ocorreu ou não foi definida).
        if data_entrada_excel == data_entrada_datetime and (not data_saida_excel) and vaga_reservada == str(numero_vaga) and bloco_reservado == bloco:
            
            # Se todas as condições forem atendidas, significa que a vaga 
            # desejada já está reservada.
            # Uma mensagem de aviso é exibida para o usuário.
            messagebox.showwarning("Reserva já Existe", "Essa vaga não está disponível.")
            
            # A variável 'vaga_encontrada' é marcada como True, indicando 
            # que a vaga já está ocupada.
            vaga_encontrada = True
            
            # Fecha a janela atual do painel.
            janela_painel.destroy()    

            # Chama a função 'exibir_painel_vagas' para atualizar o painel com as vagas disponíveis.
            exibir_painel_vagas()
            
            # Isso impede que uma reserva duplicada seja criada para a mesma vaga e período.
            return
            
            # Interrompe o loop, pois a vaga desejada já foi encontrada e está indisponível.
            break
            

        # Se uma vaga já reservada foi encontrada, a função é encerrada imediatamente.
        # Isso impede que uma reserva duplicada seja criada para a mesma vaga e período.
        if vaga_encontrada:
            return

        # Inicializa 'linha_vazia' com o valor 2, que será o ponto de 
        # partida para encontrar uma linha vazia na planilha.
        # Isso é útil para adicionar uma nova reserva na planilha.
        linha_vazia = 2

    
    # Este loop procura uma linha vazia na planilha "Reservas" para 
    # inserir os novos dados de reserva.
    while planilha.cell(row=linha_vazia, column=1).value is not None:
        
        # Incrementa 'linha_vazia' até encontrar uma linha onde a primeira
        # célula (coluna 1) está vazia (None).
        linha_vazia += 1

    # Após encontrar uma linha vazia, os detalhes da reserva são inseridos nessa linha.
    # Atribui o CPF do cliente à primeira coluna da nova linha.
    planilha.cell(row=linha_vazia, column=1).value = cpf_cliente
    
    # Atribui o nome do cliente à segunda coluna.
    planilha.cell(row=linha_vazia, column=2).value = nome_cliente
    
    # Atribui a data de entrada à terceira coluna.
    planilha.cell(row=linha_vazia, column=3).value = data_entrada
    
    # Atribui a hora de entrada à quarta coluna.
    planilha.cell(row=linha_vazia, column=4).value = hora_entrada
    
    # Insere um traço nas colunas de data e hora de saída, pois ainda não foram definidas.
    planilha.cell(row=linha_vazia, column=5).value = "-"
    planilha.cell(row=linha_vazia, column=6).value = "-"
    
    # Atribui o número da vaga à sétima coluna.
    planilha.cell(row=linha_vazia, column=7).value = numero_vaga
    
    # Atribui o bloco à oitava coluna.
    planilha.cell(row=linha_vazia, column=8).value = bloco
    
    # Inicializa a nona coluna (total a pagar) com o valor 0, indicando
    # que ainda não foi calculado.
    planilha.cell(row=linha_vazia, column=9).value = 0
    
    # Define o status da reserva para 'Reservado'.
    planilha.cell(row=linha_vazia, column=10).value = status   

    # Salva as alterações na planilha de Excel.
    workbook.save("Dados.xlsx")

    # Exibe uma mensagem de sucesso informando que a reserva foi salva.
    # show_message("Reserva salva com sucesso!")

    # Fecha a janela atual do painel.
    janela_painel.destroy()    

    # Chama a função 'exibir_painel_vagas' para atualizar o painel com as vagas disponíveis.
    exibir_painel_vagas()
    
    

# Função que escolhe a vaga automaticamente
def escolher_vaga_automaticamente():

    # Define variáveis globais para acesso e modificação dentro desta função.
    global entry_widgets, frame_cadastro

    # Obtém os dados inseridos nos campos de entrada pelo usuário.
    cpf_cliente = widgets_entrada["CPF"].get()  # Pega o CPF do cliente.
    nome_cliente = widgets_entrada["Nome"].get()  # Pega o nome do cliente.
    data_entrada = widgets_entrada["Data de Entrada"].get()  # Pega a data de entrada.
    hora_entrada = widgets_entrada["Hora de Entrada"].get()  # Pega a hora de entrada.
    numero_vaga = widgets_entrada["Número da Vaga"].get()  # Pega o número da vaga, se especificado.
    bloco = widgets_entrada["Seção"].get()  # Pega o bloco, se especificado.
    status = "Reservado"  # Define o status da reserva como 'Reservado'.

    # Carrega a planilha do Excel e acessa a aba 'Reservas'.
    planilha = openpyxl.load_workbook("Dados.xlsx")
    aba_reservas = planilha["Reservas"]

    # Converte a string da data de entrada em um objeto datetime para comparação.
    data_entrada_datetime = datetime.strptime(data_entrada, "%d/%m/%Y")

    # Inicializa variáveis para armazenar a vaga e o bloco disponíveis.
    vaga_disponivel = None  # Inicialmente, nenhuma vaga está selecionada.
    bloco_disponivel = None  # Inicialmente, nenhum bloco está selecionado.


    # Verifica se o usuário não especificou um número de vaga e
    # um bloco específico para a reserva.
    if not numero_vaga and not bloco:

        # Cria uma lista para armazenar as informações das vagas que já estão ocupadas.
        vagas_ocupadas = []

        # Itera por cada linha da aba 'Reservas' na planilha, começando da 
        # segunda linha (ignorando cabeçalhos).
        for linha in aba_reservas.iter_rows(min_row=2, values_only=True):

            # Inicializa variáveis para armazenar as datas de entrada e
            # saída, começando com None (sem valor).
            data_entrada_excel = None
            data_saida_excel = None

            # Verifica se a célula correspondente à data de entrada na 
            # linha atual contém algum valor.
            if linha[2] and str(linha[2]).strip():

                # Tenta converter a data de entrada de uma string para um objeto datetime.
                try:
                    
                    data_entrada_excel = datetime.strptime(str(linha[2]), "%d/%m/%Y")
                    
                except ValueError:
                    
                    # Caso a conversão falhe (formato incorreto, por exemplo), a
                    # exceção é capturada e o loop continua.
                    pass


            # Verifica se a célula correspondente à data de saída na
            # linha atual contém algum valor.
            if linha[4] and str(linha[4]).strip():

                
                try:
                    
                    # Tenta converter a data de saída de uma string para um objeto datetime.
                    data_saida_excel = datetime.strptime(str(linha[4]), "%d/%m/%Y")
                    
                except ValueError:
                    
                    # Se a conversão falhar, ignora a exceção e 
                    # continua o loop, mantendo data_saida_excel como None.
                    pass

            # Verifica duas condições para determinar se a vaga está ocupada:
            # 1. Se a data de saída é None e a data de entrada registrada é a mesma que a data de entrada da reserva.
            # 2. Ou se a data de saída registrada é a mesma que a data de entrada da reserva.
            if (data_saida_excel is None and data_entrada_excel.date() == data_entrada_datetime.date()) or (data_saida_excel and data_saida_excel.date() == data_entrada_datetime.date()):

                # Se alguma das condições acima for verdadeira, significa que a vaga
                # está ocupada na data desejada.
                # Adiciona o número da vaga (obtido da sétima coluna 
                #da linha) e o bloco (obtido da oitava coluna da linha) à lista
                # de vagas ocupadas.
                vagas_ocupadas.append((linha[6], linha[7]))


        
        # Itera sobre uma lista de blocos possíveis, que são 'A', 'B', 'C' e 'D'.
        for bloco_candidato in ['A', 'B', 'C', 'D']:

            # Dentro de cada bloco, itera sobre os números das vagas, que vão de 1 a 5.
            for vaga_candidata in range(1, 6):

                # Converte o número da vaga candidata para string para facilitar a
                # comparação com os dados da planilha.
                vaga_candidata = str(vaga_candidata)

                # Verifica se a combinação de bloco e vaga candidata não
                # está na lista de vagas ocupadas.
                # 'vagas_ocupadas' é uma lista de tuplas, onde cada tupla 
                # contém o número da vaga e o bloco correspondente.
                if (vaga_candidata, bloco_candidato) not in vagas_ocupadas:

                    # Se a vaga não estiver ocupada, atribui o número da vaga e
                    # o bloco à variável 'vaga_disponivel' e 'bloco_disponivel', respectivamente.
                    # Isso significa que encontramos uma vaga livre e podemos atribuí-la para a reserva.
                    vaga_disponivel = vaga_candidata
                    bloco_disponivel = bloco_candidato

                    # Uma vez que uma vaga disponível foi encontrada, 
                    # interrompe o loop interno para não continuar procurando
                    # vagas no mesmo bloco.
                    break

            # Verifica se uma vaga disponível foi encontrada no loop interno.
            # Se sim, interrompe também o loop externo para não continuar 
            # procurando em outros blocos.
            if vaga_disponivel:
                break


        # Se não encontramos uma vaga disponível após percorrer todos os blocos e vagas
        if not vaga_disponivel:

            # Exibe uma mensagem de aviso informando que todas as vagas estão ocupadas
            messagebox.showwarning("Todas as Vagas Ocupadas", "Não há vagas disponíveis.")

            # Interrompe a execução da função
            return

        # Define o número da vaga e o bloco para a reserva
        # com a vaga e o bloco disponíveis encontrados
        numero_vaga = vaga_disponivel
        bloco = bloco_disponivel

    # Senão se o número da vaga e o bloco já foram especificados, inicia
    # a verificação para garantir que a vaga não está ocupada.
    else:

        # Itera sobre todas as linhas da aba "Reservas" na planilha, 
        # começando da segunda linha, pois a primeira contém os cabeçalhos.
        for linha in aba_reservas.iter_rows(min_row=2, values_only=True):

            # Inicializa as variáveis para armazenar as datas de entrada e saída da reserva como None. 
            # Isso serve para tratar o caso em que esses campos podem estar vazios na planilha.
            data_entrada_excel = None
            data_saida_excel = None

            # Verifica se há uma data de entrada registrada na linha
            # atual (coluna correspondente ao índice 2) e se essa data não é uma string vazia.
            if linha[2] and str(linha[2]).strip():
                
                
                try:
                    
                    # Tenta converter o valor da coluna de data de 
                    # entrada (índice 2) para um objeto datetime.
                    data_entrada_excel = datetime.strptime(str(linha[2]), "%d/%m/%Y")
                                                           
                except ValueError:
                                                           
                    # Se a conversão falhar (por exemplo, devido a um formato de
                    # data incorreto), a exceção é capturada e a execução do loop continua.
                    # Isso evita que o programa pare devido a um erro na conversão da data.
                    pass


            # Verifica se a data de saída na linha atual não é nula ou uma string vazia
            if linha[4] and str(linha[4]).strip():

                
                try:
                                                           
                    # Tenta converter a data de saída para um objeto datetime
                    data_saida_excel = datetime.strptime(str(linha[4]), "%d/%m/%Y")
                                                           
                except ValueError:  # Caso a conversão falhe, não faz nada
                    pass

            # Verifica se a vaga e o bloco na linha atual da planilha 
            # correspondem à vaga e ao bloco que o usuário deseja reservar.
            # Além disso, verifica se as datas de entrada e saída da reserva 
            # atual na planilha coincidem com a data desejada pelo usuário.
            if linha[6] == numero_vaga and linha[7] == bloco and ((data_saida_excel is None and data_entrada_excel.date() == data_entrada_datetime.date()) or (data_saida_excel and data_saida_excel.date() == data_entrada_datetime.date())):
                
                # Se as condições forem verdadeiras, significa que a vaga
                # especificada pelo usuário já está ocupada.

                # Exibe uma mensagem de aviso para o usuário indicando que a vaga selecionada já está ocupada.
                # Isso informa o usuário que ele precisa escolher outra vaga ou outro período para a reserva.
                messagebox.showwarning("Vaga Ocupada", "A vaga selecionada não está disponível.")

                # Interrompe a execução da função, pois a vaga desejada já está ocupada.
                # Isso evita que o sistema prossiga com uma reserva inválida.
                return


    # Encontra a primeira linha vazia na aba "Reservas"
    linha_vazia = 2
    
    # Um loop while que continua até encontrar uma linha vazia na planilha 'Reservas'.
    while aba_reservas.cell(row=linha_vazia, column=1).value is not None:
        
        # Incrementa 'linha_vazia' para checar a próxima linha.
        linha_vazia += 1

    # Uma vez que uma linha vazia é encontrada, insere os detalhes da reserva nas células apropriadas.
    aba_reservas.cell(row=linha_vazia, column=1).value = cpf_cliente  # Insere o CPF do cliente.
    aba_reservas.cell(row=linha_vazia, column=2).value = nome_cliente  # Insere o nome do cliente.
    aba_reservas.cell(row=linha_vazia, column=3).value = data_entrada  # Insere a data de entrada da reserva.
    aba_reservas.cell(row=linha_vazia, column=4).value = hora_entrada  # Insere a hora de entrada.
    aba_reservas.cell(row=linha_vazia, column=5).value = "-"  # Marca a data de saída como indefinida.
    aba_reservas.cell(row=linha_vazia, column=6).value = "-"  # Marca a hora de saída como indefinida.
    aba_reservas.cell(row=linha_vazia, column=7).value = numero_vaga  # Insere o número da vaga reservada.
    aba_reservas.cell(row=linha_vazia, column=8).value = bloco  # Insere o bloco onde a vaga está localizada.
    aba_reservas.cell(row=linha_vazia, column=9).value = 0  # Insere 0 como valor inicial para o total a pagar.
    aba_reservas.cell(row=linha_vazia, column=10).value = status  # Insere o status da reserva (geralmente "Reservado").

    # Salva as alterações feitas na planilha Excel.
    planilha.save("Dados.xlsx")

    # Limpa os campos de entrada na interface gráfica após o salvamento dos dados.
    for widget in widgets_entrada.values():
        
        # Limpa cada widget de entrada.
        widget.delete(0, tk.END)

    # Exibe uma mensagem de informação indicando que a reserva foi efetuada com sucesso.
    messagebox.showinfo("Reserva Efetuada", "Reserva efetuada com sucesso.")

    # Fecha a janela atual do painel
    janela_painel.destroy()    
    
    # reabre o painel de vagas para atualizar as informações exibidas.
    exibir_painel_vagas()
        
        
   
    

def reservas():
    
    print("Tela Reservas...")

def tela_clientes():
    
    print("Exibir Tela Clientes...")




# Cria uma instância da classe Tk
janela = tk.Tk()

# Define o título da janela
janela.title("Sistema de Pedidos")

# Configura o fundo da janela como branco
janela.configure(bg="white")

# Define a janela em modo de tela cheia
janela.attributes("-fullscreen", True)

# Importa as classes Tk e Label da biblioteca tkinter.
from tkinter import Tk, Label

# Importa as classes ImageTk e Image da biblioteca Python Imaging Library (PIL).
from PIL import ImageTk, Image

# Inicia um bloco try-except para capturar possíveis exceções
# durante a execução do código.
try:
                                           
    # Usa a função open do módulo Image para carregar a imagem 'estacionamento.jpg' do disco.
    imagem_carregada = Image.open("estacionamento.jpg")

    # Converte a imagem para um formato compatível com o Tkinter, criando um objeto PhotoImage.
    imagem = ImageTk.PhotoImage(imagem_carregada)

    # Cria um objeto Label na janela 'janela'. Este Label será usado para exibir a imagem.
    rotulo_imagem = Label(janela)

    # Armazena uma referência da imagem no rótulo. Isso é necessário 
    # porque o Tkinter pode descartar a imagem se não houver uma referência forte para ela.
    rotulo_imagem.image = imagem

    # Configura o rótulo para exibir a imagem.
    rotulo_imagem.configure(image=imagem)

    # Adiciona o rótulo à janela, usando o gerenciador de layout pack.
    rotulo_imagem.pack()

# Se ocorrer uma exceção durante a execução do bloco try, ela é 
# capturada aqui. A exceção é impressa na saída padrão.
except Exception as e:
    print(f"Erro ao carregar a imagem: {e}")

    
# Cria um rótulo (Label) na janela 'janela' com uma mensagem de boas-vindas.
# Define a fonte do texto como Arial de tamanho 16 e o fundo do rótulo como branco.
mensagem_label = tk.Label(janela, 
                          text="Bem-vindo ao Controle de Vagas Estacionamento", 
                          font=("Arial", 16), bg="white")

# Adiciona o rótulo à janela, usando o gerenciador de layout pack.
# A opção 'pady' adiciona um espaçamento vertical de 50 pixels acima e abaixo do rótulo.
mensagem_label.pack(pady=50)

    
# Cria um frame (container) na janela 'janela' para organizar os botões.
# O fundo do frame é definido como branco.
button_frame = tk.Frame(janela, bg="white")

# Adiciona o frame à janela, usando o gerenciador de layout pack.
button_frame.pack()


# Cria um botão no 'button_frame'.
# Define o texto do botão, a fonte, o comando a ser executado quando o botão é clicado,
# e a largura do botão.
# O comando 'exibir_painel_vagas' deve ser uma função definida em outro lugar no código.
botao_painel = tk.Button(button_frame, 
                         text="Exibir Painel de Vagas", 
                         font=("Arial", 12),
                         command=exibir_painel_vagas, 
                         width=40)

# Adiciona o botão ao frame, usando o gerenciador de layout pack.
# A opção 'side=tk.TOP' posiciona o botão no topo do frame.
# 'padx' e 'pady' adicionam espaçamentos horizontal e vertical, respectivamente.
# 'ipadx' e 'ipady' adicionam espaçamentos internos no botão, horizontal e verticalmente.
botao_painel.pack(side=tk.TOP, padx=50, pady=10, ipadx=20, ipady=10)

# Cria um botão para a tela de clientes no 'button_frame'.
# O texto do botão é "Clientes", e a fonte definida é Arial tamanho 12.
# O comando 'tela_clientes' é chamado quando o botão é clicado, e a largura do botão é definida como 40.
botao_clientes = tk.Button(button_frame, 
                           text="Clientes", 
                           font=("Arial", 12),
                           command=tela_clientes,
                           width=40)

# Adiciona o botão 'botao_clientes' ao frame, alinhando-o no topo.
# Define espaçamentos horizontal e vertical de 50 e 10 pixels, respectivamente.
# 'ipadx' e 'ipady' adicionam espaçamento interno no botão, horizontal e verticalmente.
botao_clientes.pack(side=tk.TOP, padx=50, pady=10, ipadx=20, ipady=10)

# Cria um botão para a tela de reservas no mesmo frame.
# O texto do botão é "Reservas", a fonte é Arial tamanho 12, e a largura é 40.
# O comando 'reservas' é associado a este botão para ser executado quando clicado.
botao_Reservas = tk.Button(button_frame, 
                           text="Reservas", 
                           font=("Arial", 12), 
                           command=reservas, 
                           width=40)

# Posiciona o botão 'botao_Reservas' no frame, alinhando-o no topo.
# Define os mesmos espaçamentos horizontal e vertical que os outros botões.
botao_Reservas.pack(side=tk.TOP, padx=50, pady=10, ipadx=20, ipady=10)


    
# Cria um botão para sair do sistema.
# O texto do botão é "Sair", a fonte é Arial tamanho 12, e a largura é definida como 40.
# O comando 'janela.destroy' é associado a este botão, o que fechará a janela quando clicado.
sair_button = tk.Button(button_frame, 
                        text="Sair",                    
                        font=("Arial", 12), 
                        command=janela.destroy, 
                        width=40)

# Posiciona o botão 'sair_button' no frame, alinhando-o no topo.
# Mantém o mesmo padrão de espaçamentos dos botões anteriores.
# 'padx' e 'pady' adicionam espaçamentos horizontal e vertical, respectivamente.
# 'ipadx' e 'ipady' adicionam espaçamentos internos no botão, horizontal e verticalmente.
sair_button.pack(side=tk.TOP, padx=50, pady=10, ipadx=20, ipady=10)
    
# Inicia o loop principal da aplicação
janela.mainloop()
    
