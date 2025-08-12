# Importa tudo (*) do módulo tkinter. Tkinter é uma 
# biblioteca padrão do Python usada para criar interfaces gráficas.
from tkinter import *

# Importa o módulo tkinter e o renomeia como 'tk', uma 
# convenção comum para facilitar a referência a este módulo.
import tkinter as tk

# Do módulo tkinter, importa classes específicas. Tk é a 
# classe base para a criação de janelas, Frame é um contêiner para outros widgets,
# Label é um widget que exibe texto ou imagens, Entry é um 
# widget para entrada de texto de linha única,
# Button é um widget de botão e messagebox é um módulo para caixas de diálogo padrão.
from tkinter import Tk, Frame, Label, Entry, Button, messagebox

# Importa o módulo ttk do tkinter, que oferece acesso a 
# widgets com um visual melhorado e aprimorado.
from tkinter import ttk

# Importa a biblioteca openpyxl, usada para ler e escrever em arquivos do Excel.
import openpyxl

# Importa a biblioteca pandas e a renomeia como 'pd'. Pandas é uma
# biblioteca popular para análise de dados.
import pandas as pd

# Do módulo openpyxl, importa especificamente as classes Workbook e load_workbook.
# Workbook é usado para criar novos arquivos do Excel, e load_workbook é
# usado para abrir arquivos existentes.
from openpyxl import Workbook, load_workbook

# Importa as classes date e datetime do módulo datetime. 
# 'date' é usada para trabalhar com datas (ano, mês, dia), enquanto
# 'datetime' é para datas e horas.
from datetime import date, datetime

# Declaração de três variáveis globais. As variáveis globais são acessíveis em todo o código, não apenas no escopo onde foram declaradas.
entry_data_pesquisa = None  # Variável para armazenar um widget de entrada de dados (para pesquisa).
window_painel = None  # Variável para armazenar uma janela do painel (janela principal ou um painel específico).
registros_vagas = {}  # Dicionário para armazenar registros de vagas (informações sobre vagas de estacionamento).


# Define a função 'exibir_painel_vagas'. Esta função não tem
# parâmetros e é usada para exibir uma janela de painel de estacionamento.
def exibir_painel_vagas():

    # Declara variáveis como globais para que as alterações feitas nessas variáveis dentro da função sejam refletidas globalmente.
    # Essas variáveis são usadas para armazenar widgets e outras informações importantes para a interface gráfica.
    global entrada_data_pesquisa, janela_painel, registros_vagas, frame_reserva, widgets_entrada
    
    janela_painel = Toplevel()
    
    janela_painel.title("Painel de Estacionamento")
    
    

def reservas():
    
    print("Tela Reservas...")

def tela_clientes():
    
    print("Exibir Tela Clientes...")




# Cria uma instância da classe Tk
janela = tk.Tk()

# Define o título da janela
janela.title("Sistema de Pedidos")

# Configura o fundo da janela como branco
janela.configure(bg="white")

# Define a janela em modo de tela cheia
janela.attributes("-fullscreen", True)

# Importa as classes Tk e Label da biblioteca tkinter.
from tkinter import Tk, Label

# Importa as classes ImageTk e Image da biblioteca Python Imaging Library (PIL).
from PIL import ImageTk, Image

# Inicia um bloco try-except para capturar possíveis exceções
# durante a execução do código.
try:
                                           
    # Usa a função open do módulo Image para carregar a imagem 'estacionamento.jpg' do disco.
    imagem_carregada = Image.open("estacionamento.jpg")

    # Converte a imagem para um formato compatível com o Tkinter, criando um objeto PhotoImage.
    imagem = ImageTk.PhotoImage(imagem_carregada)

    # Cria um objeto Label na janela 'janela'. Este Label será usado para exibir a imagem.
    rotulo_imagem = Label(janela)

    # Armazena uma referência da imagem no rótulo. Isso é necessário 
    # porque o Tkinter pode descartar a imagem se não houver uma referência forte para ela.
    rotulo_imagem.image = imagem

    # Configura o rótulo para exibir a imagem.
    rotulo_imagem.configure(image=imagem)

    # Adiciona o rótulo à janela, usando o gerenciador de layout pack.
    rotulo_imagem.pack()

# Se ocorrer uma exceção durante a execução do bloco try, ela é 
# capturada aqui. A exceção é impressa na saída padrão.
except Exception as e:
    print(f"Erro ao carregar a imagem: {e}")

    
# Cria um rótulo (Label) na janela 'janela' com uma mensagem de boas-vindas.
# Define a fonte do texto como Arial de tamanho 16 e o fundo do rótulo como branco.
mensagem_label = tk.Label(janela, 
                          text="Bem-vindo ao Controle de Vagas Estacionamento", 
                          font=("Arial", 16), bg="white")

# Adiciona o rótulo à janela, usando o gerenciador de layout pack.
# A opção 'pady' adiciona um espaçamento vertical de 50 pixels acima e abaixo do rótulo.
mensagem_label.pack(pady=50)

    
# Cria um frame (container) na janela 'janela' para organizar os botões.
# O fundo do frame é definido como branco.
button_frame = tk.Frame(janela, bg="white")

# Adiciona o frame à janela, usando o gerenciador de layout pack.
button_frame.pack()


# Cria um botão no 'button_frame'.
# Define o texto do botão, a fonte, o comando a ser executado quando o botão é clicado,
# e a largura do botão.
# O comando 'exibir_painel_vagas' deve ser uma função definida em outro lugar no código.
botao_painel = tk.Button(button_frame, 
                         text="Exibir Painel de Vagas", 
                         font=("Arial", 12),
                         command=exibir_painel_vagas, 
                         width=40)

# Adiciona o botão ao frame, usando o gerenciador de layout pack.
# A opção 'side=tk.TOP' posiciona o botão no topo do frame.
# 'padx' e 'pady' adicionam espaçamentos horizontal e vertical, respectivamente.
# 'ipadx' e 'ipady' adicionam espaçamentos internos no botão, horizontal e verticalmente.
botao_painel.pack(side=tk.TOP, padx=50, pady=10, ipadx=20, ipady=10)

# Cria um botão para a tela de clientes no 'button_frame'.
# O texto do botão é "Clientes", e a fonte definida é Arial tamanho 12.
# O comando 'tela_clientes' é chamado quando o botão é clicado, e a largura do botão é definida como 40.
botao_clientes = tk.Button(button_frame, 
                           text="Clientes", 
                           font=("Arial", 12),
                           command=tela_clientes,
                           width=40)

# Adiciona o botão 'botao_clientes' ao frame, alinhando-o no topo.
# Define espaçamentos horizontal e vertical de 50 e 10 pixels, respectivamente.
# 'ipadx' e 'ipady' adicionam espaçamento interno no botão, horizontal e verticalmente.
botao_clientes.pack(side=tk.TOP, padx=50, pady=10, ipadx=20, ipady=10)

# Cria um botão para a tela de reservas no mesmo frame.
# O texto do botão é "Reservas", a fonte é Arial tamanho 12, e a largura é 40.
# O comando 'reservas' é associado a este botão para ser executado quando clicado.
botao_Reservas = tk.Button(button_frame, 
                           text="Reservas", 
                           font=("Arial", 12), 
                           command=reservas, 
                           width=40)

# Posiciona o botão 'botao_Reservas' no frame, alinhando-o no topo.
# Define os mesmos espaçamentos horizontal e vertical que os outros botões.
botao_Reservas.pack(side=tk.TOP, padx=50, pady=10, ipadx=20, ipady=10)


    
# Cria um botão para sair do sistema.
# O texto do botão é "Sair", a fonte é Arial tamanho 12, e a largura é definida como 40.
# O comando 'janela.destroy' é associado a este botão, o que fechará a janela quando clicado.
sair_button = tk.Button(button_frame, 
                        text="Sair",                    
                        font=("Arial", 12), 
                        command=janela.destroy, 
                        width=40)

# Posiciona o botão 'sair_button' no frame, alinhando-o no topo.
# Mantém o mesmo padrão de espaçamentos dos botões anteriores.
# 'padx' e 'pady' adicionam espaçamentos horizontal e vertical, respectivamente.
# 'ipadx' e 'ipady' adicionam espaçamentos internos no botão, horizontal e verticalmente.
sair_button.pack(side=tk.TOP, padx=50, pady=10, ipadx=20, ipady=10)
    
# Inicia o loop principal da aplicação
janela.mainloop()
    
