import pandas as opcoesDoPanda
import os
from openpyxl.styles import Color, PatternFill, Font, Border, Side
from openpyxl.styles import colors
from openpyxl.cell import Cell
#Caminho onde estão os arquivos
#caminhoArquivos = r"C:\Users\Aluno\Desktop\Curso RPA\Consolindando_python_excel\Excel"
caminhoArquivos = r"C:\python_projetos\python_rpa_projetos\Consolidando_python_excel"

#Variável onde estão todos os arquivos
listaArquivos = os.listdir(caminhoArquivos)

print(listaArquivos)

#Listando todos os arquivos + o caminho
listaCaminhoEArquivoExcel = [caminhoArquivos + '\\' + arquivo for arquivo in listaArquivos if arquivo[-4:] == 'xlsx' ]

print("----- #### -------- ###### ---------")
print("----- #### -------- ###### ---------")
print(listaCaminhoEArquivoExcel)

#Criando um DataFrame para trabalhar com os dados dos arquivos
dadosArquivo = opcoesDoPanda.DataFrame()

#Copiando todos os dados dos arquivos em DadosArquivo
listaDados = []  # Lista para armazenar os DataFrames temporariamente
for arquivo in listaCaminhoEArquivoExcel:
    dados = opcoesDoPanda.read_excel(arquivo)
    #dadosArquivo = dadosArquivo.append(dados) descontinuado append
    # Adicionando o DataFrame à lista
    listaDados.append(dados)
dadosArquivo = opcoesDoPanda.concat(listaDados, ignore_index=True)

#Criando uma nova planilha e passando os dados dos arquivos
dadosArquivo.to_excel(r"C:\python_projetos\python_rpa_projetos\Consolidando_python_excel\Arquivo Consolidado.xlsx")
##########################################################################################
from openpyxl import load_workbook
from openpyxl import Workbook
caminhoArquivoDadosSistema = r"C:\\python_projetos\\python_rpa_projetos\\Consolidando_python_excel\\Arquivo Consolidado.xlsx"
planilhaDadosSistema = load_workbook(filename=caminhoArquivoDadosSistema)

#Seleciona a sheet1
sheetPlanilhaDadosSistema = planilhaDadosSistema['Sheet1']

#Deleta a coluna A
sheetPlanilhaDadosSistema.delete_cols(1)

#Renomeia o nome da sheet
sheetPlanilhaDadosSistema.title = 'Dados Consolidados'
#Aumenta a largura das colunas A e B
sheetPlanilhaDadosSistema.column_dimensions['A'].width = 35
sheetPlanilhaDadosSistema.column_dimensions['B'].width = 40



#Criando preenchimento cor Azul

'''corCinza = PatternFill(start_color='00FFFFFF',
                    end_color='00FFFFFF',
                    fill_type='solid')
'''
corAzul = PatternFill(start_color='00FFCCFF',
                    end_color='00FFCCFF',
                    fill_type='solid')                    
#Criando preenchimento cor Amarelo
corAmarelo = PatternFill(start_color='00FFFFCC',
                    end_color='00FFFFCC',
                    fill_type='solid')

bfFina = Side(style='thin', color='00000000')
borda=Border(left=bfFina, right=bfFina, top=bfFina, bottom=bfFina)
sheetPlanilhaDadosSistema['A1'].fill = corAzul
sheetPlanilhaDadosSistema['B1'].fill = corAzul
sheetPlanilhaDadosSistema['C1'].fill = corAzul

sheetPlanilhaDadosSistema['A1'].border = borda
sheetPlanilhaDadosSistema['B1'].border = borda
sheetPlanilhaDadosSistema['C1'].border = borda
#salva a planilha com as alterações
planilhaDadosSistema.save(filename=caminhoArquivoDadosSistema)

os.startfile(caminhoArquivoDadosSistema)


