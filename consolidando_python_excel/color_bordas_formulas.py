import pandas as opcoesDoPanda
import os

#Caminho onde estão os arquivos
#caminhoArquivos = r"C:\Users\Aluno\Desktop\Curso RPA\Consolindando_python_excel\Excel"
caminhoArquivos = r"C:\python_projetos\python_rpa_projetos\Consolidando_python_excel"

#Variável onde estão todos os arquivos
listaArquivos = os.listdir(caminhoArquivos)

print(listaArquivos)

#Listando todos os arquivos + o caminho
listaCaminhoEArquivoExcel = [caminhoArquivos + '\\' + arquivo for arquivo in listaArquivos if arquivo[-4:] == 'xlsx' ]



#Criando um DataFrame para trabalhar com os dados dos arquivos
dadosArquivo = opcoesDoPanda.DataFrame()

#Copiando todos os dados dos arquivos em DadosArquivo
listaDados = []  # Lista para armazenar os DataFrames temporariamente
for arquivo in listaCaminhoEArquivoExcel:
    dados = opcoesDoPanda.read_excel(arquivo)
    #dadosArquivo = dadosArquivo.append(dados) descontinuado append
    # Adicionando o DataFrame à lista
    listaDados.append(dados)
dadosArquivo = opcoesDoPanda.concat(listaDados, ignore_index=True)

#Criando uma nova planilha e passando os dados dos arquivos
dadosArquivo.to_excel(r"C:\python_projetos\python_rpa_projetos\Consolidando_python_excel\Arquivo Consolidado.xlsx")
##########################################################################################
from openpyxl import load_workbook
from openpyxl import workbook

from openpyxl.styles import Color, PatternFill, Font, Border, Side
from openpyxl.styles import colors
from openpyxl.cell import Cell

caminhoArquivoDadosSistema = "C:\python_projetos\python_rpa_projetos\Consolidando_python_excel\Arquivo Consolidado.xlsx"
planilhaDadosSistema = load_workbook(filename=caminhoArquivoDadosSistema)

#Seleciona a sheet1
sheetPlanilhaDadosSistema = planilhaDadosSistema['Sheet1']

#Deleta a coluna A
sheetPlanilhaDadosSistema.delete_cols(1)

#Renomeia o nome da sheet
sheetPlanilhaDadosSistema.title = 'Dados Consolidados'

#Aumenta a largura das colunas A e B
sheetPlanilhaDadosSistema.column_dimensions['A'].width = 35
sheetPlanilhaDadosSistema.column_dimensions['B'].width = 40

#Criando preenchimento cor Cinza fraco
corCinza = PatternFill(start_color='00FFFFFF',
                    end_color='00FFFFFF',
                    fill_type='solid')

#Criando preenchimento cor Amarela
corAmerelo = PatternFill(start_color='00FFFFCC',
                    end_color='00FFFFCC',
                    fill_type='solid')

#Coloca borda preta na célula
bfFina = Side(border_style='thin', color='000000')
borda = Border(left=bfFina, right=bfFina, top=bfFina, bottom=bfFina)

#Colocando cores e borda e fundo nas celulas A1, B1 e C1
sheetPlanilhaDadosSistema['A1'].fill = corAmerelo
sheetPlanilhaDadosSistema['B1'].fill = corAmerelo
sheetPlanilhaDadosSistema['C1'].fill = corAmerelo
sheetPlanilhaDadosSistema['A1'].border = borda
sheetPlanilhaDadosSistema['B1'].border = borda
sheetPlanilhaDadosSistema['C1'].border = borda

#Dando um for linha por linha para pintar as bordas e o fundo
for linha in range(2, len(sheetPlanilhaDadosSistema['A']) + 1):
        
        #Criando celula A2, B2, C2 e assim por diante de acordo com o número da linha
        celulaColunaA = 'A' + str(linha)
        celulaColunaB = 'B' + str(linha)
        celulaColunaC = 'C' + str(linha)
                   
        #Colorindo o fundo de cinza claro das linhas
        sheetPlanilhaDadosSistema[celulaColunaA].fill = corCinza
        sheetPlanilhaDadosSistema[celulaColunaB].fill = corCinza
        sheetPlanilhaDadosSistema[celulaColunaC].fill = corCinza
                   
        #Colocando a borda preta nas linhas
        sheetPlanilhaDadosSistema[celulaColunaA].border = borda
        sheetPlanilhaDadosSistema[celulaColunaB].border = borda
        sheetPlanilhaDadosSistema[celulaColunaC].border = borda
        
#Achando a ultima linha e somando + 1
ultimaLinhaMaisUm = linha + 1
print(ultimaLinhaMaisUm)

#Desenhando a coluna C + a última linha para usar na variável
celulaUltimaLinha = 'C' + str(ultimaLinhaMaisUm)

#Desenhando a fúrmula da soma
formulaSoma = "=SUM(C2:C" + str(linha) + ")"

#Imprimindo a fórmula de soma na planilha
sheetPlanilhaDadosSistema[celulaUltimaLinha] = formulaSoma

#Salva a planilha com as alterações
planilhaDadosSistema.save(filename=caminhoArquivoDadosSistema)

#Abre a planilha
os.startfile(caminhoArquivoDadosSistema)