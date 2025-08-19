#Tela Login - Interface
#Cria conexão com o Excel e Loga
#Criando Tela de Transações
#Depósito parte 1
#Depósito parte 2 e Saque
import tkinter as tk #Importa a biblioteca para criar a interface gráfica
from tkinter import ttk #Importa o módulo ttk do Tkinter para obter estilos ou intens mais avançados
import tkinter.font as font #Para personalizar as fontes
from tkinter import messagebox #Exibe caixas de diálogo
import pandas as pd #Para trabalhar com os dados em formato tabular

from datetime import datetime
import os

from openpyxl import load_workbook

def abrir_janela_operacoes(nome, conta, cpf):
    
    janela_operacoes = tk.Tk() #Cria a instancia do Tk para criar a janela
    janela_operacoes.title("Operações") #Define o título da janela
    janela_operacoes.geometry("300x500") #Define as dimensões da janela

    #Define a cor de fundo da janela
    janela_operacoes.configure(bg="#FFFFFF") #Cor branco)
    
    #Cria o rótulo
    label_nome = tk.Label(janela_operacoes, 
                         text="Nome: " + nome,
                         font=custom_font, 
                         bg="#FFFFFF")
    label_nome.pack(pady=10) #Adicional o rótulo à janela com um espaçamento vertical de 10

    
    #Cria o rótulo
    label_conta = tk.Label(janela_operacoes, 
                         text="Número da Conta: " + conta,
                         font=custom_font, 
                         bg="#FFFFFF")
    label_conta.pack(pady=10) #Adicional o rótulo à janela com um espaçamento vertical de 10

    
    saldo = 0.0
    
    saldo_label = tk.Label(janela_operacoes, 
                         text="Saldo: R$ {:.2f}".format(saldo),
                         font=custom_font, 
                         bg="#FFFFFF")
    saldo_label.pack(pady=10) #Adicional o rótulo à janela com um espaçamento vertical de 10

    
    textbox = tk.Entry(janela_operacoes,
                         font="Arial 30", 
                         bg="#FFFFFF")
    textbox.pack(pady=10) #Adicional o Entry à janela com um espaçamento vertical de 10

    
    def sair():
        
        #Fecho a tela de Operações
        janela_operacoes.destroy()
        
    def calcular_saldo(conta):
        
        arquivo_excel = r"C:\python_projetos\python_rpa_projetos\tk_caixaEletronico\Base_Dados.xlsx"
        
        #Verifica se o arquivo Excel existe
        if os.path.exists(arquivo_excel):
            
            try:
                
                #Lê o arquivo Excel e carrega a planilha "Transações" em um DataFrame
                df_transacoes = pd.read_excel(arquivo_excel, sheet_name="Transações")
                
                #Calcula o saldo da conta somando os valores das transações da conta fornecida
                saldo = df_transacoes.loc[df_transacoes['Conta'] == int(conta), 'Valor'].sum()
                   
                
            except Exception as e:
                
                messagebox.showerror("Erro", "Erro ao ler a planilha!")
                
        else:
            
            print("Arquivo não encontrado!")
            
        
        return saldo
        
    #Calcula o saldo da conta
    saldo = calcular_saldo(conta)
    
    #Imprimo o saldo atual que o cliente tem na conta
    saldo_label.config(text="Saldo: R$ {:.2f}".format(saldo))
            
    def depositar():
        
        nonlocal saldo #Permite o acesso a variável da função externa
        
        #strip - remove os espaços branco
        valor = float(textbox.get().strip())
        
        #Atualiza o saldo adicionando o valor do depósito
        saldo += valor
        
        #Limpa o conteúdo
        textbox.delete(0, "end")
        
        saldo_label.config(text="Saldo: R$ {:.2f}".format(saldo))
        
        #Obtem a data atual
        data_atual = datetime.now().strftime("%d/%m/%Y")
        
        #nome, conta, cpf
        dados_transacao = {
            'Conta' : int(conta),
            'Nome' : str(nome),
            'CPF' : cpf,
            'Tipo' : "Depósito",
            'Valor' : valor,
            'Data' : data_atual
        }
        
        arquivo_excel = r"C:\python_projetos\python_rpa_projetos\tk_caixaEletronico\Base_Dados.xlsx"
        
        #Verifica se o arquivo Excel existe
        if os.path.exists(arquivo_excel):
            
            #Carrega o arquivo de excel existente
            workbook = load_workbook(arquivo_excel)
            
            if 'Transações' in workbook.sheetnames:
                
                #Seleciona a sheet de Transações
                sheet_transacoes = workbook['Transações']
                
            else:
                
                sheet_transacoes = workbook.create_sheet('Transações')
                
        else:
            
            print("Arquivo não encontrado!")
            
        #Adiciona a nova linha a sheet transações
        nova_linha = [dados_transacao['Conta'], dados_transacao['Nome'], dados_transacao['CPF'],
                      dados_transacao['Tipo'], dados_transacao['Valor'], dados_transacao['Data']]
        sheet_transacoes.append(nova_linha)
        
        #Salva o arquivo do excel com as alterações
        workbook.save(arquivo_excel)
        
        #Exibe a mensagem
        messagebox.showinfo("Déposito", "Depósito realizado com sucesso!")
        
    def sacar():
        
        nonlocal saldo #Permite o acesso a variável da função externa
        
        #strip - remove os espaços branco
        valor = float(textbox.get().strip())
        
        #verifica se o valor do saque está dentro do limite
        #e se há saldo suficiente na conta
        if valor <= 2000 and valor <= saldo:
            
            #Atualiza o saldo adicionando o valor do depósito
            saldo -= valor

            #Limpa o conteúdo
            textbox.delete(0, "end")

            saldo_label.config(text="Saldo: R$ {:.2f}".format(saldo))

            #Obtem a data atual
            data_atual = datetime.now().strftime("%d/%m/%Y")

            #nome, conta, cpf
            dados_transacao = {
                'Conta' : int(conta),
                'Nome' : str(nome),
                'CPF' : cpf,
                'Tipo' : "Saque",
                'Valor' : valor,
                'Data' : data_atual
            }

            arquivo_excel = r"C:\python_projetos\python_rpa_projetos\tk_caixaEletronico\Base_Dados.xlsx"

            #Verifica se o arquivo Excel existe
            if os.path.exists(arquivo_excel):

                #Carrega o arquivo de excel existente
                workbook = load_workbook(arquivo_excel)

                if 'Transações' in workbook.sheetnames:

                    #Seleciona a sheet de Transações
                    sheet_transacoes = workbook['Transações']

                else:

                    sheet_transacoes = workbook.create_sheet('Transações')

            else:

                print("Arquivo não encontrado!")

            #Adiciona a nova linha a sheet transações
            nova_linha = [dados_transacao['Conta'], dados_transacao['Nome'], dados_transacao['CPF'],
                          dados_transacao['Tipo'], dados_transacao['Valor'], dados_transacao['Data']]
            sheet_transacoes.append(nova_linha)

            #Salva o arquivo do excel com as alterações
            workbook.save(arquivo_excel)

            #Exibe a mensagem
            messagebox.showinfo("Saque", "Saque realizado com sucesso!")
            
        elif valor > 2000:
            
            messagebox.showinfo("Limite Excedido", "O valor máximo de saque é R$ 2.000")
            
        else:
            
            messagebox.showinfo("Saldo Insuficiente", "Saldo Insuficiente para realizar o saque")
            
    
    #Adicionando os botões na tela
    depositar_button = tk.Button(janela_operacoes, 
                         text="Depositar",
                         font="Arial 20", 
                         bg="#FFFFFF", 
                         command=depositar)
    depositar_button.pack(pady=10) #Adicional o botão à janela com um espaçamento vertical de 10

    
    sacar_button = tk.Button(janela_operacoes, 
                         text="Sacar",
                         font="Arial 20", 
                         bg="#FFFFFF", 
                         command=sacar)
    sacar_button.pack(pady=10) #Adicional o botão à janela com um espaçamento vertical de 10

    
    transacoes_button = tk.Button(janela_operacoes, 
                         text="Transações",
                         font="Arial 20", 
                         bg="#FFFFFF")
    transacoes_button.pack(pady=10) #Adicional o botão à janela com um espaçamento vertical de 10

    
    sair_button = tk.Button(janela_operacoes, 
                         text="Sair",
                         font="Arial 20", 
                         bg="#FFFFFF",
                         command=sair)
    sair_button.pack(pady=10) #Adicional o botão à janela com um espaçamento vertical de 10

    #Crio a janela na tela
    janela_operacoes.mainloop()
    
def Verificar_login(entry_conta, entry_senha):
    
    #Obtém os valores digitados pelo usuário no campo de entrada de dados
    conta = entry_conta.get()
    senha = entry_senha.get()
    
    #Limpa os campos de entrada de dados
    entry_conta.delete(0, tk.END)
    entry_senha.delete(0, tk.END)
    
    arquivo_excel = r"C:\python_projetos\python_rpa_projetos\tk_caixaEletronico\Base_Dados.xlsx"
    
    try:
        
        dados = pd.read_excel(arquivo_excel, sheet_name="Usuários")
        
        #Itera sobre as linhas da sheet de usuários
        for index, row in dados.iterrows():
            
            #if - se
            if str(row['Conta']) == str(conta) and str(row['Senha']) == str(senha):
                
                nome = row['Nome']
                cpf = row['CPF']
                
                #Fecho a tela de Login
                janela_login.destroy()
                
                #Abre a tela de operações
                abrir_janela_operacoes(nome, conta, cpf)
                
                break
    
    except Exception as e:
        
        messagebox.showerror("Erro", "Erro ao ler a planilha!")

janela_login = tk.Tk() #Cria a instancia do Tk para criar a janela
janela_login.title("Caixa Eletrônico") #Define o título da janela
janela_login.geometry("300x300") #Define as dimensões da janela

#Define a cor de fundo da janela
janela_login.configure(bg="#FFFFFF") #Cor branco

#Definindo uma variavel com a fonte padrão
custom_font = font.Font(family="Arial", size=16)

#Criando um campo de texto
label_conta = tk.Label(janela_login, 
                        text="Número da Conta: ",
                        font=custom_font, 
                        bg="#FFFFFF")
label_conta.pack(pady=10) #pack - Cria embaixo e centraliza 

#Campo de entrada de dados
entry_conta = tk.Entry(janela_login,
                        font=custom_font, 
                        bg="#FFFFFF")
entry_conta.pack(pady=5) #pack - Cria embaixo e centraliza

#-------------------------------

#Criando um campo de texto
label_senha = tk.Label(janela_login, 
                        text="Senha: ",
                        font=custom_font, 
                        bg="#FFFFFF")
label_senha.pack(pady=10) #pack - Cria embaixo e centraliza 

#Campo de entrada de dados
entry_senha = tk.Entry(janela_login,
                        show="*",
                        font=custom_font, 
                        bg="#FFFFFF")
entry_senha.pack(pady=5) #pack - Cria embaixo e centraliza

botao_entrar = tk.Button(janela_login, 
                        text="Entrar",
                        font=custom_font, 
                        bg="#FFFFFF",
                        command=lambda: Verificar_login(entry_conta, entry_senha)) #lambda - Cria uma função anônima que chama a Verificar_login
botao_entrar.pack(pady=10) #pack - Cria embaixo e centraliza


#Inicializa a janela na tela
janela_login.mainloop()