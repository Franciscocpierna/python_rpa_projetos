#Aula 3
# Importar as bibliotecas necessárias
import tkinter as tk  # Importa a biblioteca tkinter e a renomeia como tk
from tkinter import messagebox  # Importa a classe messagebox do módulo tkinter
from tkinter import scrolledtext  # Importa a classe scrolledtext do módulo tkinter
import os  # Importa o módulo os
import win32com.client as win32  # Importa o módulo win32com.client e o renomeia como win32
import openpyxl  # Importa o módulo openpyxl

#Definir a lista de perguntas
perguntas = [
    {
        "pergunta": "Qual é a forma correta de criar uma variável em Python?",
        "opcoes": ["a) var = 10", "b) 10 = var", "c) VAR = 10", "d) $var = 10"],
        "resposta": "a",
        "explicacao": "Em Python, a forma correta de criar uma variável é utilizando o nome da variável seguido pelo operador de atribuição (=) e o valor que será atribuído à variável."
    },
    {
        "pergunta": "Qual é o resultado da expressão 3 + 4 * 2?",
        "opcoes": ["a) 14", "b) 11", "c) 10", "d) 7"],
        "resposta": "b",
        "explicacao": "Em Python, a ordem de precedência dos operadores é levada em consideração. Neste caso, a multiplicação tem precedência sobre a adição, então a expressão é avaliada como 3 + (4 * 2) = 11."
    },
    {
        "pergunta": "Qual é o método utilizado para obter o tamanho de uma lista em Python?",
        "opcoes": ["a) len()", "b) size()", "c) count()", "d) get_size()"],
        "resposta": "a",
        "explicacao": "O método len() é utilizado para obter o tamanho de uma lista em Python. Ele retorna o número de elementos presentes na lista."
    },
    {
        "pergunta": "Qual é o operador de atribuição em Python?",
        "opcoes": ["a) =", "b) ==", "c) +=", "d) -="],
        "resposta": "a",
        "explicacao": "O operador de atribuição em Python é o sinal de igual (=). Ele é utilizado para atribuir um valor a uma variável."
    },
    {
        "pergunta": "Qual é o resultado da expressão 'hello' + 'world'?",
        "opcoes": ["a) hello", "b) world", "c) helloworld", "d) hello world"],
        "resposta": "c",
        "explicacao": "Quando se utiliza o operador de concatenação (+) com duas strings em Python, elas são concatenadas, ou seja, unidas. Neste caso, 'hello' + 'world' resulta em 'helloworld'."
    },
    {
        "pergunta": "Qual é a estrutura de repetição utilizada em Python para percorrer uma sequência?",
        "opcoes": ["a) while", "b) for", "c) repeat", "d) loop"],
        "resposta": "b",
        "explicacao": "A estrutura de repetição for é utilizada em Python para percorrer uma sequência, como uma lista, uma string, entre outros. Ela permite executar um bloco de código para cada elemento da sequência."
    },
    {
        "pergunta": "Qual é o método utilizado para imprimir um valor no console em Python?",
        "opcoes": ["a) print()", "b) input()", "c) display()", "d) show()"],
        "resposta": "a",
        "explicacao": "O método print() é utilizado para imprimir um valor no console em Python. Ele exibe o valor especificado como argumento na saída padrão."
    },
    {
        "pergunta": "Qual é o tipo de dado utilizado para armazenar uma sequência de caracteres em Python?",
        "opcoes": ["a) string", "b) number", "c) boolean", "d) list"],
        "resposta": "a",
        "explicacao": "O tipo de dado utilizado para armazenar uma sequência de caracteres em Python é chamado de string. Strings são representadas entre aspas simples ('') ou aspas duplas (\")."
    },
    {
        "pergunta": "Qual é o resultado da expressão 2 ** 3?",
        "opcoes": ["a) 6", "b) 8", "c) 2", "d) 23"],
        "resposta": "b",
        "explicacao": "O operador ** é utilizado em Python para realizar a exponenciação. Neste caso, 2 ** 3 é igual a 8, pois representa 2 elevado à potência de 3."
    },
    {
        "pergunta": "Qual é o operador utilizado para verificar se dois valores são iguais em Python?",
        "opcoes": ["a) =", "b) ==", "c) +=", "d) -="],
        "resposta": "b",
        "explicacao": "O operador de comparação utilizado para verificar se dois valores são iguais em Python é o operador ==. Ele retorna True se os valores forem iguais e False caso contrário."
    }
]

def iniciar_teste():
    
    # Permite acessar as variáveis nome_completo e email definidas fora da função
    global nome_completo, email
    
    # Obtém o nome completo e o email inseridos pelos usuários a partir dos campos de entrada na interface gráfica
    nome_completo = nome_entry.get()
    email = email_entry.get()
    
    #if - se
    if nome_completo.strip() == "" or email.strip() == "":
        
        # Mostra um aviso caso algum dos campos esteja vazio
        messagebox.showwarning("Aviso", "Por favor, preencha todos os campos antes de começar o teste.")

    else:
        
        # Se todos os campos estiverem preenchidos, destrói a tela inicial e exibe a primeira pergunta
        #Fecho o Frame da janela principal
        #Assim eu fecho o frame subindo as pergunta e opções
        tela_inicial.destroy()
        
        mostrar_pergunta()
        
def mostrar_pergunta():
    
    """
    pergunta_label: É a variável que representa o rótulo (label) onde a pergunta 
    será exibida. Presumivelmente, essa variável está vinculada a um rótulo na 
    interface gráfica.

    .config(): É um método para configurar as propriedades de um widget 
    da interface gráfica. Permite alterar várias opções, como texto, fonte, 
    cor etc., do widget especificado.

    text=perguntas[pergunta_atual]["pergunta"]: Estamos usando a opção text 
    do método config() para definir o texto que será exibido no rótulo (pergunta_label). 
    O texto é obtido a partir da lista de perguntas (perguntas), na posição 
    pergunta_atual, e acessando a chave "pergunta" do dicionário correspondente
    à pergunta atual.
    """
    pergunta_label.config(text=perguntas[pergunta_atual]["pergunta"])
    
    for i in range(len(opcao_buttons)):
        
        #Define o texto das opções de respostas nos botões
        """
        opcao_buttons[i]: É um elemento da lista opcao_buttons que representa 
        um botão de opção na interface gráfica. Presumivelmente, essa lista contém 
        vários botões de opção, e estamos acessando o elemento na posição i.

        .config(): É um método para configurar as propriedades de um widget da 
        interface gráfica. Permite alterar várias opções, como texto, fonte, cor 
        etc., do widget especificado.

        text=perguntas[pergunta_atual]["opcoes"][i]: Estamos usando a opção text 
        do método config() para definir o texto que será exibido no botão de opção 
        (opcao_buttons[i]). O texto é obtido a partir da lista de perguntas (perguntas), 
        na posição pergunta_atual, e acessando a chave "opcoes" do dicionário correspondente 
        à pergunta atual. Em seguida, acessamos o elemento na posição i da lista de opções 
        de resposta para essa pergunta.
        """
        opcao_buttons[i].config(text=perguntas[pergunta_atual]["opcoes"][i])
    
#Função para verficar a resposta dada pelo usuário
def verificar_resposta(resposta):
    
    # Permite acessar as variáveis pontuacao e pergunta_atual definidas fora da função
    global pontuacao, pergunta_atual 
    # Verifica se a resposta dada pelo usuário está correta
    
    if resposta == perguntas[pergunta_atual]["resposta"]:
        
        # Se a resposta estiver correta, exibe uma caixa de diálogo de informação com a mensagem "Resposta correta!"
        messagebox.showinfo("Resposta", "Resposta correta!")
        
        pontuacao += 1  # Incrementa a pontuação do usuário
        
    else:
        
        # Se a resposta estiver incorreta, obtém a resposta correta e a opção correta correspondente
        resposta_correta = perguntas[pergunta_atual]["resposta"]
        
        """
        perguntas[pergunta_atual]: A variável perguntas é uma lista que armazena as 
        informações de todas as perguntas do teste. A indexação pergunta_atual é usada 
        para acessar a pergunta atual que está sendo processada.

        ['opcoes']: Após acessar a pergunta atual, estamos usando a chave 'opcoes' 
        para obter a lista de opções de resposta disponíveis para essa pergunta. 
        Supomos que cada pergunta possua um dicionário com informações, e 'opcoes' é a 
        chave que guarda a lista de opções.

        ord(resposta_correta) - ord('a'): A variável resposta_correta contém a letra 
        correspondente à resposta correta para a pergunta atual 
        (por exemplo, 'a', 'b', 'c', 'd'). A função ord() retorna o valor inteiro que 
        representa o código Unicode de um caractere. Aqui, estamos convertendo a letra em 
        um valor numérico. A subtração de ord('a') é usada para obter um índice numérico 
        baseado em zero, ou seja, 'a' se torna 0, 'b' se torna 1, 'c' se torna 2 e assim por 
        diante.

        perguntas[pergunta_atual]['opcoes'][ord(resposta_correta) - ord('a')]: Com base 
        no índice calculado, estamos acessando a opção correta de resposta para a pergunta 
        atual. Isso significa que estamos selecionando a opção correta da lista de opções 
        disponíveis para essa pergunta.
        """
        opcao_correta = perguntas[pergunta_atual]["opcoes"][ord(resposta_correta) - ord('a')]
    
        # Exibe uma caixa de diálogo de informação com a mensagem "Resposta incorreta! A resposta correta era: {opcao_correta}"
        messagebox.showinfo("Resposta", f"Resposta incorreta! A resposta correta era: {resposta_correta}")
        
    respostas_usuarios.append(resposta) # Adiciona a resposta dada pelo usuário à lista de respostas_usuarios
    pergunta_atual += 1  # Atualiza a pergunta_atual para a próxima pergunta
    
    #Verifica se ainda há perguntas a serem exibidas
    if pergunta_atual < len(perguntas):
        
        mostrar_pergunta() #Se houver mais perguntas, exibe na tela a próxima pergunta
    
    else:
        
        exibe_tela_final()
        pergunta_atual=0
        respostas_usuarios.clear()
        
def exibe_tela_final():
    
    #Chama a função para salvar os dados no arquivo de Excel
    salvar_dados_excel()
    
    janela_final = tk.Toplevel(janela_principal)  # Cria uma nova janela para exibir a tela final
    janela_final.title("Resultado Final")  # Define o título da janela
    janela_final.attributes('-fullscreen', True)  # Define a janela em modo de tela cheia

    barra_rolagem = tk.Scrollbar(janela_final)  # Cria uma barra de rolagem vertical
    barra_rolagem.pack(side=tk.RIGHT, fill=tk.Y)  # Posiciona a barra de rolagem no lado direito e preenche o eixo Y

    """
    janela_final: É a janela na qual a área de texto será exibida. Presumivelmente, 
    essa variável representa a janela final da interface gráfica.

    yscrollcommand=barra_rolagem.set: O parâmetro yscrollcommand está sendo 
    configurado para utilizar uma barra de rolagem vertical (barra_rolagem) com a área de 
    texto. Isso permite que o usuário role o conteúdo verticalmente dentro da área de 
    texto quando o texto excede a área visível.

    width=80: Define a largura da área de texto como 80 caracteres. Isso determina 
    quantos caracteres podem ser exibidos em uma linha antes que uma quebra de linha 
    automática seja aplicada.

    height=15: Define a altura da área de texto como 15 linhas. Isso determina a 
    quantidade de linhas visíveis na área de texto antes que a rolagem vertical seja ativada.

    font=("Arial", 20): Define a fonte do texto exibido na área de texto como Arial 
    com tamanho 20. Isso especifica o tipo de fonte e seu tamanho.
    """
    area_texto = scrolledtext.ScrolledText(janela_final, 
                                           yscrollcommand=barra_rolagem.set, 
                                           width=80, height=15, 
                                           font=("Arial", 20))
    area_texto.pack(pady=10)
    
    #Cria uma variável para armazenar o total de acertos
    total_acertos = 0
    
    #Loop para percorrer as perguntas e exibir as informações na área de texto
    for i in range(len(perguntas)):
        
        #pergunta corrente
        pergunta = perguntas[i]
        
        #Pego a resposta correta e armazeno em resposta_correta
        resposta_correta = pergunta["resposta"]
        
        resposta_usuario = "Não respondida"
        
        #Verifico se a pergunta doi respondida pelo usuário
        if i < len(respostas_usuarios):
            
            #Armarmazeno a resposta de respostas_usuarios na varivael
            resposta_usuario = respostas_usuarios[i]
            
        #armazena em explicacao o que tem dentro da lista de pergunta
        #No indice de explicacao
        explicacao = pergunta["explicacao"]
        
        # Insere as informações da pergunta na área de texto
        area_texto.insert(tk.END, f"Pergunta {i + 1}:\n")
        area_texto.insert(tk.END, pergunta["pergunta"] + "\n")
        
        """
        area_texto: É a área de texto na qual o texto será inserido. Presumivelmente, 
        essa variável representa a área de texto na janela final da interface gráfica.

        insert(tk.END, ...): É um método da área de texto que permite inserir texto 
        em uma posição específica. Nesse caso, tk.END indica que o texto será inserido no 
        final da área de texto.

        f"Resposta correta: {perguntas[i]['opcoes'][ord(resposta_correta) - ord('a')]} \n": 
        É a string que será inserida na área de texto. 
        
        Vamos analisar cada parte dela:

        "Resposta correta: ": É uma parte fixa do texto que será exibido. Indica que a 
        linha seguinte fornecerá a resposta correta.

        {perguntas[i]['opcoes'][ord(resposta_correta) - ord('a')}: Essa expressão recupera a 
        opção correta da pergunta atual. 
        
        Explicando detalhadamente:
        
            perguntas[i]: Acessa a pergunta atual, identificada pelo índice i.
            ['opcoes']: Acessa a lista de opções da pergunta.
            [ord(resposta_correta) - ord('a')]: Obtém o índice da opção correta usando o 
            valor da variável resposta_correta. A variável resposta_correta contém uma 
            letra representando a resposta correta, e ord(resposta_correta) - ord('a') calcula 
            o deslocamento necessário para obter o índice correto na lista de opções. Isso 
            assume que as opções estão em ordem alfabética, onde 'a' representa a primeira 
            opção, 'b' a segunda e assim por diante.

        \n: Insere uma quebra de linha após o texto da resposta correta, para formatar adequadamente o texto na área de texto.
        """
        area_texto.insert(tk.END, f"Resposta correta: {perguntas[i]['opcoes'][ord(resposta_correta) - ord('a')]} \n")
        
        """
        area_texto: É a área de texto na qual o texto será inserido. Presumivelmente, 
        essa variável representa a área de texto na janela final da interface gráfica.

        insert(tk.END, ...): É um método da área de texto que permite inserir texto em 
        uma posição específica. Nesse caso, tk.END indica que o texto será inserido no 
        final da área de texto.

        f"Sua resposta: {perguntas[i]['opcoes'][ord(resposta_usuario) - ord('a')]} \n": 
        É a string que será inserida na área de texto. 
        
        Vamos analisar cada parte dela:

        "Sua resposta: ": É uma parte fixa do texto que será exibido. Indica que a 
        linha seguinte fornecerá a resposta do usuário.

        {perguntas[i]['opcoes'][ord(resposta_usuario) - ord('a')}: Essa expressão recupera a 
        opção selecionada pelo usuário para a pergunta atual. 
        
        Explicando detalhadamente:
        
            perguntas[i]: Acessa a pergunta atual, identificada pelo índice i.
            ['opcoes']: Acessa a lista de opções da pergunta.
            [ord(resposta_usuario) - ord('a')]: Obtém o índice da opção selecionada pelo 
            usuário usando o valor da variável resposta_usuario. A variável resposta_usuario 
            contém uma letra representando a opção escolhida pelo usuário, 
            e ord(resposta_usuario) - ord('a') calcula o deslocamento necessário para obter o 
            índice correto na lista de opções. Isso assume que as opções estão em ordem 
            alfabética, onde 'a' representa a primeira opção, 'b' a segunda e 
            assim por diante.

        \n: Insere uma quebra de linha após o texto da resposta do usuário, para formatar 
        adequadamente o texto na área de texto.
        """
        area_texto.insert(tk.END, f"Sua resposta: {perguntas[i]['opcoes'][ord(resposta_usuario) - ord('a')]} \n")
        area_texto.insert(tk.END, explicacao + "\n\n")

                          
        if resposta_usuario == resposta_correta:
        
            total_acertos += 1
            
    resultado = "Aprovado(a)" if total_acertos >= 6 else "Reprovado(a)"  # Verifica se o resultado é aprovado ou reprovado
    area_texto.insert(tk.END, f"Total de acertos: {total_acertos} de {len(perguntas)}\n")
    area_texto.insert(tk.END, f"Resultado: {resultado}\n")  # Insere o total de acertos e o resultado na área de texto

    barra_rolagem.config(command=area_texto.yview)  # Configura a barra de rolagem para funcionar com a área de texto
    
    #Exibe o botão para fechar a tela
    botao_fechar = tk.Button(janela_final,
                            text="Fechar",
                            font=("Arial", 20),
                            command = janela_final.destroy)
    botao_fechar.pack(pady=10)
    
    #Botão para criar o email do resultado final
    botao_email = tk.Button(janela_final,
                            text="Enviar Resultado",
                            font=("Arial", 20),
                            command = criar_email)
    botao_email.pack(pady=10)
            
    
    
def salvar_dados_excel():
    
    #Carrega o arquivo de excel
    wb = openpyxl.load_workbook(r'C:\python_projetos\python_rpa_projetos\tk_sistema_simulaTesteEmprego\Dados.xlsx')
    planilha_respostas = wb['Respostas'] #Acessa a planilha 'Respostas'
    
    #Determina a proxima linha disponivel para inserir os dados
    linha = planilha_respostas.max_row + 1
    
    planilha_respostas.cell(row=linha, column=1).value = nome_completo
    planilha_respostas.cell(row=linha, column=2).value = email
    
    #Insere as respostas dadas pelo usuário nas colunas a partir da coluna 3
    for i, resposta in enumerate(respostas_usuarios):
        
        #Adiciona as respostas dadas pelo usuários nas colunas 
        planilha_respostas.cell(row=linha, column=3+i).value = resposta
        
    pontuacao_total = (pontuacao / len(perguntas)) * 100
        
    planilha_respostas.cell(row=linha, column=len(perguntas)+3).value = pontuacao_total
    planilha_respostas.cell(row=linha, column=len(perguntas)+4).value = "Aprovado(a)" if pontuacao_total >= 60 else "Reprovado(a)"

    #Salva o arquivo do Excel
    wb.save(r'C:\python_projetos\python_rpa_projetos\tk_sistema_simulaTesteEmprego\Dados.xlsx')
    wb.close() #Fecha o arquivo do Excel
    
def criar_email():
    
    global nome_completo, email, pontuacao
    
    resultado_teste = ""
    
    #Loop para percorrer as perguntas e montar o resultado do teste
    for i in range(len(perguntas)):
    
        #Pego a pergunta atual da posição que estiver passando
        pergunta = perguntas[i]
        resposta_correta = pergunta["resposta"]
        resposta_usuario = "Não respostida"
        
        #Verifico se a pergunta doi respondida pelo usuário
        if i < len(respostas_usuarios):
            
            #Armarmazeno a resposta de respostas_usuarios na varivael
            resposta_usuario = respostas_usuarios[i]
        
            resultado_teste += f"Pergunta {i + 1}:\n"
            resultado_teste += pergunta["pergunta"] + "\n"
            resultado_teste += f"Resposta correta: {perguntas[i]['opcoes'][ord(resposta_correta) - ord('a')]} \n"
            resultado_teste += f"Sua resposta: {perguntas[i]['opcoes'][ord(resposta_usuario) - ord('a')]} \n"
            resultado_teste += pergunta["explicacao"] + "\n\n"
        
    pontuacao_total = (pontuacao / len(perguntas)) * 100
        
    if pontuacao_total >= 60:
            
        resultado = "Aprovado(a)"
            
    else:
            
        resultado = "Reprovado(a)"
            
    #Assunto do email
    subject = "Resultado do Teste de Python"
        
    #Corpo do email
    body = f"Nome Completo: {nome_completo}\n\n" + resultado_teste + f"Resultado Final: {resultado}"
    
    outlook = win32.Dispatch("Outlook.Application")  # Cria uma instância do Outlook
    namespace = outlook.GetNamespace("MAPI")  # Obtém o namespace do Outlook
    drafts_folder = namespace.GetDefaultFolder(16)  # Obtém a pasta de Rascunhos do Outlook

    email_item = drafts_folder.Items.Add()  # Cria um novo item de email na pasta de Rascunhos
    email_item.Subject = subject  # Define o assunto do email
    email_item.Body = body  # Define o corpo do email
    email_item.To = email  # Define o destinatário do email
    email_item.Save()  # Salva o rascunho do email
    #.Send()

    messagebox.showinfo("Email Criado", f"Um rascunho de email foi criado no Outlook com os resultados do teste.\n\nResultado: {resultado}")
    
# Inicializar variáveis
pontuacao = 0  # Variável para armazenar a pontuação do usuário
pergunta_atual = 0  # Variável para acompanhar a pergunta atual exibida
respostas_usuarios = []  # Lista para armazenar as respostas dadas pelos usuários
nome_completo = ""  # Variável para armazenar o nome completo do usuário
email = ""  # Variável para armazenar o email do usuário

janela_principal = tk.Tk()  # Cria a janela principal
janela_principal.title("Avaliação de Python")  # Define o título da janela
janela_principal.attributes('-fullscreen', True)  # Define a janela em modo de tela cheia

largura_tela = janela_principal.winfo_screenwidth()  # Obtém a largura da tela
altura_tela = janela_principal.winfo_screenheight()  # Obtém a altura da tela

tela_inicial = tk.Frame(janela_principal)  # Cria um frame para a tela inicial
tela_inicial.pack(pady=altura_tela * 0.2)  # Posiciona o frame na janela principal com um espaçamento vertical

# Cria um rótulo para o nome completo
label_nome = tk.Label(tela_inicial, 
                      text="Nome Completo:", 
                      font=("Arial", 20))  
label_nome.pack()  # Posiciona o rótulo no frame

#Campo de entrada de dados
nome_entry = tk.Entry(tela_inicial, 
                      font=("Arial", 20))
nome_entry.pack()  # Posiciona o rótulo no frame

# Cria um rótulo para o nome completo
label_email = tk.Label(tela_inicial, 
                      text="Email:", 
                      font=("Arial", 20))  
label_email.pack()  # Posiciona o rótulo no frame

#Campo de entrada de dados
email_entry = tk.Entry(tela_inicial, 
                      font=("Arial", 20))
email_entry.pack()  # Posiciona o rótulo no frame

#Cria o botão
button_iniciar = tk.Button(tela_inicial, 
                          text="Inicial Teste",
                          font=("Arial", 20),
                          command = iniciar_teste)
button_iniciar.pack(pady=altura_tela * 0.05)  # Posiciona o button_iniciar no frame

# Cria um rótulo para exibir a pergunta
pergunta_label = tk.Label(janela_principal, 
                      text="", 
                      font=("Arial", 20))  
pergunta_label.pack(pady=altura_tela * 0.05)  # Posiciona o rótulo na janela_principal

#Lista vazia para armazenar os botões de opção
opcao_buttons = []

#Cria 4 botões
for posicao in range(4):
    
    opcao_button = tk.Button(janela_principal,
                            text="",
                            width = int(largura_tela * 0.4),
                            command = lambda posicao = posicao: verificar_resposta(chr(ord('a') + posicao)), #ord('a') calcula o deslocamento necessário para obter o índice correto na lista
                            font = ("Arial", 20))
    opcao_button.pack(pady=altura_tela * 0.01)  # Posiciona o botão na janela_principal
    opcao_buttons.append(opcao_button) #Adiciona o botão à lista de opções
    
#Cria o menu na janela principal
barra_menu = tk.Menu(janela_principal)
barra_menu.add_command(label="Sair",
                      command = janela_principal.destroy) # Aciona o camando para sair do sistema
janela_principal.config(menu=barra_menu) #Configura a barra de menu na janela_principal

#Inicial o loop que carrega sistema
janela_principal.mainloop()
